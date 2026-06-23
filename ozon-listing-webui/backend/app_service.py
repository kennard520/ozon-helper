from __future__ import annotations

import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]   # tools/ozon-listing-webui/
REPO = ROOT.parents[1]
AUTH_ROOT = REPO / ".auth"
PROFILE_1688 = AUTH_ROOT / "1688_profile"
FRONTEND_DIST = ROOT / "frontend" / "dist"
_FONT_PATH = str(Path(__file__).resolve().parent / "assets" / "Montserrat-VF.ttf")  # 俄语信息图字体（随镜像）
_GEN_SIZE = "1024x1536"   # 全站统一生图尺寸(gpt-image 竖版,生成时即指定,不裁)；与 gen_image.DEFAULT_SIZE 一致

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import time  # noqa: E402
from urllib.parse import urlparse  # noqa: E402

from backend.catalog import Catalog  # noqa: E402
from backend.drafts import (  # noqa: E402
    collected_chars,
    create_draft_from_url,
    dimension_warnings,
    match_chars_to_attributes,
    missing_required_attributes,
    normalize_category_attrs,
    BRAND_ATTR_ID,
    NO_BRAND,
    split_collection_value,
    to_ozon_import_item,
    utc_now_iso,
    validate_draft,
)
from backend.ozon_client_adapter import (  # noqa: E402
    build_client,
    get_attribute_values,
    get_category_attributes,
    get_import_info,
    publish_items,
    search_attribute_values,
)
from backend.oss import OssClient  # noqa: E402
from backend.media_rehost import rehost_draft_media, needs_rehost  # noqa: E402
from backend.store import Store  # noqa: E402
from backend.settings_migrate import normalize_stores, migrate_ai  # noqa: E402
import backend.media as _media  # noqa: E402
import backend.ai_video as ai_video  # noqa: E402
import urllib.request  # noqa: E402


def _to_int(value: object) -> int:
    try:
        return int(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return 0


def _parse_dims_mm(text: object) -> tuple | None:
    """'190×190×340 mm' / '20.5x20.5x30.5 cm' → (L,W,H) 毫米(int)。取前3个数，按单位换算到 mm；不足3个→None。"""
    import re  # noqa: PLC0415
    s = str(text or "")
    nums = re.findall(r"\d+(?:\.\d+)?", s)
    if len(nums) < 3:
        return None
    vals = [float(x) for x in nums[:3]]
    low = s.lower()
    if ("cm" in low) or ("см" in low) or ("厘米" in low):
        vals = [v * 10 for v in vals]   # cm→mm
    return tuple(int(round(v)) for v in vals)


def _parse_volume_ml(text: object) -> int | None:
    """'4 L' / '4000 ml' / '4升' → 毫升(int)。识别不出数字→None。"""
    import re  # noqa: PLC0415
    s = str(text or "")
    m = re.search(r"\d+(?:\.\d+)?", s)
    if not m:
        return None
    v = float(m.group())
    low = s.lower()
    if ("ml" in low) or ("мл" in low) or ("毫升" in low):
        return int(round(v))
    if ("l" in low) or ("л" in low) or ("升" in low):   # 升/литр → ×1000
        return int(round(v * 1000))
    return int(round(v))


# 物理量属性关键词(按 Ozon 属性名判定)：重量/尺寸/容量——这些由代码按单位确定填，不让 AI 猜
_WEIGHT_KW = ("重量", "вес", "масса")
_DIM_KW = ("尺寸", "размер", "габарит")
_VOL_KW = ("容量", "体积", "容积", "объ", "вмести", "вмещ")

OZON_ATTRIBUTE_LANGUAGES = {"DEFAULT", "EN", "RU", "TR", "ZH_HANS"}


def _attr_language(language: str | None) -> str:
    lang = str(language or "ZH_HANS").strip().upper()
    if lang in {"ZH", "ZH-HANS", "ZH_CN", "ZH-CN"}:
        return "ZH_HANS"
    return lang if lang in OZON_ATTRIBUTE_LANGUAGES else "ZH_HANS"


def _models_url(base: str) -> str:
    """OpenAI 兼容 /v1/models 端点：容忍 base 带/不带 /v1（同 chat 端点处理）。"""
    b = str(base or "").rstrip("/")
    if b.endswith("/v1"):
        b = b[:-3].rstrip("/")
    return b + "/v1/models"


def _is_country_attr(attr: dict) -> bool:
    """是否「原产国/制造国」属性（写死中国，发布时强制填，不在 UI 让用户填）。
    名按中文(原产国/制造国/生产国/产地)或俄文(страна/произв)匹配。"""
    nm = str(attr.get("name") or "")
    low = nm.lower()
    return ("原产国" in nm or "制造国" in nm or "生产国" in nm or "产地" in nm
            or "страна" in low or "произв" in low)


def _download_bytes(url: str, *, timeout: int = 60) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    return urllib.request.urlopen(req, timeout=timeout).read()


def _has_cjk(text: object) -> bool:
    """是否含中日韩表意文字（CJK 统一汉字）。用于拦未本地化的 1688 草稿。"""
    return any("一" <= ch <= "鿿" for ch in str(text or ""))


def _img_type_from_label(label: object) -> str:
    """候选图 label/slot → 图集类型(白底/主图/细节/场景/尺寸/卖点/包装/其他)，供图集分类与排序。"""
    s = str(label or "")
    for kw, t in (("白底", "白底"), ("主图", "白底"), ("细节", "细节"), ("场景", "场景"),
                  ("尺寸", "尺寸"), ("卖点", "卖点"), ("包装", "包装")):
        if kw in s:
            return t
    return "其他"


class App:
    def __init__(self) -> None:
        self.store = Store()
        self.catalog = Catalog(store=self.store, language="ZH_HANS")
        # 俄语树用来对齐采集回来的俄语类目路径，做自动匹配
        self.catalog_ru = Catalog(store=self.store, language="RU")
        self._ensure_auth_bootstrap()

    # ---------- 鉴权（多用户）----------
    def _ensure_auth_bootstrap(self) -> None:
        """首次启动：生成稳定 JWT 密钥 + 建默认管理员 admin/admin(user_id=1，承接旧数据)。"""
        settings = self.store.get_settings()
        if not settings.get("jwt_secret"):
            self.store.save_settings({"jwt_secret": os.urandom(32).hex()})
        if self.store.count_users() == 0:
            from backend.auth import hash_password  # noqa: PLC0415
            self.store.create_user("admin", hash_password("admin"), role="admin")

    def auth_secret(self) -> str:
        return str(self.store.get_settings().get("jwt_secret") or "")

    def login(self, username: str, password: str) -> dict:
        from backend.auth import verify_password, make_token  # noqa: PLC0415
        user = self.store.get_user_by_username((username or "").strip())
        if not user or not verify_password(password or "", user["password_hash"]):
            raise ValueError("用户名或密码错误")
        if user.get("status") and user["status"] != "active":
            raise ValueError("账号已停用")
        token = make_token(user["id"], self.auth_secret())
        return {"token": token, "user": self._public_user(user)}

    def user_from_token(self, token: str) -> dict | None:
        from backend.auth import decode_token  # noqa: PLC0415
        payload = decode_token(token or "", self.auth_secret())
        if not payload:
            return None
        user = self.store.get_user_by_id(int(payload.get("sub", 0)))
        return self._public_user(user) if user else None

    @staticmethod
    def _public_user(user: dict) -> dict:
        return {"id": user["id"], "username": user["username"], "role": user.get("role", "user")}

    # ---------- 用户管理（仅 admin）----------
    def admin_list_users(self) -> dict:
        out = []
        for u in self.store.list_users():
            cnt = len(self.store.get_settings(u["id"]).get("ozon_stores") or [])
            out.append({**u, "store_count": cnt})
        return {"users": out}

    def admin_create_user(self, username: str, password: str, max_stores: int = 1) -> dict:
        username = (username or "").strip()
        if len(username) < 3:
            raise ValueError("用户名至少 3 个字符")
        if len(password or "") < 6:
            raise ValueError("密码至少 6 位")
        if self.store.get_user_by_username(username):
            raise ValueError("用户名已存在")
        from backend.auth import hash_password  # noqa: PLC0415
        u = self.store.create_user(username, hash_password(password),
                                   role="user", max_stores=max(1, int(max_stores)))
        return {**self._public_user(u), "max_stores": u["max_stores"], "store_count": 0}

    def admin_update_user(self, actor: dict, user_id: int,
                          max_stores: int | None = None,
                          status: str | None = None,
                          password: str | None = None) -> dict:
        target = self.store.get_user_by_id(int(user_id))
        if not target:
            raise ValueError("用户不存在")
        if status is not None:
            status = str(status).strip()
            if status not in ("active", "disabled"):
                raise ValueError("status 只能是 active/disabled")
            if status == "disabled":
                if int(actor["id"]) == int(user_id):
                    raise ValueError("不能禁用自己")
                if target.get("role") == "admin":
                    active_admins = [u for u in self.store.list_users()
                                     if u.get("role") == "admin" and (u.get("status") or "active") == "active"]
                    if len(active_admins) <= 1:
                        raise ValueError("不能禁用最后一个管理员")
            self.store.set_status(user_id, status)
        if max_stores is not None:
            self.store.set_max_stores(user_id, max(1, int(max_stores)))
        if password is not None:
            if len(password) < 6:
                raise ValueError("密码至少 6 位")
            from backend.auth import hash_password  # noqa: PLC0415
            self.store.set_password_hash(user_id, hash_password(password))
        u = self.store.get_user_by_id(int(user_id))
        cnt = len(self.store.get_settings(u["id"]).get("ozon_stores") or [])
        return {**self._public_user(u), "max_stores": u["max_stores"],
                "status": u["status"], "store_count": cnt}

    def admin_delete_user(self, actor: dict, user_id: int) -> dict:
        target = self.store.get_user_by_id(int(user_id))
        if not target:
            raise ValueError("用户不存在")
        if int(actor["id"]) == int(user_id):
            raise ValueError("不能删除自己")
        if target.get("role") == "admin":
            raise ValueError("不能删除管理员账号")
        self.store.delete_user(int(user_id))
        return {"deleted": True, "id": int(user_id)}

    # ---------- 钱包 ----------
    def wallet_state(self) -> dict:
        """当前用户钱包：账户 + 最近流水。"""
        return {"account": self.store.get_account(), "txns": self.store.list_txns()}

    def wallet_recharge(self, amount: float, remark: str = "") -> dict:
        try:
            amount = float(amount)
        except (TypeError, ValueError):
            raise ValueError("金额必须是数字")
        if amount <= 0:
            raise ValueError("金额必须大于 0")
        account = self.store.recharge(amount, remark=remark or "充值")
        return {"account": account}

    def presign_media(self, items: list) -> dict:
        """给插件签一批媒体的预签名 OSS 上传地址（服务级共享桶，内容哈希去重）。"""
        oss = OssClient(self.store.get_settings())
        if not oss.configured():
            raise ValueError("未配置 OSS（服务级），无法签发上传地址")
        return {"results": oss.presign_items(items or [])}

    def publish_fee(self) -> float:
        """单次发布扣费，读全局设置 publish_fee，默认 0（不收费，旧行为不变）。"""
        try:
            return float(self.store.get_settings().get("publish_fee") or 0)
        except (TypeError, ValueError):
            return 0.0

    def _auto_match_category(self, scraped: dict) -> None:
        """采集时按俄语类目路径自动猜 description_category_id + type_id。
        叶子名常有歧义(如 наушники=耳机/毛皮耳罩)，故取多候选、按候选完整路径与
        面包屑(含上级 Электроника/Одежда 等)的词重合度打分，选最吻合的那个消歧。"""
        path = str(scraped.get("category_path") or "").strip()
        if not path:
            return
        try:
            client = build_client(self.store.get_settings())
        except Exception:  # noqa: BLE001
            return  # 没配 key 就跳过
        segments = [s.strip() for s in path.split("/") if s.strip()]
        # 商品自己声明的「Тип(类型)」属性值——最精准的类型信号，优先用它搜
        type_seeds = [
            str(a.get("value")).strip()
            for a in (scraped.get("attributes") or [])
            if isinstance(a, dict) and str(a.get("name") or "").strip() in ("Тип", "Тип товара")
            and a.get("value")
        ]
        # 搜索种子：先 Тип 值(精准) → 再面包屑叶子往父回溯
        seeds = type_seeds + list(reversed(segments))
        if not seeds:
            return
        # 面包屑里所有词(俄语词形容错：长词去尾)，用于给候选路径打分消歧(电子 vs 服装)
        crumb_words = set()
        for s in segments:
            for w in s.lower().split():
                if len(w) >= 3:
                    crumb_words.add(w[:-1] if len(w) >= 5 else w)

        def _score(hit: dict) -> int:
            p = str(hit.get("path") or "").lower()
            return sum(1 for w in crumb_words if w in p)

        # 按种子顺序找，第一个有候选的种子那层选最佳(路径与面包屑域最吻合者)
        for seg in seeds:
            try:
                hits = self.catalog_ru.search(client, seg, limit=30)
            except Exception:  # noqa: BLE001
                return
            cands = [h for h in hits if h.get("description_category_id") and h.get("type_id")]
            if not cands:
                continue
            best = max(cands, key=_score)   # 路径与面包屑重合最多者(消歧)；并列取第一
            scraped["category_id"] = str(best["description_category_id"])
            scraped["type_id"] = str(best["type_id"])
            return

    def search_category(self, query: str, limit: int = 500) -> dict:
        # 本地缓存在就纯离线搜，不碰网络/不要求 API key；缓存为空才首次拉取
        client = None
        if not self.catalog.has_cache():
            client = build_client(self.store.get_settings())
        return {"results": self.catalog.search(client, query, limit=limit)}

    def category_tree(self) -> dict:
        # 本地缓存优先，无缓存才需 API key 首次拉取
        client = None
        if not self.catalog.has_tree_cache():
            client = build_client(self.store.get_settings())
        return {"tree": self.catalog.tree(client)}

    def resolve_category(self, cat_id: int, type_id: int) -> dict:
        # 按 ID 反查可读类目名（给前端回显用），本地缓存优先
        client = None
        if not self.catalog.has_cache():
            client = build_client(self.store.get_settings())
        leaf = self.catalog.find_leaf(client, cat_id, type_id)
        return {"leaf": leaf}

    def state(self) -> dict:
        self._migrate_ai_platforms()   # 一次性:旧 AI 配置 → 平台列表(已迁移则空转)
        settings = self.store.get_settings()
        _ai = migrate_ai(settings)

        def _ai_public(blk: dict) -> dict:
            out = {"engine": blk["engine"], "api_base": blk["api_base"],
                   "model": blk["model"], "api_key_saved": bool(blk["api_key"])}
            if "multimodal" in blk:
                out["multimodal"] = bool(blk["multimodal"])
            return out

        def _slot_public(kind: str) -> dict:
            out = _ai_public(_ai[f"ai_{kind}"])
            raw = settings.get(f"ai_{kind}") if isinstance(settings.get(f"ai_{kind}"), dict) else {}
            out["platform"] = str(raw.get("platform") or "")   # 新结构:槽引用的平台名
            if raw.get("model"):
                out["model"] = str(raw.get("model"))
            return out

        return {
            "settings": {
                "ozon_client_id": settings.get("ozon_client_id", ""),
                "ozon_api_key_saved": bool(settings.get("ozon_api_key")),
                "rub_cny": settings.get("rub_cny", ""),
                "rub_cny_at": settings.get("rub_cny_at", ""),
                # 合同货币：草稿 price 字段按 RUB 录入，发布时换算成合同币提交（默认 CNY）
                "contract_currency": settings.get("contract_currency", "CNY"),
                # 翻译引擎（功能③）：manual / glossary / remote；key 不回传，只回传是否已存
                "translate_engine": settings.get("translate_engine", "manual"),
                "translate_api_base": settings.get("translate_api_base", ""),
                "translate_model": settings.get("translate_model", ""),
                "translate_api_key_saved": bool(settings.get("translate_api_key")),
                # AI 卡片应用模式（功能③）：true=自动应用 / false=人工确认（默认）
                "ai_auto_apply": bool(settings.get("ai_auto_apply")),
                # 采集后自动发布到 Ozon：true=采集即自动发 / false=只建草稿（默认）
                "auto_publish": bool(settings.get("auto_publish")),
                # 卡片生成聊天引擎：remote=沿用翻译引擎配置(DeepSeek) / agnes=Agnes-2.0-Flash
                "ai_chat_provider": settings.get("ai_chat_provider", "remote"),
                # agnes 时是否把商品图发给模型做图片理解（标题/属性更贴图）
                "ai_card_vision": bool(settings.get("ai_card_vision")),
                # Agnes AI（聊天/生图/生视频共用一个 key）；key 不回传，只回传是否已存
                "agnes_api_base": settings.get("agnes_api_base", ""),
                "agnes_api_key_saved": bool(settings.get("agnes_api_key")),
                "agnes_chat_model": settings.get("agnes_chat_model", ""),
                "agnes_image_model": settings.get("agnes_image_model", ""),
                "agnes_video_model": settings.get("agnes_video_model", ""),
                "ozon_stores": [
                    {"name": st["name"], "client_id": st["client_id"],
                     "is_default": st["is_default"], "api_key_saved": bool(st["api_key"])}
                    for st in normalize_stores(settings)
                ],
                "last_publish_store": settings.get("last_publish_store", ""),
                "oss_endpoint": settings.get("oss_endpoint", ""),
                "oss_bucket": settings.get("oss_bucket", ""),
                "oss_access_key_id": settings.get("oss_access_key_id", ""),
                "oss_public_base": settings.get("oss_public_base", ""),
                "oss_access_key_secret_saved": bool(settings.get("oss_access_key_secret")),
                "ai_text": _slot_public("text"),
                "ai_image": _slot_public("image"),
                "ai_video": _slot_public("video"),
                "ai_multimodal": _slot_public("multimodal"),
                # 平台列表(只地址+Key，配一次多用途复用)；key 不回传，只回 key_saved
                "ai_platforms": [
                    {"name": str(p.get("name") or ""),
                     "base": str(p.get("base") or p.get("api_base") or ""),
                     "key_saved": bool(p.get("key") or p.get("api_key"))}
                    for p in (settings.get("ai_platforms") or [])
                    if isinstance(p, dict) and str(p.get("name") or "").strip()
                ],
                "translate_mode": _ai["translate_mode"],
            },
            "paths": {
                "db": str(self.store.path),
                "auth_root": str(AUTH_ROOT),
                "profile_1688": str(PROFILE_1688),
            },
            "status": {
                "ozon_api": "connected" if settings.get("ozon_client_id") and settings.get("ozon_api_key") else "not_configured",
                "profile_1688": "present" if PROFILE_1688.exists() else "not_created",
            },
        }

    def _migrate_ai_platforms(self) -> None:
        """一次性迁移：旧 ai_text/ai_image/... 的 {engine,api_base,api_key} → ai_platforms(地址+Key)
        + 槽改成引用平台名。已有 ai_platforms 则跳过。保留各槽 model + ai_text.multimodal。"""
        s = self.store.get_settings()
        if s.get("ai_platforms"):
            return
        olds = {}
        for kind in ("text", "multimodal", "image", "video"):
            slot = s.get(f"ai_{kind}")
            if (isinstance(slot, dict) and str(slot.get("api_base") or "").strip()
                    and not str(slot.get("platform") or "").strip()):
                olds[kind] = slot
        if not olds:
            return

        def _name(base: str, idx: int) -> str:
            h = base.lower()
            if "agnes" in h:
                return "Agnes"
            if "gptplus5" in h or "gptplus" in h:
                return "GPTPlus5"
            if "openai" in h:
                return "OpenAI"
            if "deepseek" in h:
                return "DeepSeek"
            return f"平台{idx + 1}"

        platforms: list[dict] = []
        sig_to_name: dict = {}
        new_slots: dict = {}
        for kind, slot in olds.items():
            base = str(slot.get("api_base") or "").strip()
            key = str(slot.get("api_key") or "").strip()
            sig = (base, key)
            if sig not in sig_to_name:
                nm = _name(base, len(platforms))
                while any(p["name"] == nm for p in platforms):
                    nm += "2"
                sig_to_name[sig] = nm
                platforms.append({"name": nm, "base": base, "key": key})
            blk = {"platform": sig_to_name[sig], "model": str(slot.get("model") or "").strip()}
            if kind == "text":
                blk["multimodal"] = bool(slot.get("multimodal"))
            new_slots[f"ai_{kind}"] = blk
        self.store.save_settings({"ai_platforms": platforms, **new_slots})

    @staticmethod
    def _clean_secret(value: object) -> str:
        # 去掉粘贴时常带进来的反引号/引号/空白（曾导致 Invalid Api-Key）
        return str(value or "").strip().strip("`'\"").strip()

    def save_settings(self, payload: dict) -> dict:
        allowed: dict = {}
        if "ozon_client_id" in payload:
            allowed["ozon_client_id"] = self._clean_secret(payload.get("ozon_client_id"))
        if payload.get("contract_currency"):
            allowed["contract_currency"] = str(payload["contract_currency"]).strip().upper()
        if payload.get("ozon_api_key"):
            allowed["ozon_api_key"] = self._clean_secret(payload["ozon_api_key"])
        # 翻译引擎（功能③）：引擎名 / API base / model 非空才存；key 走 _clean_secret 且非空才存
        if payload.get("translate_engine"):
            allowed["translate_engine"] = str(payload["translate_engine"]).strip().lower()
        if payload.get("translate_api_base"):
            allowed["translate_api_base"] = str(payload["translate_api_base"]).strip()
        if payload.get("translate_model"):
            allowed["translate_model"] = str(payload["translate_model"]).strip()
        if payload.get("translate_api_key"):
            allowed["translate_api_key"] = self._clean_secret(payload["translate_api_key"])
        # AI 卡片自动应用开关：False 是有意义的值，用 is not None 判断（不能 falsy 判断，否则关不掉）
        if payload.get("ai_auto_apply") is not None:
            allowed["ai_auto_apply"] = bool(payload["ai_auto_apply"])
        # 采集后自动发布开关：False 是有意义的值，用 is not None 判断（不能 falsy 判断，否则关不掉）
        if payload.get("auto_publish") is not None:
            allowed["auto_publish"] = bool(payload["auto_publish"])
        # Agnes AI：provider/base/model 非空才存（空字段不覆盖已存值）；key 走 _clean_secret
        if payload.get("ai_chat_provider"):
            allowed["ai_chat_provider"] = str(payload["ai_chat_provider"]).strip().lower()
        if payload.get("ai_card_vision") is not None:
            allowed["ai_card_vision"] = bool(payload["ai_card_vision"])
        if payload.get("agnes_api_base"):
            allowed["agnes_api_base"] = str(payload["agnes_api_base"]).strip()
        if payload.get("agnes_api_key"):
            allowed["agnes_api_key"] = self._clean_secret(payload["agnes_api_key"])
        for k in ("agnes_chat_model", "agnes_image_model", "agnes_video_model"):
            if payload.get(k):
                allowed[k] = str(payload[k]).strip()
        # 汇率：随时间戳一起持久化（CLAUDE.md 硬规矩，不存无日期汇率）
        if payload.get("rub_cny") not in (None, ""):
            try:
                allowed["rub_cny"] = float(payload["rub_cny"])
                allowed["rub_cny_at"] = utc_now_iso()
            except (TypeError, ValueError):
                pass
        if payload.get("ozon_stores") is not None:
            from backend.settings_migrate import normalize_stores, mirror_of  # noqa: PLC0415
            prev = {str(st.get("client_id")): str(st.get("api_key") or "")
                    for st in (self.store.get_settings().get("ozon_stores") or [])}
            incoming = []
            for st in payload["ozon_stores"]:
                cid = self._clean_secret(st.get("client_id"))
                name = str(st.get("name") or "").strip()
                key = self._clean_secret(st.get("api_key")) or prev.get(cid, "")
                if cid and name:
                    incoming.append({"name": name, "client_id": cid, "api_key": key,
                                     "is_default": bool(st.get("is_default"))})
            stores = normalize_stores({"ozon_stores": incoming})
            # 店铺配额：非 admin 按 max_stores 限制店铺数（admin 豁免）。后端强制。
            from backend.store import current_user_id  # noqa: PLC0415
            actor = self.store.get_user_by_id(current_user_id.get())
            if actor and actor.get("role") != "admin":
                limit = int(actor.get("max_stores") or 1)
                if len(stores) > limit:
                    raise ValueError(f"最多 {limit} 个店铺，请联系管理员调整上限")
            allowed["ozon_stores"] = stores
            mcid, mkey = mirror_of(stores)
            allowed["ozon_client_id"] = mcid
            allowed["ozon_api_key"] = mkey
        if payload.get("translate_mode"):
            allowed["translate_mode"] = str(payload["translate_mode"]).strip().lower()
        # 平台列表：地址+Key 配一次，多用途复用。key 留空 = 沿用同名平台已存 key。
        if payload.get("ai_platforms") is not None:
            prev_plats = {str(p.get("name") or ""): str(p.get("key") or p.get("api_key") or "")
                          for p in (self.store.get_settings().get("ai_platforms") or [])
                          if isinstance(p, dict)}
            out_plats = []
            for p in payload["ai_platforms"]:
                if not isinstance(p, dict):
                    continue
                name = str(p.get("name") or "").strip()
                if not name:
                    continue
                key = self._clean_secret(p.get("key") or p.get("api_key")) or prev_plats.get(name, "")
                out_plats.append({"name": name,
                                  "base": str(p.get("base") or p.get("api_base") or "").strip(),
                                  "key": key})
            allowed["ai_platforms"] = out_plats
        for kind in ("ai_text", "ai_image", "ai_video", "ai_multimodal"):
            blk = payload.get(kind)
            if isinstance(blk, dict):
                prev_blk = self.store.get_settings().get(kind) or {}
                prev_key = str(prev_blk.get("api_key") or "") if isinstance(prev_blk, dict) else ""
                blk_out = {
                    "engine": str(blk.get("engine") or "").strip().lower(),
                    "api_base": str(blk.get("api_base") or "").strip(),
                    "api_key": self._clean_secret(blk.get("api_key")) or prev_key,
                    "model": str(blk.get("model") or "").strip(),
                    "platform": str(blk.get("platform") or "").strip(),   # 新:槽引用的平台名
                }
                if kind == "ai_text":
                    blk_out["multimodal"] = bool(blk.get("multimodal"))
                allowed[kind] = blk_out
        for k in ("oss_endpoint", "oss_bucket", "oss_access_key_id", "oss_public_base"):
            if payload.get(k) is not None:
                allowed[k] = str(payload[k]).strip()
        if payload.get("oss_access_key_secret"):
            allowed["oss_access_key_secret"] = self._clean_secret(payload["oss_access_key_secret"])
        if payload.get("last_publish_store") is not None:
            allowed["last_publish_store"] = str(payload["last_publish_store"]).strip()
        if allowed:
            self.store.save_settings(allowed)
        return self.state()

    def _settings_for_store(self, store_client_id: str | None) -> dict:
        """返回目标店的 settings：client_id/api_key 替换成目标店，其余全局字段（rub_cny/
        contract_currency 等）不变。空或 == 主店 → 主店；额外店在 ozon_stores 里按 client_id
        匹配；找不到抛 ValueError。"""
        s = self.store.get_settings()
        cid = str(store_client_id or "").strip()
        if not cid or cid == str(s.get("ozon_client_id") or "").strip():
            return s
        for st in s.get("ozon_stores") or []:
            if str(st.get("client_id") or "").strip() == cid:
                return {**s, "ozon_client_id": cid, "ozon_api_key": str(st.get("api_key") or "")}
        raise ValueError(f"目标店未配置: {cid}")

    def list_drafts(self, *, status: str = "all", page: int = 1, page_size: int = 20,
                    store_client_id: str | None = None) -> dict:
        """草稿绑定店：store_client_id 非 None 时只返回该店草稿（计数同）。
        没传 store_client_id → 回退用户默认店(settings.ozon_client_id)，只查这一个店；
        否则"不带店=全店混列"会和前端"带当前店"的请求结果打架(草稿忽有忽无)。
        没配默认店才退回不过滤(兼容)。"""
        if store_client_id is None:
            default_store = str((self.store.get_settings() or {}).get("ozon_client_id") or "")
            store_client_id = default_store or None
        scid = None if store_client_id is None else str(store_client_id or "")
        drafts, total = self.store.list_drafts_page(
            status=status, page=page, page_size=page_size, store_client_id=scid)
        return {
            "drafts": drafts,
            "total": total,
            "page": max(1, int(page)),
            "page_size": max(1, min(int(page_size), 200)),
            "counts": self.store.count_by_status(store_client_id=scid),
        }

    def copy_draft_to_store(self, draft_id: int, target_client_id: str) -> dict:
        """把草稿复制到另一个店：克隆内容（标题/类目/属性/媒体/尺寸/描述/采购信息），
        重置店级字段（store_client_id=目标店、状态重算、清 ozon_product_id/库存/仓库/发布响应）。
        目标店已有同来源草稿则拦截，不重复建。"""
        from backend.drafts import utc_now_iso  # noqa: PLC0415
        target = str(target_client_id or "").strip()
        if not target:
            return {"ok": False, "error": "未指定目标店铺"}
        src = self.store.get_draft(draft_id)
        if src is None:
            raise KeyError(f"draft {draft_id} not found")
        if str(src.get("store_client_id") or "") == target:
            return {"ok": False, "error": "目标店与当前店相同"}
        dup = self.store.find_by_source_url(src["source_url"], None, target)
        if dup:
            return {"ok": False, "error": "该店已存在此来源商品", "draft_id": dup["id"]}
        now = utc_now_iso()
        clone = {
            **src,
            "store_client_id": target,
            "ozon_product_id": None,
            "stock": 0,
            "warehouse_id": None,
            "publish_response": None,
            "status": "ready",          # insert_draft 会按校验结果改成 invalid（缺字段时）
            "created_at": now,
            "updated_at": now,
        }
        clone.pop("id", None)
        saved = self.store.insert_draft(clone)
        return {"ok": True, "draft": saved}

    def update_draft(self, draft_id: int, payload: dict) -> dict:
        current = self.store.get_draft(draft_id)
        if current is None:
            raise KeyError(f"draft {draft_id} not found")
        normalized = dict(payload or {})
        brand_name = str(normalized.get("brand_name") or current.get("brand_name") or "").strip()
        brand_id = normalized.get("brand_id", current.get("brand_id"))
        normalized["brand_id"] = brand_id if brand_name == NO_BRAND and _to_int(brand_id) > 0 else None
        normalized["brand_name"] = NO_BRAND
        draft = self.store.update_draft(draft_id, normalized)
        return {"draft": draft}

    def batch_publish(self, ids: list, store_client_id: str | None = None) -> dict:
        """批量发布：逐个调 publish（各自校验/扣费/媒体托管），单个失败不影响其它。
        返回 {results:[{id, published, errors}], published, failed}。"""
        results = []
        ok = 0
        for did in ids or []:
            try:
                r = self.publish(int(did), store_client_id)
                published = bool(r.get("published"))
                if published:
                    ok += 1
                results.append({"id": did, "published": published,
                                "errors": r.get("errors") or [], "warnings": r.get("warnings") or []})
            except Exception as exc:  # noqa: BLE001
                results.append({"id": did, "published": False, "errors": [str(exc)]})
        return {"results": results, "published": ok, "failed": len(results) - ok}

    def batch_update_drafts(self, ids: list, patch: dict) -> dict:
        """批量给多个草稿打同一个补丁（目前用于批量设库存/仓库）。
        只放行安全的批量字段；逐条复用 update_draft，并保留各自原 status——
        否则未传 status 时 update_draft 会重算，把已发布草稿误降级为 ready。"""
        allowed = {k: v for k, v in (patch or {}).items() if k in ("stock", "warehouse_id")}
        if not allowed:
            raise ValueError("没有可批量设置的字段（仅支持 stock / warehouse_id）")
        updated, errors = [], []
        for did in ids or []:
            try:
                cur = self.store.get_draft(int(did))
                if cur is None:
                    errors.append({"id": did, "error": "草稿不存在"})
                    continue
                # 保留原 status，避免批量改库存/仓库时被重算降级
                draft = self.store.update_draft(int(did), {**allowed, "status": cur.get("status")})
                updated.append(draft)
            except Exception as exc:  # noqa: BLE001
                errors.append({"id": did, "error": str(exc)})
        return {"updated": updated, "errors": errors}

    def translate_draft(self, draft_id: int) -> dict:
        from backend.translate import get_engine  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        settings = self.store.get_settings()
        from backend.settings_migrate import migrate_ai, ai_config  # noqa: PLC0415
        _tm = migrate_ai(settings)["translate_mode"]
        # ai translate_mode 需要 key；无 key 时降级 manual（避免抛 RuntimeError）
        if _tm == "ai" and not ai_config(settings, "text")["key"]:
            _tm = "manual"
        engine = get_engine(_tm, settings)
        title = engine.translate(str(draft.get("ozon_title") or ""))
        desc = engine.translate(str(draft.get("description") or ""))
        updated = self.store.update_draft(draft_id, {"ozon_title": title, "description": desc})
        still = _has_cjk(title) or _has_cjk(desc)
        note = "" if not still else "仍含中文：manual 引擎只占位，请配置 remote 引擎或手动翻译"
        return {"draft": updated, "engine": engine.name, "still_cjk": still, "note": note}

    def _validate_and_build_item(self, draft: dict, store_settings: dict | None = None) -> tuple[list[str], dict | None]:
        """共享校验 + item 构建逻辑（publish 与 publish_preview 共用）。

        返回 (errors, item)：
        - errors: 阻断性错误列表（同 publish 的拦截规则）；非空时 item 为 None
        - item: to_ozon_import_item 产出 + 币种换算后的 import item（pre-media-swap）

        注意：
        - 此方法不写 DB（不改 status、不写 validation_errors）——由调用方决定是否持久化
        - 媒体上传不在此方法里发生；media 检测仅验证 company_id / is_logged_in（is_ok 检查）
        - 返回的 item 仍使用 draft 里的原始 media URL（未上传替换），仅供预览；
          publish 会在校验通过后自行完成 upload + swap
        """
        errors: list[str] = list(validate_draft(draft))
        warnings: list[str] = list(dimension_warnings(draft))   # 克重/尺寸缺失=软警告，不拦发布
        # 1688 来源若标题/描述仍含中文，说明还没本地化（功能③的 AI 中译俄）——拦截
        if draft.get("source_platform") == "1688" and (
            _has_cjk(draft.get("ozon_title")) or _has_cjk(draft.get("description"))
        ):
            errors.append("1688 商品未本地化（标题/描述仍含中文），请先翻译成俄语再发布")
        # 类目必填属性：缺了只警告、不阻断——发到 Ozon 后在后台补(Ozon 卡片好编辑)
        cat, typ = str(draft.get("category_id") or "").strip(), str(draft.get("type_id") or "").strip()
        if cat and typ:
            try:
                attrs = self._category_attrs(int(cat), int(typ))
                for m in missing_required_attributes(draft, attrs):
                    warnings.append(f"缺必填属性：{m['name']}")
            except ValueError:
                pass  # 未配置 API key 等：不阻断
            except Exception as exc:  # noqa: BLE001
                errors.append(f"类目必填属性校验失败，无法确认是否缺失（请重试）：{exc}")
        # 合同货币汇率校验
        settings = store_settings if store_settings is not None else self.store.get_settings()
        currency = str(settings.get("contract_currency") or "CNY").upper()
        if not settings.get("rub_cny"):
            errors.append("未配置 RUB/CNY 汇率，无法换算合同币价格，请先在设置里填写汇率")
        if errors:
            return errors, warnings, None
        # 构建 item（用原始 URL，不做 upload/swap）
        item = to_ozon_import_item(draft)
        # 内部 price/old_price 统一为 CNY 人民币。
        if currency == "RUB":
            # 合同币种=卢布 → CNY 换算成 RUB（rub_cny = CNY/RUB）
            rate = float(settings.get("rub_cny"))
            item["currency_code"] = "RUB"
            item["price"] = str(round(float(item["price"] or 0) / rate, 2))
            if item.get("old_price"):
                item["old_price"] = str(round(float(item["old_price"]) / rate, 2))
        else:
            # 合同币种=CNY → 价格已是人民币，直接发，不换算
            item["currency_code"] = currency
        return [], warnings, item

    def publish_preview(self, draft_id: int, store_client_id: str | None = None) -> dict:
        """预览将要发布的内容，不发送任何请求，不写 DB（无副作用）。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        target_store = str(draft.get("store_client_id") or "") or store_client_id
        errors, warnings, item = self._validate_and_build_item(draft, self._settings_for_store(target_store))
        if errors:
            return {"ok": False, "errors": errors, "warnings": warnings, "summary": None}
        # 构建摘要（item 已含换算后价格）
        images = item.get("images") or []
        attributes = item.get("attributes") or []
        has_video = bool(draft.get("video_url"))
        summary = {
            "offer_id": item.get("offer_id"),
            "name": item.get("name"),
            "description_len": len(str(item.get("description_category_id") and draft.get("description") or "")),
            "category_id": item.get("description_category_id"),
            "type_id": item.get("type_id"),
            "price": item.get("price"),
            "old_price": item.get("old_price"),
            "currency_code": item.get("currency_code"),
            "dims_mm": {
                "depth": item.get("depth"),
                "width": item.get("width"),
                "height": item.get("height"),
            },
            "weight_g": item.get("weight"),
            "images_count": len(images),
            "attributes_count": len(attributes),
            "has_video": has_video,
        }
        return {"ok": True, "errors": [], "warnings": warnings, "summary": summary}

    def publish_preflight(self, draft_id: int) -> dict:
        """发布前核对清单：error=硬拦(不让发) / warn=建议 / verify=看图识别需人工核对 / passed=已就绪。
        硬拦复用 validate_draft;软项=尺寸密度警告 + 理解层标'图片识别'的数字 + 标题长度/图片数。"""
        from backend.drafts import dimension_warnings, loads_json, validate_draft  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = sr if isinstance(sr, dict) else {}
        und = sr.get("understanding") if isinstance(sr.get("understanding"), dict) else {}
        conf = und.get("confidence") if isinstance(und.get("confidence"), dict) else {}
        specs = und.get("specs") if isinstance(und.get("specs"), dict) else {}

        checks: list = []
        for e in validate_draft(draft):                       # 🔴 硬拦
            checks.append({"severity": "error", "label": e})
        for w in dimension_warnings(draft):                    # 🟡 尺寸/密度建议
            checks.append({"severity": "warn", "label": w})
        for key, c in conf.items():                            # 🟡 看图识别·待核对的数字(只列 specs 里有值的)
            if str(c) == "图片识别" and str(key).startswith("specs."):
                field = str(key).split(".", 1)[1]
                val = specs.get(field, "")
                if str(val).strip():
                    checks.append({"severity": "verify",
                                   "label": f"核对 {field}: {val}（看图识别,易错,发布前请确认）"})
        title = str(draft.get("ozon_title") or "")
        if len(title) > 150:
            checks.append({"severity": "warn", "label": f"标题偏长({len(title)} 字),建议精简到 150 内"})
        imgs = draft.get("images") or []
        if len(imgs) < 3:
            checks.append({"severity": "warn", "label": f"图片偏少({len(imgs)} 张),Ozon 建议多张主图/细节"})

        passed: list = []
        if title and not any("标题" in c["label"] for c in checks if c["severity"] == "error"):
            passed.append(f"标题就绪({len(title)} 字)")
        if str(draft.get("brand_name") or "") == "Нет бренда":
            passed.append("品牌 = 无品牌 ✓")
        try:
            if float(draft.get("price") or 0) > 0:
                passed.append(f"售价 {draft.get('price')} ✓")
        except (TypeError, ValueError):
            pass
        if (draft.get("weight_g") or 0) and (draft.get("length_mm") or 0):
            passed.append("尺寸/重量已填 ✓")

        blocking = sum(1 for c in checks if c["severity"] == "error")
        return {"ok": True, "checks": checks, "passed": passed,
                "blocking": blocking, "can_publish": blocking == 0}

    def publish(self, draft_id: int, store_client_id: str | None = None) -> dict:
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        if str(draft.get("media_status") or "done") == "pending":
            raise ValueError("图片还在上传，请稍候再发布")
        # 草稿绑定店：优先发到草稿自带店；旧草稿(无 store_client_id)回退入参/默认店
        target_store = str(draft.get("store_client_id") or "") or store_client_id
        store_settings = self._settings_for_store(target_store)
        # 写死：品牌=无品牌、原产国=中国(覆盖任何采集/AI 值；要改去 Ozon 后台改)
        draft = self._ensure_fixed_attrs(draft)
        # 媒体托管：插件已把媒体传到卖家自己的 Ozon 店铺(ir.ozone.ru)时，全是 Ozon 原生链接、
        # 无需 OSS——跳过。只有存在非 Ozon 媒体(老路径/手动加图)时才用 OSS 兜底；未配 OSS 才硬拦。
        oss = OssClient(store_settings)
        rehost_stats = {"uploaded": 0, "failed": 0}
        if needs_rehost(draft):
            if not oss.configured():
                errs = ["有非 Ozon 来源的图片/视频需托管：请用插件把媒体传到你的 Ozon 店铺，或在设置里配置阿里云 OSS"]
                updated = self.store.update_draft(draft_id, {"status": "invalid", "validation_errors": errs})
                return {"published": False, "draft": updated, "errors": errs}
            draft, rehost_stats = rehost_draft_media(draft, oss.upload_remote)
            # OSS URL 持久化回草稿（再发幂等；展示也变 OSS 图）
            self.store.update_draft(draft_id, {
                "images": draft.get("images") or [], "video_url": draft.get("video_url") or "",
                "source_raw": draft.get("source_raw") or {}})
        errors, warnings, item = self._validate_and_build_item(draft, store_settings)
        if errors:
            updated = self.store.update_draft(draft_id, {"status": "invalid", "validation_errors": errors})
            return {"published": False, "draft": updated, "errors": errors, "warnings": warnings}
        if rehost_stats.get("failed"):
            warnings = [*warnings, f"{rehost_stats['failed']} 个媒体未能上传到 OSS，已沿用原链接（Ozon 可能抓不到）"]
        # 发布扣费（publish_fee>0 才扣；余额不足直接拦下不发布）
        fee = self.publish_fee()
        if fee > 0 and not self.store.deduct(fee, biz_no=f"publish:{draft_id}", remark="发布商品"):
            errs = ["余额不足，请先充值后再发布"]
            updated = self.store.update_draft(draft_id, {"status": "invalid", "validation_errors": errs})
            return {"published": False, "draft": updated, "errors": errs, "warnings": warnings}
        try:
            response = publish_items(store_settings, [item])
        except Exception as exc:  # noqa: BLE001  Ozon 报错(如图片URL无效)别透成 500，落库+回前端
            if fee > 0:  # 调用失败把刚扣的费退回，不白扣
                self.store.refund(fee, biz_no=f"publish-refund:{draft_id}", remark="发布失败退款")
            msg = f"Ozon 拒收: {exc}"
            updated = self.store.update_draft(draft_id, {"status": "invalid", "validation_errors": [msg]})
            return {"published": False, "draft": updated, "errors": [msg], "warnings": warnings}
        self.store.save_settings({"last_publish_store": str(store_settings.get("ozon_client_id") or "")})
        task_id = ((response.get("result") or {}).get("task_id"))
        poll: dict = {}
        final_status = "draft"
        item_errors: list[str] = []
        if task_id:
            for _ in range(10):
                time.sleep(2)
                try:
                    info = get_import_info(store_settings, task_id)
                except Exception as exc:  # noqa: BLE001
                    poll = {"error": str(exc)}
                    break
                poll_items = (info.get("result") or {}).get("items") or []
                statuses = [it.get("status") for it in poll_items]
                if statuses and not any(s in ("pending", "not_started", "") for s in statuses):
                    poll = info
                    # ⚠️ Ozon 即使 status='imported' 也可能带 level='error' 的错误：卡片建了(有 product_id)
                    # 但不可售(如必填属性空)，不会出现在后台正常商品里。有 error 级错误 → 视为失败，
                    # 并把 Ozon 的具体原因带回前端，否则会误报"已发布"而后台实际没有。
                    for it in poll_items:
                        for e in (it.get("errors") or []):
                            if str(e.get("level") or "").lower() == "error":
                                aid = e.get("attribute_id")
                                desc = str(e.get("description") or e.get("code") or "未知错误")
                                item_errors.append(f"[属性{aid}] {desc}" if aid else desc)
                    # skipped = Ozon 跳过(变体单条发常见：变体须走「整组发布」；或内容无变化)。
                    # 既非成功也非失败，单独标记并提示，避免误报"已发布"或"失败"。
                    if statuses and all(s == "skipped" for s in statuses) and not item_errors:
                        final_status = "skipped"
                        item_errors.append("Ozon 跳过未更新：若是变体商品请用「整组发布」；单品则说明内容无变化")
                        break
                    ok = all(s in ("imported", "skipped") for s in statuses) and not item_errors
                    final_status = "published" if ok else "failed"
                    break
            else:
                poll = {"warning": "poll timeout — Ozon 仍在异步处理，稍后用 task_id 查"}
        updated = self.store.update_draft(draft_id, {
            "status": final_status,
            "validation_errors": item_errors or None,
            "publish_response": {"import": response, "poll": poll, "warnings": warnings,
                                 "store_client_id": str(store_settings.get("ozon_client_id") or ""),
                                 "rehost": rehost_stats},
            "offer_id": item["offer_id"],
        })
        if fee > 0 and final_status == "failed":   # Ozon 校验没过、卡片不可售 → 退发布费
            self.store.refund(fee, biz_no=f"publish-refund:{draft_id}", remark="发布未成功(Ozon 校验)退款")
        return {"published": final_status == "published", "draft": updated, "response": response,
                "poll": poll, "task_id": task_id, "warnings": warnings, "rehost": rehost_stats,
                "errors": item_errors}

    def pull_ozon_products(self, visibility: str = "ALL") -> dict:
        from backend.ozon_client_adapter import (  # noqa: PLC0415
            list_ozon_products, get_ozon_info, get_ozon_attributes, get_ozon_descriptions,
            ozon_to_draft)
        errors: list = []
        try:
            listing = list_ozon_products(self.store.get_settings(), visibility)
        except Exception as exc:  # noqa: BLE001
            return {"pulled": 0, "drafts": self.store.list_drafts(), "errors": [str(exc)]}
        offer_ids = [str(it.get("offer_id")) for it in listing if it.get("offer_id")]
        info = get_ozon_info(self.store.get_settings(), offer_ids) if offer_ids else {}
        try:
            attrs = get_ozon_attributes(self.store.get_settings(), offer_ids) if offer_ids else {}
        except Exception as exc:  # noqa: BLE001
            attrs = {}
            errors.append(f"属性拉取失败(尺寸/属性留空): {exc}")
        try:
            descriptions = get_ozon_descriptions(self.store.get_settings(), offer_ids) if offer_ids else {}
        except Exception as exc:  # noqa: BLE001
            descriptions = {}
            errors.append(f"描述拉取失败(描述留空): {exc}")
        pulled = 0
        for oid in offer_ids:
            if oid not in info:
                continue
            d = ozon_to_draft(info[oid], attrs.get(oid))
            # 补全描述：/v3/product/info/list 不含描述，单独接口拉取
            desc_text = descriptions.get(oid, "")
            if desc_text:
                d["description"] = desc_text
            existing = self.store.find_by_offer_id(oid)
            if existing:
                # 再拉已存在草稿：非破坏式合并，只补空字段 + 刷新 Ozon 权威身份，
                # 绝不覆盖用户手编的 description/price/supplier 等（否则每次拉单都丢编辑）。
                self.store.update_draft(existing["id"], self._merge_pulled_into_existing(existing, d))
            else:
                self.store.insert_draft(self._normalize_ozon_draft(d, oid))
            pulled += 1
        return {"pulled": pulled, "drafts": self.store.list_drafts(), "errors": errors}

    @staticmethod
    def _merge_pulled_into_existing(existing: dict, pulled: dict) -> dict:
        """再拉已存在草稿时，算出"只补空、不覆盖"的最小更新集（纯函数，可测）。

        - 用户可能手编的字段：仅当现有草稿该字段为空/缺失时才用拉来的值填补；
          已有非空值则一律保留，不传进更新集（让 store 维持原值）。
        - Ozon 权威身份/元数据：用户不会手改 → 总是用拉来的最新值刷新。
        - status 显式设为 "published"：从 Ozon 拉回的商品已上架，
          update_draft 的 status 重算逻辑若不传 status 会将其降级为 "ready"，
          故此处明确传入 "published" 保持已上架状态。
        """
        def _empty(v: object) -> bool:
            if v is None:
                return True
            if isinstance(v, str):
                return v.strip() == ""
            if isinstance(v, (list, tuple, dict)):
                return len(v) == 0
            return False

        patch: dict = {}
        # 用户可能手编的字段：只在现有为空时补
        user_editable = (
            "description", "ozon_title", "price", "old_price", "stock",
            "supplier", "purchase_url", "purchase_note", "cost_cny",
            "category_id", "type_id",
        )
        for k in user_editable:
            if k in pulled and _empty(existing.get(k)) and not _empty(pulled.get(k)):
                patch[k] = pulled[k]

        # 图片：仅当现有草稿一张都没有时才用拉来的
        if _empty(existing.get("images")) and not _empty(pulled.get("images")):
            patch["images"] = pulled["images"]

        # 视频：仅当现有草稿没有视频时才用拉来的（Ozon 重新托管的 v.ozone.ru URL）
        if _empty(existing.get("video_url")) and not _empty(pulled.get("video_url")):
            patch["video_url"] = pulled["video_url"]

        # Ozon 权威身份/元数据：总是刷新（用户不会手编）
        if pulled.get("ozon_product_id") is not None:
            patch["ozon_product_id"] = pulled["ozon_product_id"]
        patch["source"] = pulled.get("source", "ozon")
        patch["offer_id"] = existing.get("offer_id") or pulled.get("offer_id")
        # 从 Ozon 拉回的商品已在平台上线，显式保持 published 状态（不让 update_draft 重算降级）
        patch["status"] = "published"
        return patch

    @staticmethod
    def _normalize_ozon_draft(draft: dict, offer_id: str) -> dict:
        """补齐 store.insert_draft 要求但纯映射不产出的键。
        source_url 在 drafts 表是 UNIQUE 主键约束（且 insert 按它去重），
        Ozon 商品无 1688 链接，故用合成唯一键 ozon://product/<offer_id>。"""
        now = utc_now_iso()
        draft.setdefault("purchase_url", "")
        draft.setdefault("purchase_note", "")
        draft["brand_id"] = None
        draft["brand_name"] = NO_BRAND
        draft.setdefault("cost_cny", None)
        draft.setdefault("video_url", "")
        draft.setdefault("local_images", [])
        draft.setdefault("validation_errors", [])
        draft.setdefault("publish_response", None)
        draft["source_url"] = f"ozon://product/{offer_id}"
        draft["created_at"] = now
        draft["updated_at"] = now
        return draft

    def category_attributes(self, cat_id: int, type_id: int, language: str = "ZH_HANS") -> dict:
        lang = _attr_language(language)
        attrs = self._category_attrs(cat_id, type_id, language=lang)
        return {"result": attrs, "required": [a for a in attrs if a.get("is_required")], "language": lang}

    def _category_attrs(self, cat_id: int, type_id: int, language: str = "ZH_HANS") -> list[dict]:
        lang = _attr_language(language)
        # 本地优先：缓存命中直接用；没有再调 Ozon 拉取并写回缓存
        cached = self.store.load_category_attrs(cat_id, type_id, lang)
        # 旧缓存缺新增字段（description / category_dependent / is_aspect）→ 视为过期重拉。
        # category_dependent 用于排除"类型"；is_aspect 用于标出「区别特征」(变体维度)。
        if cached and not ("description" in cached[0] and "category_dependent" in cached[0]
                           and "is_aspect" in cached[0] and "max_value_count" in cached[0]):
            cached = None
        if cached is not None:
            return cached
        raw = get_category_attributes(self.store.get_settings(), cat_id, type_id, language=lang)
        attrs = normalize_category_attrs(raw)
        # 只缓存非空结果：空大概率是 API 抽风，缓存了会永久屏蔽必填校验
        if attrs:
            self.store.save_category_attrs(cat_id, type_id, attrs, lang)
        return attrs

    def required_check(self, draft_id: int, language: str = "ZH_HANS") -> dict:
        """结构校验 + 类目必填属性校验，给 UI 标记/提示用。"""
        lang = _attr_language(language)
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        errors = list(validate_draft(draft))
        required: list[dict] = []
        optional: list[dict] = []
        aspects: list[dict] = []
        missing: list[dict] = []
        cat, typ = str(draft.get("category_id") or "").strip(), str(draft.get("type_id") or "").strip()
        if cat and typ:
            try:
                attrs = self._category_attrs(int(cat), int(typ), language=lang)
                # 只排除「类型/Тип」这一个属性：它由 type_id 表达，单独填会显示一个填不了的"缺类型"。
                # ⚠️ 不能把所有 category_dependent 全删——「专为/颜色/材料/碗的类型」等真·必填/可选项
                # 也是 category_dependent(字典随类目变)，删了会导致必填项不显示、发布时被 Ozon 拒。
                attrs = [a for a in attrs if not (
                    a.get("category_dependent")
                    and str(a.get("name") or "").strip().lower() in ("类型", "тип", "type")
                )]
                # 品牌(85)写死「无品牌」、原产国写死「中国」——发布时强制填，不在属性区展示让用户填，
                # 要改自己去 Ozon 改。故从所有展示组里排除这两个。
                attrs = [a for a in attrs if int(a.get("id") or 0) != 85 and not _is_country_attr(a)]
                # 「区别特征」(变体维度 is_aspect，如商品颜色)单列一组、置顶突出——合并成一张卡时
                # 各变体靠它区分，最关键，不该埋在折叠里。从必填/可选里拆出，避免重复展示。
                aspects = [a for a in attrs if a.get("is_aspect")]
                required = [a for a in attrs if a.get("is_required") and not a.get("is_aspect")]
                # 可选属性：非必填、非变体维度的全列出，供人工补充
                optional = [a for a in attrs if not a.get("is_required") and not a.get("is_aspect")]
                missing = missing_required_attributes(draft, attrs)
            except Exception as exc:  # noqa: BLE001
                # 没配 key / 拉取失败：不阻断，只提示
                return {"errors": errors, "required": [], "optional": [], "aspects": [], "missing": [],
                        "attr_warning": f"类目必填属性未校验：{exc}"}
        for m in missing:
            errors.append(f"缺必填属性：{m['name']}")
        return {"errors": errors, "required": required, "optional": optional, "aspects": aspects,
                "missing": missing, "language": lang}

    def _ensure_attr_values(self, cat: int, typ: int, attr_id: int,
                            language: str = "RU") -> tuple[list[dict], bool]:
        """确保本地有该属性的全部字典值。返回 (values, oversized)。
        命中缓存直接返回；未命中拉全量并存库；oversized 时 values 存空、走实时搜；
        拉取失败返回 ([], False)（不写缓存，调用方回退实时 search）。
        language：缓存按语言分别存(RU 用于发布时解析；ZH_HANS 用于前端下拉显示中文)。
        注：本方法使用「全量字典值缓存」(category_attr_values_cache 表, save/load_attr_values)，
        区别于前端 LIKE 模糊搜用的 attribute_values_cache 表 (save_attribute_values/find_attribute_values)。"""
        cached = self.store.load_attr_values(cat, typ, attr_id, language=language)
        if cached is not None:
            return cached
        try:
            out = get_attribute_values(self.store.get_settings(), cat, typ, attr_id,
                                       language=language, max_total=2000)
        except Exception:  # noqa: BLE001
            return ([], False)
        oversized = bool(out.get("oversized"))
        store_values: list[dict] = [] if oversized else (out.get("values") or [])
        self.store.save_attr_values(cat, typ, attr_id, store_values, oversized, language=language)
        return (store_values, oversized)

    def attribute_value_options(self, cat: int, typ: int, attr_id: int,
                                language: str = "ZH_HANS") -> dict:
        """某属性的**全量字典选项**(前端下拉用)：先查 DB 缓存，缺了拉 Ozon 全量并回写。
        oversized(字典过大>2000)时 values 为空、前端回退实时搜。"""
        lang = _attr_language(language)
        values, oversized = self._ensure_attr_values(int(cat), int(typ), int(attr_id), language=lang)
        opts = [{"id": int(v.get("id") or 0), "value": str(v.get("value") or "")}
                for v in (values or []) if int(v.get("id") or 0)]
        return {"values": opts, "oversized": oversized, "language": lang}

    @staticmethod
    def _local_match_value(values: list[dict], text: str) -> dict | None:
        """在缓存的字典值里按俄文精确匹配（strip+lower 相等）。
        命中返回 {"dictionary_value_id": id, "value": value}，否则 None。"""
        t = str(text or "").strip().lower()
        if len(t) < 2:
            return None
        for v in values:
            if str(v.get("value") or "").strip().lower() == t:
                vid = int(v.get("id") or 0)
                if vid:
                    return {"dictionary_value_id": vid, "value": str(v.get("value") or text)}
        return None

    def _resolve_values(self, cat: int, typ: int, attr_id: int, texts: list[str],
                        is_collection: bool) -> list[dict]:
        """把文本值解析成 Ozon 字典值 [{dictionary_value_id, value}]。
        **只走实时 search，不缓存俄文字典**(界面只缓存中文选项)——AI 出的属性值是俄文、
        且 Ozon 中文字典不全(如 薄荷绿/哑光白 中文库查无)，故匹配用实时搜，结果不落库。
        auto_map 已并发调用本方法，实时搜的额外耗时被并发摊平。"""
        out: list[dict] = []
        seen: set[int] = set()
        for t in (texts if is_collection else texts[:1]):
            if len(str(t).strip()) < 2:
                continue
            hit = self._search_one_value(cat, typ, attr_id, t)
            if not hit:
                continue
            vid = int(hit["dictionary_value_id"])
            if vid and vid not in seen:
                seen.add(vid)
                out.append({"dictionary_value_id": vid, "value": hit["value"]})
        return out

    def _search_one_value(self, cat: int, typ: int, attr_id: int, text: str) -> dict | None:
        """实时搜单个值（兜底）。精确优先，否则取第一个结果。失败/无结果返回 None。"""
        try:
            resp = search_attribute_values(self.store.get_settings(), cat, typ, attr_id, text, language="RU")
            res = resp.get("result") or []
        except Exception:  # noqa: BLE001
            return None
        if not res:
            return None
        t = str(text).strip().lower()
        hit = next((r for r in res if str(r.get("value") or "").strip().lower() == t), res[0])
        vid = _to_int(hit.get("id"))
        if not vid:
            return None
        return {"dictionary_value_id": vid, "value": str(hit.get("value") or text)}

    def _resolve_pairs_concurrent(self, cat: int, typ: int,
                                  tasks: list[tuple[int, list[str], bool]]) -> dict[int, list[dict]]:
        """并发解析多个字典属性的值。tasks = [(attr_id, texts, is_collection), ...]。
        返回 {attr_id: [{dictionary_value_id, value}, ...]}。本地命中不走网络，未命中并发实时搜。"""
        from concurrent.futures import ThreadPoolExecutor  # noqa: PLC0415
        import contextvars  # noqa: PLC0415
        if not tasks:
            return {}

        def _one(task: tuple[int, list[str], bool]) -> tuple[int, list[dict]]:
            aid, texts, is_coll = task
            return (aid, self._resolve_values(cat, typ, aid, texts, is_coll))

        # ThreadPoolExecutor 不会把父线程的 ContextVar(current_user_id)带进子线程，否则子线程
        # get_settings() 会拿到默认用户、用错 Ozon 凭证。给每个任务复制一份独立的父线程 context
        # （独立副本才能被多个子线程并发 run），在该 context 里执行解析。
        ctxs = [contextvars.copy_context() for _ in tasks]
        with ThreadPoolExecutor(max_workers=min(16, len(tasks))) as ex:
            return dict(ex.map(lambda c, t: c.run(_one, t), ctxs, tasks))

    def list_ai_models(self, kind: str = "", base: str = "", key: str = "",
                       platform: str = "") -> dict:
        """查接口可用模型(GET /v1/models，OpenAI 兼容)。来源优先级：
        ① 传了 platform(平台名) → 用该平台已存的地址+Key；② 传了 base/key → 用传的;
        ③ 都没传 → 用 kind 槽解析出的地址/Key。便于"选平台配模型"和"边配平台边试"两种。"""
        import json as _json  # noqa: PLC0415
        import urllib.request  # noqa: PLC0415
        from backend.settings_migrate import ai_config, ai_platforms  # noqa: PLC0415
        b = str(base or "").strip()
        key2 = str(key or "").strip()
        if str(platform or "").strip():
            p = ai_platforms(self.store.get_settings()).get(str(platform).strip())
            if p:
                b = b or p["base"]
                key2 = key2 or p["key"]
        if not b and kind:
            cfg = ai_config(self.store.get_settings(),
                            kind if kind in ("text", "image", "video", "multimodal") else "text")
            b = b or str(cfg.get("base") or "")
            key2 = key2 or str(cfg.get("key") or "")
        if not b:
            return {"models": [], "error": "未配置接口地址"}
        import time as _time  # noqa: PLC0415
        url = _models_url(b)
        req = urllib.request.Request(
            url, headers={"Authorization": f"Bearer {key2}", "User-Agent": "Mozilla/5.0"})
        data = None
        for attempt in range(3):   # SSL 瞬断/超时退避重试
            try:
                with urllib.request.urlopen(req, timeout=20) as resp:  # noqa: S310
                    data = _json.loads(resp.read())
                break
            except Exception as exc:  # noqa: BLE001
                if attempt < 2:
                    _time.sleep(1.2 * (attempt + 1))
                    continue
                return {"models": [], "error": f"拉取失败(已重试): {exc}"}
        items = data.get("data") if isinstance(data, dict) else data
        models = sorted({str(m.get("id")) for m in (items or [])
                         if isinstance(m, dict) and m.get("id")})
        return {"models": models}

    def recognize_category(self, draft_id: int) -> dict:
        """AI 识别 Ozon 类别(description_category_id + type_id)并写入草稿。
        **特征是按类别来的**，故这是「特征值识别(auto_map)」的前置步骤。
        复用 navigate_category(只做类别下钻)，比「智能草案」轻——不生成文案/属性。"""
        from backend.ai_card import build_profile, navigate_category  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        raw = draft.get("source_raw") if isinstance(draft.get("source_raw"), dict) else {}
        settings = self.store.get_settings()
        profile = build_profile(raw or {}, understanding=(raw or {}).get("understanding"))
        nav = navigate_category(self._category_roots_zh(settings),
                                self._card_chat(settings, draft), profile)
        if not nav or not nav.get("type_id"):
            return {"ok": False, "matched": False,
                    "note": "AI 没识别出类别，请手动选类目或用「智能草案」"}
        cat = int(nav["description_category_id"])
        typ = int(nav["type_id"])
        path = " / ".join(nav.get("path") or [])
        patch = {"category_id": str(cat), "type_id": str(typ)}
        if path:
            patch["category_path"] = path
        updated = self.store.update_draft(draft_id, patch)
        return {"ok": True, "matched": True, "category_id": cat, "type_id": typ,
                "category_path": path, "category_fallback": nav.get("category_fallback"),
                "draft": updated}

    def _physical_facts(self, draft: dict) -> dict:
        """从草稿 + 理解 specs 提取确定的物理量：净重(g)、包装尺寸(mm)、产品尺寸(mm)、容量(ml)。"""
        facts: dict = {}
        wg = _to_int(draft.get("weight_g"))
        if wg:
            facts["weight_g"] = wg
        L, W, H = (_to_int(draft.get("length_mm")), _to_int(draft.get("width_mm")),
                   _to_int(draft.get("height_mm")))
        if L and W and H:
            facts["pkg_dims_mm"] = (L, W, H)
        sr = draft.get("source_raw") if isinstance(draft.get("source_raw"), dict) else {}
        und = sr.get("understanding") if isinstance(sr.get("understanding"), dict) else {}
        specs = und.get("specs") if isinstance(und.get("specs"), dict) else {}
        pd = _parse_dims_mm(specs.get("尺寸"))
        if pd:
            facts["prod_dims_mm"] = pd
        vol = _parse_volume_ml(specs.get("容量"))
        if vol:
            facts["volume_ml"] = vol
        return facts

    @staticmethod
    def _is_physical_attr(attr: dict) -> bool:
        """是否「重量/尺寸/容量」类数字物理属性(代码确定填，不让 AI 猜)。仅限非字典。"""
        if attr.get("dictionary_id"):
            return False
        name = str(attr.get("name") or "").lower()
        return any(k in name for k in (_WEIGHT_KW + _DIM_KW + _VOL_KW))

    def _physical_attr_value(self, attr: dict, facts: dict) -> str | None:
        """按属性名/描述的单位从 facts 算出该填的值；无法判定/无数据→None。
        重量→克(净/毛都用 weight_g)；尺寸→含"包装"用包装尺寸否则产品尺寸，按 mm/cm 输出 LxWxH；容量→ml。"""
        name = str(attr.get("name") or "").lower()
        if any(k in name for k in _WEIGHT_KW):
            wg = facts.get("weight_g")
            return str(wg) if wg else None
        if any(k in name for k in _DIM_KW):
            # "包装"只在名字里判：描述常含"不带包装"会误判(如 #4382「不带包装的尺寸」是产品尺寸)
            is_pkg = ("包装" in name) or ("упаковк" in name)
            dims = facts.get("pkg_dims_mm") if is_pkg else (facts.get("prod_dims_mm") or facts.get("pkg_dims_mm"))
            if not dims:
                return None
            if ("厘米" in name) or ("см" in name) or ("cm" in name):   # mm→cm
                return "x".join((str(v // 10) if v % 10 == 0 else f"{v / 10:.1f}") for v in dims)
            return "x".join(str(int(v)) for v in dims)
        if any(k in name for k in _VOL_KW):
            ml = facts.get("volume_ml")
            return str(int(ml)) if ml else None
        return None

    @staticmethod
    def _is_annotation_attr(attr: dict) -> bool:
        """「简介/Аннотация」类长文本营销字段——复用文案步骤的俄语描述，不让属性 AI 逐模型乱翻。"""
        if attr.get("dictionary_id"):
            return False
        name = str(attr.get("name") or "").lower().strip()
        return ("简介" in name) or ("аннотац" in name) or (name == "описание")

    def ai_fill_attributes(self, draft_id: int) -> dict:
        """用 AI 按草稿**当前类目**填属性(不重新导航类目，避免 AI 选错类目)。
        字典(选项)属性：把该属性的**全部合法选项**(id+俄文值)连同「单选/多选+上限」一起发给 AI，
        让它从选项里选、直接返回 dictionary_value_id —— 比旧「AI 自由写俄语词→逐值实时搜」准得多，
        也省掉实时搜。选项超大(oversized 或 >OPTION_CAP)回退「自由写俄语+实时搜」。
        非字典按 type 提示 AI：String→俄语文本 / Boolean→true·false。
        **重量/尺寸/容量等物理数字不让 AI 猜**：从 AI 清单剔除，由 _physical_* 按 Ozon 单位
        (名字/描述里写明 mm/cm/克/ml)从草稿+specs 确定填——单位永不出错、毛重不漏。
        已特殊处理的 9048/23171/85 不动；其余 AI 填的覆盖/补充。"""
        import json as _json  # noqa: PLC0415
        from backend.ai_card import _SYS_ATTRS_PICK, _extract_json, build_profile  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        cat = int(str(draft.get("category_id") or "0") or 0)
        typ = int(str(draft.get("type_id") or "0") or 0)
        if not cat or not typ:
            return {"error": "请先选好类目（类别识别）再用 AI 填特征"}
        raw = draft.get("source_raw") if isinstance(draft.get("source_raw"), dict) else {}
        settings = self.store.get_settings()
        profile = build_profile(raw or {}, understanding=(raw or {}).get("understanding"))
        all_attrs = [a for a in (self._category_attrs(cat, typ) or []) if int(a.get("id") or 0) != 85]
        facts = self._physical_facts(draft)   # 物理量(重量/尺寸/容量)由代码确定填，下面从 AI 清单剔除
        required = [a for a in all_attrs if a.get("is_required")]
        ordered = required + [a for a in all_attrs if not a.get("is_required")]
        OPTION_CAP = 150   # 选项数 ≤ 此值才整列发给 AI；超过则回退自由写+实时搜(绝不截断选项)
        brief: list[dict] = []
        opt_index: dict = {}   # aid -> {value_id: 俄文值}；AI 选 id 后据此回填 value(免实时搜)
        for a in ordered[:80]:
            aid = int(a.get("id") or 0)
            if not aid:
                continue
            if self._is_physical_attr(a):
                continue   # 物理数字属性不发给 AI（下面代码按单位确定填）
            if self._is_annotation_attr(a):
                continue   # 简介(Аннotация)复用文案描述，不发给 AI
            item = {"id": aid, "name": a.get("name"), "required": bool(a.get("is_required")),
                    "hint": str(a.get("description") or "")[:100]}
            t = str(a.get("type") or "").strip().lower()
            if a.get("dictionary_id"):
                values, oversized = self._ensure_attr_values(cat, typ, aid, language="RU")
                if values and not oversized and len(values) <= OPTION_CAP:
                    opt_index[aid] = {int(v["id"]): str(v.get("value") or "")
                                      for v in values if int(v.get("id") or 0)}
                    item["kind"] = "select_many" if a.get("is_collection") else "select_one"
                    if a.get("is_collection") and a.get("max_value_count"):
                        item["max_values"] = int(a["max_value_count"])
                    item["options"] = [{"id": vid, "value": val} for vid, val in opt_index[aid].items()]
                else:
                    item["kind"] = "text_ru"   # 超大字典/选项过多 → 自由写俄语，事后实时搜
            elif t in ("decimal", "integer", "float", "number"):
                item["kind"] = "number"
            elif t == "boolean":
                item["kind"] = "boolean"
            else:
                item["kind"] = "text_ru"
            brief.append(item)
        chat = self._card_chat(settings, draft)
        try:
            card = _extract_json(chat(_SYS_ATTRS_PICK,
                "Attribute list:\n" + _json.dumps(brief, ensure_ascii=False)
                + "\n\nProduct:\n" + profile))
        except Exception as exc:  # noqa: BLE001
            return {"error": f"AI 属性输出解析失败: {exc}"}
        by_meta = {int(a["id"]): a for a in all_attrs if a.get("id") is not None}
        new_attrs: list[dict] = []
        mapped: list[dict] = []
        unmapped: list[dict] = []
        for ca in (card.get("attributes") or []):
            try:
                aid = int(ca.get("id"))
            except (TypeError, ValueError):
                continue
            meta = by_meta.get(aid)
            if not meta:
                continue
            if self._is_physical_attr(meta):
                continue   # 物理属性只由代码填，忽略 AI 对它的任何输出
            if self._is_annotation_attr(meta):
                continue   # 简介复用文案描述，忽略 AI 输出
            if aid in opt_index:   # 选项属性：AI 返回 value_ids，按 id 直接回填(无需实时搜)
                ids = ca.get("value_ids")
                if not isinstance(ids, list):
                    ids = [ids] if ids not in (None, "") else []
                vals = []
                for i in ids:
                    try:
                        vid = int(i)
                    except (TypeError, ValueError):
                        continue
                    if vid in opt_index[aid]:
                        vals.append({"dictionary_value_id": vid, "value": opt_index[aid][vid]})
                if not meta.get("is_collection"):
                    vals = vals[:1]
                if vals:
                    new_attrs.append({"id": aid, "values": vals})
                    mapped.append({"id": aid, "name": meta.get("name"),
                                   "value": ", ".join(v["value"] for v in vals)})
                else:
                    unmapped.append({"id": aid, "name": meta.get("name"),
                                     "value": str(ca.get("value_ids") or ca.get("value") or "")})
                continue
            val = str(ca.get("value") or ca.get("value_ru") or "").strip()   # 容错:有的模型用 value_ru
            if not val:
                continue
            if meta.get("dictionary_id"):   # 超大字典回退：自由值→实时搜
                texts = [x.strip() for x in val.replace("，", ",").split(",") if x.strip()]
                resolved = self._resolve_values(cat, typ, aid, texts, bool(meta.get("is_collection")))
                if resolved:
                    new_attrs.append({"id": aid, "values": resolved})
                    mapped.append({"id": aid, "name": meta.get("name"), "value": val})
                else:
                    unmapped.append({"id": aid, "name": meta.get("name"), "value": val})
            else:   # 自由文本/数字/布尔
                new_attrs.append({"id": aid, "values": [{"value": val}]})
                mapped.append({"id": aid, "name": meta.get("name"), "value": val})
        # 物理数字属性(重量/尺寸/容量)由代码按 Ozon 单位确定填——不经 AI，单位永不出错、毛重不漏
        for a in all_attrs:
            if not self._is_physical_attr(a):
                continue
            v = self._physical_attr_value(a, facts)
            if not v:
                continue
            aid = int(a["id"])
            new_attrs.append({"id": aid, "values": [{"value": v}]})
            mapped.append({"id": aid, "name": a.get("name"), "value": v})
        # 简介(Аннотация)复用文案步骤生成的俄语描述，避免属性 AI 逐模型乱翻/重复；文案未跑则留空
        desc = str(draft.get("description") or "").strip()
        if desc:
            for a in all_attrs:
                if self._is_annotation_attr(a):
                    aid = int(a["id"])
                    new_attrs.append({"id": aid, "values": [{"value": desc}]})
                    mapped.append({"id": aid, "name": a.get("name"), "value": desc[:80]})
        keep = {9048, 23171, BRAND_ATTR_ID}  # 型号名/标签/品牌另有处理，AI 不动
        ai_ids = {int(a["id"]) for a in new_attrs if a.get("id") is not None} - keep
        existing = [a for a in (draft.get("attributes") or []) if isinstance(a, dict)]
        merged = [a for a in existing if not (a.get("id") is not None and int(a.get("id")) in ai_ids)]
        merged += [a for a in new_attrs if int(a.get("id")) not in keep]
        updated = self.store.update_draft(draft_id, {"attributes": merged})
        return {"ok": True, "draft": updated, "mapped_count": len(mapped),
                "mapped": mapped, "unmapped": unmapped}

    def auto_map_attributes(self, draft_id: int) -> dict:
        """采集特征(俄文) → 按俄文名对到类目属性(也取俄文) → 解析字典值，自动填进上架属性。
        属性 ID 与语言无关，填进去后中文界面的必填校验照样会减少。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        cat = str(draft.get("category_id") or "").strip()
        typ = str(draft.get("type_id") or "").strip()
        if not cat or not typ:
            return {"error": "请先选好类目再自动填充"}
        cat_i, typ_i = int(cat), int(typ)
        # 采集数据是俄文，属性表也取俄文才能按名对上；走 RU 维度缓存（_category_attrs 已带
        # 缓存+写回），避免每次采集都跨境拉一遍俄文属性表（实测每次省 ~5s）。不污染界面 ZH 缓存。
        meta = self._category_attrs(cat_i, typ_i, language="RU")
        pairs = match_chars_to_attributes(collected_chars(draft), meta)

        # 现有上架格式条目（{id, values}），按 id 索引，便于覆盖/合并
        raw = draft.get("attributes") if isinstance(draft.get("attributes"), list) else []
        publish_by_id: dict[int, dict] = {}
        passthrough: list[dict] = []
        for a in raw:
            if isinstance(a, dict) and "id" in a and "values" in a:
                publish_by_id[_to_int(a.get("id"))] = a
            elif isinstance(a, dict):
                passthrough.append(a)  # 采集来的 {name,value} 参考，原样保留

        mapped: list[dict] = []
        unmapped: list[dict] = []
        dict_tasks: list[tuple[int, list[str], bool]] = []     # (aid, texts, is_collection)
        dict_meta: dict[int, dict] = {}                        # aid -> {name, text}
        free_text: list[tuple[int, str, str]] = []             # (aid, text, name)
        for p in pairs:
            attr, ch = p["attr"], p["char"]
            aid = _to_int(attr.get("id"))
            if aid == BRAND_ATTR_ID:
                continue
            text = str(ch.get("value") or "")
            if attr.get("dictionary_id"):
                dict_tasks.append((aid, split_collection_value(text), bool(attr.get("is_collection"))))
                dict_meta[aid] = {"name": attr.get("name"), "text": text}
            else:  # 自由文本属性
                free_text.append((aid, text, str(attr.get("name") or "")))

        resolved = self._resolve_pairs_concurrent(cat_i, typ_i, dict_tasks)
        for aid, m in dict_meta.items():
            values = resolved.get(aid) or []
            if not values:
                unmapped.append({"id": aid, "name": m["name"], "value": m["text"]})
                continue
            publish_by_id[aid] = {"id": aid, "values": values}
            mapped.append({"id": aid, "name": m["name"],
                           "value": " , ".join(v["value"] for v in values)})
        for aid, text, name in free_text:
            publish_by_id[aid] = {"id": aid, "values": [{"value": text}]}
            mapped.append({"id": aid, "name": name, "value": text})

        self._force_country_to_china(cat_i, typ_i, meta, publish_by_id, mapped)  # 原产国一律中国
        self._fill_model_name(meta, draft, publish_by_id, mapped)  # 型号名称：单品自动填唯一随机值
        new_attributes = passthrough + list(publish_by_id.values())
        brand_id = draft.get("brand_id") if str(draft.get("brand_name") or "").strip() == NO_BRAND else None
        patch: dict = {
            "attributes": new_attributes,
            "brand_id": brand_id if _to_int(brand_id) > 0 else None,
            "brand_name": NO_BRAND,
        }
        updated = self.store.update_draft(draft_id, patch)
        return {"draft": updated, "mapped": mapped, "unmapped": unmapped,
                "mapped_count": len(mapped)}

    def _force_country_to_china(self, cat: int, typ: int, meta: list[dict],
                                publish_by_id: dict, mapped: list[dict]) -> None:
        """原产国一律「中国」：类目里只要有原产国属性(俄文名含 Страна)，就把它的值强制
        设成 Китай(中国)、覆盖竞品采到的任何值；竞品没采到也主动填。与采集内容无关。"""
        for attr in meta:
            if "страна" not in str(attr.get("name") or "").lower():
                continue
            caid = _to_int(attr.get("id"))
            if not caid or caid == BRAND_ATTR_ID:
                continue
            cvals = self._resolve_values(cat, typ, caid, ["Китай"], False)
            if cvals:
                publish_by_id[caid] = {"id": caid, "values": cvals}
                mapped.append({"id": caid, "name": attr.get("name"), "value": "Китай"})

    def _ensure_fixed_attrs(self, draft: dict) -> dict:
        """发布前**写死**：品牌=无品牌、原产国=中国(Китай)，覆盖采集/AI 的任何值。
        不在「特征」让用户填，要改去 Ozon 后台改。即使没跑过自动填充也保证这两项就位。"""
        cat = int(str(draft.get("category_id") or "0") or 0)
        typ = int(str(draft.get("type_id") or "0") or 0)
        if not cat or not typ:
            return draft
        try:
            meta = self._category_attrs(cat, typ, language="RU")
        except Exception:  # noqa: BLE001  拉不到属性表就不强填(不阻断发布)
            return draft
        attrs = [a for a in (draft.get("attributes") or []) if isinstance(a, dict)]
        patch: dict = {}
        # 原产国 → Китай(覆盖)
        country_id = next((_to_int(a.get("id")) for a in meta
                           if _is_country_attr(a) and _to_int(a.get("id")) != BRAND_ATTR_ID), 0)
        if country_id:
            cvals = self._resolve_values(cat, typ, country_id, ["Китай"], False)
            if cvals:
                attrs = [a for a in attrs if _to_int(a.get("id")) != country_id]
                attrs.append({"id": country_id, "values": cvals})
                patch["attributes"] = attrs
        # 品牌 → 无品牌(brand_id 没解析过才解析一次)
        if not (_to_int(draft.get("brand_id")) > 0
                and str(draft.get("brand_name") or "").strip() == NO_BRAND):
            bvals = self._resolve_values(cat, typ, BRAND_ATTR_ID, [NO_BRAND], False)
            if bvals:
                patch["brand_id"] = bvals[0]["dictionary_value_id"]
                patch["brand_name"] = NO_BRAND
        if patch:
            return self.store.update_draft(int(draft["id"]), patch)
        return draft

    def _fill_model_name(self, meta: list[dict], draft: dict,
                         publish_by_id: dict, mapped: list[dict]) -> None:
        """型号名称(9048)：单品自动填一个唯一随机值，让每个单品成独立卡片、并满足必填。
        - 已有 9048（用户填过/竞品采到）不动；
        - 类目没有 9048 这个属性则不填（不塞无效属性）；
        - 变体组草稿跳过：其 9048 在合并发布时由 variant_group 统一填（保证组内合并为一张卡）。"""
        from backend.variant_publish import MODEL_NAME_ATTR_ID  # noqa: PLC0415  # 9048
        if MODEL_NAME_ATTR_ID in publish_by_id:
            return
        if not any(_to_int(a.get("id")) == MODEL_NAME_ATTR_ID for a in (meta or [])):
            return
        sr = draft.get("source_raw")
        vg = str((sr or {}).get("variant_group") or "").strip() if isinstance(sr, dict) else ""
        if vg:
            # 变体组草稿：型号名(9048)= variant_group，组内一致 → Ozon 合并为一张卡。
            # 单条发布也填(原来 return 跳过 → 单发变体时 9048 空缺、型号名发不上去)。
            publish_by_id[MODEL_NAME_ATTR_ID] = {"id": MODEL_NAME_ATTR_ID, "values": [{"value": vg}]}
            mapped.append({"id": MODEL_NAME_ATTR_ID, "name": "型号名称", "value": vg})
            return
        import uuid  # noqa: PLC0415
        token = "M-" + uuid.uuid4().hex[:8].upper()
        publish_by_id[MODEL_NAME_ATTR_ID] = {"id": MODEL_NAME_ATTR_ID, "values": [{"value": token}]}
        mapped.append({"id": MODEL_NAME_ATTR_ID, "name": "型号名称", "value": token})

    def _no_brand_value(self, cat: int, typ: int) -> dict | None:
        """解析"无品牌"(Нет бренда)的字典值，attr 85。找不到返回 None。"""
        vals = self._resolve_values(cat, typ, 85, [NO_BRAND], False)
        return vals[0] if vals else None

    def _category_roots_ru(self, settings: dict) -> list:
        client = None if self.catalog_ru.has_tree_cache() else build_client(settings)
        return self.catalog_ru.raw_tree(client)

    def _category_roots_zh(self, settings: dict) -> list:
        """中文(ZH_HANS)类目树根，喂给 AI 做类目下钻。产品 profile 是中文，
        中文↔中文匹配远比中文↔俄语可靠：「智能喂食器」直接对上「宠物餐具→宠物自动喂食器」，
        不会被摄像头/远程等电子卖点带去「宠物小工具→驱虫器」。返回的 cat/type_id 与语言无关。"""
        client = None if self.catalog.has_tree_cache() else build_client(settings)
        return self.catalog.raw_tree(client)

    def _card_chat(self, settings: dict, draft: dict):
        """卡片/类目下钻/提示词生成用的 chat 函数。
        - engine=agnes：agnes_chat；多模态时附公网主图
        - engine=openai：deepseek_chat；多模态时附公网主图(OpenAI vision)
        多模态由 ai_text.multimodal 决定（取代旧 ai_card_vision）。"""
        from backend.settings_migrate import ai_config, migrate_ai  # noqa: PLC0415
        engine = ai_config(settings, "text")["engine"]
        multimodal = bool(migrate_ai(settings)["ai_text"].get("multimodal"))
        images = None
        if multimodal:
            from backend import agnes  # noqa: PLC0415
            sr = (draft or {}).get("source_raw") or {}
            pics = agnes.pick_public_images((draft or {}).get("images"), sr.get("detail_images")) or []
            images = pics[:1] or None      # 只发主图
        if engine == "agnes":
            from backend import agnes  # noqa: PLC0415
            return lambda s, u: agnes.agnes_chat(settings, s, u, images=images)
        from backend.ai_card import deepseek_chat  # noqa: PLC0415  单测会 monkeypatch 该属性
        return lambda s, u: deepseek_chat(settings, s, u, images=images)

    def ai_generate(self, draft_id: int) -> dict:
        """生成 AI 卡片提案，但 **不写入草稿**。

        返回:
          {"ok": True,
           "proposal": { ...所有将被应用的字段... },
           "report": {"category_fallback":..., "brand_warning":...,
                      "unmapped":..., "mapped":..., "keywords":...,
                      "category_path":...}}

        调用方（前端）应向用户展示预览；用户点「应用」后把 proposal 字段发往
        PATCH /api/drafts/{id}（update_draft），点「放弃」则丢弃。
        """
        from backend.ai_card import generate_card  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        raw = draft.get("source_raw") or {}
        if not raw:
            raw = {"title": draft.get("source_title") or draft.get("ozon_title") or "",
                   "params": draft.get("attributes") if isinstance(draft.get("attributes"), list) else [],
                   "description_text": draft.get("description") or ""}
        settings = self.store.get_settings()
        r = generate_card(
            raw,
            chat=self._card_chat(settings, draft),
            # AI 从根类目逐层下钻到末级类型（取代旧的"关键词→搜索候选→选一个"两步）。
            # 喂中文树：中文↔中文匹配，避免被电子卖点带偏类目（喂食器≠驱虫器）。
            category_roots=self._category_roots_zh(settings),
            fetch_required_attrs=self._category_attrs,
            resolve_values=self._resolve_values,
            # 理解层事实(若已生成)并入 profile：让文案基于图上卖点写，简介更厚
            understanding=(raw.get("understanding") if isinstance(raw, dict) else None),
        )
        if not r.get("ok"):
            return r
        cat, typ = int(r["category_id"]), int(r["type_id"])
        # 构造 proposal（仅包含将要被应用的字段，不写 DB）
        proposal: dict = {
            "category_id": r["category_id"],
            "type_id": r["type_id"],
            "ozon_title": r["ozon_title"],
            "description": r["description"],
            "attributes": r["attributes"],
        }
        # AI 从参数表解析到的毛重/尺寸：仅在 >0 时纳入 proposal
        for k in ("weight_g", "length_mm", "width_mm", "height_mm"):
            if _to_int(r.get(k)) > 0:
                proposal[k] = _to_int(r.get(k))
        nb = self._no_brand_value(cat, typ)
        brand_warning: str | None = None
        if nb:
            proposal["brand_id"] = nb["dictionary_value_id"]
            proposal["brand_name"] = nb["value"]
        else:
            proposal["brand_name"] = NO_BRAND
            brand_warning = "无品牌(Нет бренда)未能解析为字典值，发布前请手动确认品牌属性"
        unmapped = list(r["unmapped"])
        if brand_warning:
            unmapped = [{"id": 85, "name": "Бренд", "value": brand_warning}] + unmapped
        report = {
            "category_path": r.get("category_path"),
            "category_fallback": r.get("category_fallback", False),
            "brand_warning": brand_warning,
            "unmapped": unmapped,
            "mapped": r["mapped"],
            "keywords": [],   # 已改树形下钻，不再有关键词步骤（保留键以兼容下游）
        }
        # 组装待确认草案（含缺失必填/可选属性）
        from backend.ai_card import build_proposal_draft  # noqa: PLC0415
        try:
            attrs = self._category_attrs(cat, typ)
            required = [a for a in attrs if a.get("is_required")]
            optional = [a for a in attrs if not a.get("is_required")]
        except Exception:  # noqa: BLE001
            required, optional = [], []
        draft_json = build_proposal_draft(proposal, report, required, optional, ts=utc_now_iso())
        is_auto = bool(settings.get("ai_auto_apply"))
        self.store.set_ai_proposal(draft_id, draft_json)
        if is_auto:
            applied = self.apply_ai_proposal(draft_id)
            return {"ok": True, "mode": "applied", "draft": applied["draft"],
                    "unmapped": applied["unmapped"], "report": report}
        return {"ok": True, "mode": "draft", "proposal": draft_json, "report": report}

    def ai_copy(self, draft_id: int) -> dict:
        """只生成文案(标题/简介/标签)——**1 次 LLM 调用**，不下钻类目、不映射属性，比 ai_generate 快得多。
        结果写 ai_proposal(预览)，用户确认后应用。类目/属性走自动匹配或 Ozon 复制，单独处理。"""
        from backend.ai_card import (  # noqa: PLC0415
            NO_BRAND, _SYS_TITLE, _extract_json, build_profile, build_proposal_draft, clean_hashtags)
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        raw = draft.get("source_raw") or {}
        if not raw:
            raw = {"title": draft.get("source_title") or draft.get("ozon_title") or "",
                   "params": draft.get("attributes") if isinstance(draft.get("attributes"), list) else [],
                   "description_text": draft.get("description") or ""}
        understanding = raw.get("understanding") if isinstance(raw, dict) else None
        profile = build_profile(raw, understanding=understanding)
        chat = self._card_chat(self.store.get_settings(), draft)
        try:
            body = _extract_json(chat(_SYS_TITLE, "Product:\n" + profile))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"AI 文案输出解析失败: {exc}"}
        proposal: dict = {"ozon_title": str(body.get("ozon_title") or ""),
                          "description": str(body.get("description") or ""),
                          "attributes": []}   # 文案只出 标题/简介/标签，不提取品牌/类型
        tags_value = clean_hashtags(body.get("hashtags"))
        mapped: list = []
        if tags_value:
            proposal["attributes"].append({"id": 23171, "values": [{"value": tags_value}]})
            mapped.append({"id": 23171, "name": "#Хештеги", "value": tags_value})
        report = {"category_path": "", "category_fallback": False, "brand_warning": None,
                  "unmapped": [], "mapped": mapped, "keywords": []}
        draft_json = build_proposal_draft(proposal, report, [], [], ts=utc_now_iso())
        # 草案不含品牌/类型 → 这些 key 不进 fields，应用时 apply 跳过,沿用采集到的(品牌恒无品牌、类目有就有)
        for k in ("category_id", "type_id", "category_path", "brand_name", "brand_id"):
            (draft_json.get("fields") or {}).pop(k, None)
        self.store.set_ai_proposal(draft_id, draft_json)
        return {"ok": True, "mode": "draft", "proposal": draft_json, "report": report}

    def ai_image_prompts(self, draft_id: int, n_points: int = 3) -> dict:
        """生成 ChatGPT 出图提示词(主图 + n_points 张卖点图)，不写库。
        同时返回原始图片 URL，供用户手动上传 ChatGPT 当参考图。
        AI 未配置 → deepseek_chat 抛 RuntimeError(路由转 400)。"""
        from backend.ai_card import (  # noqa: PLC0415
            build_image_prompt_input, parse_image_prompts, _SYS_IMG_PROMPTS)
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        # 注意：不能用 `n_points or 3`——n_points=0 是 falsy 会变 3。显式判 None。
        n = max(1, min(int(n_points if n_points is not None else 3), 6))
        settings = self.store.get_settings()
        system = _SYS_IMG_PROMPTS.format(n=n)
        out = self._card_chat(settings, draft)(system, build_image_prompt_input(draft))
        parsed = parse_image_prompts(out, n)
        _sr = draft.get("source_raw") or {}
        detail_images = _sr.get("detail_images") or []
        detail_local = _sr.get("detail_local") or []   # 本地副本(避防盗链)，给缩略图显示
        return {
            "ok": True,
            "main": parsed["main"],
            "selling_points": parsed["selling_points"],
            "source_images": draft.get("images") or [],
            "local_images": draft.get("local_images") or [],
            "detail_images": detail_images,    # 源 URL，给 ChatGPT 当参考(复制全部用)
            "detail_local": detail_local,      # 本地副本，缩略图显示用(避防盗链)
        }

    def _resolve_image_input(self, url: str) -> str:
        """把前端选的源图变成 Agnes 能取到的输入：http(s) 直接用；/media/ 本地副本
        外网不可达 → 读字节转 data URI（图生图官方支持 Data URI Base64；图生视频
        官方只写了 URL，data URI 是尽力而为，被拒会以 Agnes 400 浮出来）。"""
        from backend import agnes  # noqa: PLC0415
        u = str(url or "").strip()
        if u.startswith("http"):
            return u
        if u.startswith("/media/"):
            ext = u.rsplit(".", 1)[-1].lower() if "." in u else ""
            if ext in ("mp4", "mov", "webm", "m4v"):
                raise ValueError(f"源图不能是视频文件: {u}")
            data = _media.read_media_bytes(u)
            if not data:
                raise ValueError(f"本地图读取失败: {u}")
            return agnes.to_data_uri(data, u)
        raise ValueError("源图必须是 http(s) URL 或 /media/ 本地图")

    def ai_generate_image(self, draft_id: int, *, mode: str = "text2img", prompt: str = "",
                          source_url: str | None = None, size: str | None = None,
                          as_main: bool = False) -> dict:
        """Agnes 生成商品图：text2img(营销图) / img2img(白底主图等，保构图改场景)。
        生成结果下载到本地 /media/（不依赖 Agnes URL 存活期），挂进 draft.images；
        发布时由 OSS rehost 链路把本地图上传到你的 OSS。as_main=True 插到首位当主图。"""
        from backend import agnes  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        p = str(prompt or "").strip()
        if not p:
            raise ValueError("提示词不能为空")
        m = str(mode or "text2img").strip().lower()
        settings = self.store.get_settings()
        sources = None
        if m == "img2img":
            if not str(source_url or "").strip():
                raise ValueError("图生图需要选择一张源图")
            sources = [self._resolve_image_input(str(source_url))]
        # 出图引擎 admin 可切：ai_image.engine == gptimage → 走 gen_image；默认 agnes
        from backend.settings_migrate import ai_config  # noqa: PLC0415
        imgcfg = ai_config(settings, "image")
        if imgcfg["engine"] == "gptimage":
            from backend.gen_image import (  # noqa: PLC0415
                create_image, edit_image, images_from_response)
            gcfg = self._gen_image_cfg(settings)
            tmp = None
            try:
                if str(source_url or "").strip():       # img2img：源图落临时文件给 edit
                    tmp = self._local_image_path(str(source_url))
                    resp = edit_image(gcfg, p, [tmp], size=size or _GEN_SIZE, output_format="png")
                else:                                    # text2img
                    resp = create_image(gcfg, p, size=size or _GEN_SIZE, output_format="png")
            finally:
                if tmp and os.path.exists(tmp):
                    os.remove(tmp)
            picked = images_from_response(resp)
            if not picked:
                raise RuntimeError("gptimage 未返回图片")
            data = picked[0]
            remote_url = ""
            fname = "gptimage-ai.png"
        else:
            remote_url = agnes.generate_image(settings, p, size=size or _GEN_SIZE,
                                              source_images=sources)
            data = _download_bytes(remote_url, timeout=120)
            ext = remote_url.rsplit("?", 1)[0].rsplit(".", 1)[-1].lower()
            fname = f"agnes-ai.{ext if ext in ('png', 'jpg', 'jpeg', 'webp') else 'png'}"
        if len(data) > 20 * 1024 * 1024:   # 与上传路由同口径的 20MB 上限
            raise RuntimeError(f"生成图过大({len(data) // 1024 // 1024}MB > 20MB)")
        local = _media.save_upload(f"draft-{draft_id}", fname, data)
        # 生图可能阻塞数分钟——写前重读草稿，别用进入时的旧快照盖掉期间的用户编辑/并发生成
        cur = self.store.get_draft(draft_id)
        if cur is None:
            raise KeyError(f"draft {draft_id} not found")
        images = list(cur.get("images") or [])
        patch: dict = {"images": images}
        if as_main:
            images.insert(0, local)
            # images↔local_images 是按下标配对的平行数组（前端 localMap），头插必须同步补位
            locs = list(cur.get("local_images") or [])
            if locs:
                patch["local_images"] = ["", *locs]
        else:
            images.append(local)
        updated = self.store.update_draft(draft_id, patch)
        return {"ok": True, "draft": updated, "image": local, "remote_url": remote_url}

    def _image_bytes(self, src: str) -> bytes:
        """读一张草稿图的字节：http(s)→下载；/media/→本地副本。"""
        u = str(src or "").strip()
        if u.startswith("http"):
            return _download_bytes(u)
        if u.startswith("/media/"):
            data = _media.read_media_bytes(u)
            if not data:
                raise ValueError(f"本地图读取失败: {u}")
            return data
        raise ValueError("图片必须是 http(s) URL 或 /media/ 本地图")

    def make_infographic(self, draft_id: int, *, source_index: int = 0, heading: str = "",
                         bullets: list | None = None, watermark: str = "") -> dict:
        """把草稿第 source_index 张图做成俄语信息图（底部面板 + 标题 + 要点），可选叠店铺水印，
        存 /media/ 并追加到 draft.images。标题缺省取 ozon_title。引擎无关、纯 PIL（Montserrat）。"""
        from backend.image_compose import add_watermark, compose_infographic  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        imgs = list(draft.get("images") or [])
        if not imgs:
            raise ValueError("草稿没有图片，无法生成信息图（先采集或生成主图）")
        idx = max(0, min(int(source_index), len(imgs) - 1))
        data = self._image_bytes(str(imgs[idx]))
        h = str(heading or draft.get("ozon_title") or draft.get("source_title") or "")
        bl = [str(b) for b in (bullets or []) if str(b).strip()]
        ig = compose_infographic(data, heading=h, bullets=bl, font_path=_FONT_PATH)
        wm = str(watermark or "").strip()
        if wm:
            ig = add_watermark(ig, wm, font_path=_FONT_PATH)
        local = _media.save_upload(f"draft-{draft_id}", "infographic.jpg", ig)
        # 生成可能耗时——写前重读草稿，别盖掉期间编辑
        cur = self.store.get_draft(draft_id)
        if cur is None:
            raise KeyError(f"draft {draft_id} not found")
        images = list(cur.get("images") or [])
        images.append(local)
        updated = self.store.update_draft(draft_id, {"images": images})
        return {"ok": True, "draft": updated, "image": local}

    def _local_image_path(self, src: str) -> str:
        """把一张图(http 或 /media/)落成本地临时文件，返回路径（供 gen_image 的 edit 用）。调用方负责删。"""
        import tempfile  # noqa: PLC0415
        data = self._image_bytes(src)
        fd, path = tempfile.mkstemp(suffix=".png", prefix="ozsrc-")
        try:
            os.write(fd, data)
        finally:
            os.close(fd)
        return path

    def try_copy(self, draft_id: int) -> dict:
        """Ozon 来源草稿：试官方复制(import-by-sku)。可复制→在目标店建复制卡并标记草稿；
        不可复制→返回 copyable=False（前端据此转「原创建卡」分支）。会在你店里真建一张卡。"""
        import re  # noqa: PLC0415
        from backend.listing_build import random_offer_id  # noqa: PLC0415
        from backend.ozon_client_adapter import copy_by_sku  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        # SKU：显式 sku → 商品链接末尾数字（不用 ozon_product_id，那是 Seller API 的另一种 id）
        sku = str(draft.get("sku") or "").strip()
        if not sku:
            m = re.search(r"/product/[^/?#]*?(\d+)/?(?:[?#]|$)", str(draft.get("source_url") or ""))
            sku = m.group(1) if m else ""
        if not sku.isdigit():
            raise ValueError("草稿没有可用的 Ozon SKU（只有带 Ozon 商品链接的草稿能走官方复制）")
        settings = self._settings_for_store(draft.get("store_client_id"))
        offer_id = random_offer_id()
        verdict = copy_by_sku(
            settings, sku=int(sku), offer_id=offer_id,
            price=str(draft.get("price") or "") or None,
            old_price=str(draft.get("old_price") or "") or None,
            currency_code="RUB",
            name=str(draft.get("ozon_title") or draft.get("source_title") or "")[:200] or None)
        if verdict["copyable"]:
            self.store.update_draft(draft_id, {
                "offer_id": offer_id,
                "status": "published" if verdict["status"] == "created" else "draft",
                "publish_response": {"copy": verdict},
            })
        return {"copyable": verdict["copyable"], "status": verdict["status"],
                "offer_id": offer_id if verdict["copyable"] else None,
                "task_id": verdict.get("task_id")}

    def make_rich_content(self, draft_id: int, *, image_indexes: list | None = None) -> dict:
        """把草稿图拼成 Ozon 富文本(billboard 大图序列)，存 draft.source_raw.rich_content_json。
        默认跳过主图(索引0)、其余全进；俄语文字烤在图里(先用「生成俄语信息图」做几张再来这步)。
        发布时 to_ozon_import_item 自动塞属性 11254、rewrite_item_media 把图链换 OSS 直链。"""
        from backend.drafts import loads_json  # noqa: PLC0415
        from backend.listing_build import build_rich_content  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        imgs = list(draft.get("images") or [])
        if not imgs:
            raise ValueError("草稿没有图片，无法生成富文本（先采集/生成图片）")
        if image_indexes:
            sel = [imgs[i] for i in image_indexes if isinstance(i, int) and 0 <= i < len(imgs)]
        else:
            sel = imgs[1:] if len(imgs) > 1 else imgs   # 默认跳过主图，其余进富文本
        if not sel:
            raise ValueError("没有可用于富文本的图片")
        rc = build_rich_content(sel)
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = dict(sr or {})
        sr["rich_content_json"] = rc
        updated = self.store.update_draft(draft_id, {"source_raw": sr})
        return {"ok": True, "draft": updated, "blocks": len(rc["content"])}

    def _multimodal_chat(self, settings: dict):
        """理解层多模态 chat。优先用单独配的「多模态 AI」槽;**没配则自动复用「文本 AI」**
        ——Agnes 本就多模态,翻译那套配置直接拿来看图理解,用户无需再配一个模型。
        返回 chat(system, user, image_urls) -> str。"""
        from backend.settings_migrate import ai_config  # noqa: PLC0415
        kind = "multimodal" if ai_config(settings, "multimodal").get("key") else "text"
        engine = ai_config(settings, kind)["engine"]
        if engine == "agnes":
            from backend import agnes  # noqa: PLC0415
            return lambda s, u, images: agnes.agnes_chat(settings, s, u, images=images, kind=kind)
        from backend.ai_card import deepseek_chat  # noqa: PLC0415
        return lambda s, u, images: deepseek_chat(settings, s, u, images=images, kind=kind)

    def understand_draft(self, draft_id: int, *, force: bool = False) -> dict:
        """理解层:多模态"看图理解"→ understanding,缓存 source_raw.understanding。
        force=True 重抽;已有缓存且非 force → 直接返回(cached=True)。图先经 _resolve_image_input
        转成模型可取链接(http 直用 / /media/ 转 data URI)。"""
        from backend.drafts import loads_json  # noqa: PLC0415
        from backend.understand import understand  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = dict(sr or {})
        cached = isinstance(sr.get("understanding"), dict) and bool(sr["understanding"])
        if force or not cached:
            chat_fn = self._multimodal_chat(self.store.get_settings())
            u = understand(draft, chat_fn, resolve_image=self._resolve_image_input)
            sr["understanding"] = u
        else:
            u = sr["understanding"]
        # 把理解到的尺寸/克重 + 默认定价写进结构化草稿字段（即使缓存命中也回填，便于补旧草稿）
        patch = self._autofill_from_understanding(draft, u)
        patch["source_raw"] = sr
        updated = self.store.update_draft(draft_id, patch)
        autofill = {k: v for k, v in patch.items() if k != "source_raw"}
        return {"ok": True, "understanding": u, "cached": cached and not force,
                "draft": updated, "autofill": autofill}

    def _cost_basis_cny(self, draft: dict) -> float | None:
        """进价(CNY)：cost_cny>0 优先；否则取 1688 采集的 price_display 区间最低价。"""
        try:
            c = float(draft.get("cost_cny") or 0)
            if c > 0:
                return c
        except (TypeError, ValueError):
            pass
        import re  # noqa: PLC0415
        from backend.drafts import loads_json  # noqa: PLC0415
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = sr if isinstance(sr, dict) else {}
        nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", str(sr.get("price_display") or ""))]
        return min(nums) if nums else None

    def _autofill_from_understanding(self, draft: dict, understanding: dict) -> dict:
        """理解 → 结构化字段补丁：尺寸/克重以**采集的包装件重尺为准**，理解(看图)的值仅在草稿没有时兜底
        (看图读到的多是产品尺寸而非物流箱尺寸，Ozon 要的是包装尺寸)；默认定价仅当售价为空时填。"""
        from backend.listing_build import default_pricing, parse_dims_mm, parse_weight_g  # noqa: PLC0415
        specs = (understanding or {}).get("specs") or {}
        patch: dict = {}

        def _empty(v):
            return v in (None, 0, "", "0")
        dims = parse_dims_mm(specs.get("尺寸") or specs.get("size") or "")
        if dims and all(_empty(draft.get(k)) for k in ("length_mm", "width_mm", "height_mm")):
            patch["length_mm"], patch["width_mm"], patch["height_mm"] = dims
        w = parse_weight_g(specs.get("重量") or specs.get("weight") or "")
        if w and _empty(draft.get("weight_g")):
            patch["weight_g"] = w
        if not str(draft.get("price") or "").strip():   # 不覆盖用户已填的售价
            cost = self._cost_basis_cny(draft)
            pr = default_pricing(cost) if cost else None
            if pr:
                patch["price"], patch["old_price"] = str(pr[0]), str(pr[1])
                try:
                    has_cost = float(draft.get("cost_cny") or 0) > 0
                except (TypeError, ValueError):
                    has_cost = False
                if not has_cost:
                    patch["cost_cny"] = cost   # 进价回填，便于用户看到/二次定价稳定
        return patch

    def recommend(self, draft_id: int) -> dict:
        """智能推荐(纯逻辑,不调 AI):据来源 + 缓存的 understanding → 推荐路径(复制/俄化/重做)
        + 逐图默认处理。understanding 未生成时 per_image 为空、仅按来源给路径默认。"""
        from backend.drafts import loads_json  # noqa: PLC0415
        from backend.recommend import recommend_path  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = sr if isinstance(sr, dict) else {}
        understanding = sr.get("understanding") if isinstance(sr.get("understanding"), dict) else None
        source = draft.get("source") or draft.get("source_platform") or ""
        rec = recommend_path(source=source, understanding=understanding, copyable=None)
        return {"ok": True, "recommendation": rec, "has_understanding": understanding is not None}

    def _add_candidate(self, draft_id: int, data: bytes, label: str, *, slot: str = "") -> str:
        """把一张图字节存 /media(候选 key) 并追加到 source_raw.ai_image_candidates,返回本地 URL。
        slot=图集计划槽位 id(可空)：用于把候选关联到计划槽,驱动槽位状态(待做/候选/已应用)。"""
        if len(data) > 20 * 1024 * 1024:
            raise RuntimeError("候选图过大(>20MB)")
        local = _media.save_upload(f"draft-{draft_id}-cand", "cand.png", data)
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = dict(draft.get("source_raw") or {})
        cands = list(sr.get("ai_image_candidates") or [])
        cands.append({"url": local, "angle": label, "slot": slot})
        sr["ai_image_candidates"] = cands
        self.store.update_draft(draft_id, {"source_raw": sr, "status": draft.get("status")})
        return local

    def _gen_image_cfg(self, settings: dict | None = None):
        """按 DB 的 ai_image 槽（接口地址/Key/模型）构造 GenImageConfig；字段留空回退 GPTPLUS5_* 环境变量。
        俄化/重做/AI生图统一走它 → 生图 AI 配置可全部存数据库。"""
        from backend.gen_image import GenImageConfig  # noqa: PLC0415
        from backend.settings_migrate import ai_config  # noqa: PLC0415
        s = settings if settings is not None else self.store.get_settings()
        c = ai_config(s, "image")
        return GenImageConfig(api_key=(c.get("key") or None),
                              base_url=(c.get("base") or None),
                              model=(c.get("model") or None))

    def _edit_source_image(self, draft_id: int, source_index: int, prompt: str) -> bytes:
        """对草稿第 source_index 张图做 gpt-image edit(传源图保产品一致),返回结果字节。"""
        from backend.gen_image import edit_image, images_from_response  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        imgs = list(draft.get("images") or [])
        if not imgs:
            raise ValueError("草稿没有图片（先采集/生成）")
        idx = max(0, min(int(source_index), len(imgs) - 1))
        tmp = self._local_image_path(str(imgs[idx]))
        try:
            resp = edit_image(self._gen_image_cfg(), prompt, [tmp], size=_GEN_SIZE, output_format="png")
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)
        picked = images_from_response(resp)
        if not picked:
            raise RuntimeError("出图未返回图片")
        return picked[0]

    def localize_image(self, draft_id: int, source_index: int = 0) -> dict:
        """单张俄化:把第 source_index 张图上的中文换成俄语(gpt-image edit,保图不变),结果进候选区。"""
        from backend.gen_image import LOCALIZE_PROMPT  # noqa: PLC0415
        data = self._edit_source_image(draft_id, source_index, LOCALIZE_PROMPT)
        url = self._add_candidate(draft_id, data, f"俄化#{source_index}")
        return {"ok": True, "candidate": url, "draft": self.store.get_draft(draft_id)}

    def regen_image(self, draft_id: int, source_index: int = 0, *, role: str = "",
                    heading: str = "", bullets: list | None = None) -> dict:
        """单张重做:按角色 + 俄语文字 重新生成(gpt-image edit 源图),结果进候选区。数字需 QC。"""
        from backend.gen_image import build_infographic_prompt  # noqa: PLC0415
        prompt = build_infographic_prompt(role=role, heading=heading, bullets=bullets)
        data = self._edit_source_image(draft_id, source_index, prompt)
        url = self._add_candidate(draft_id, data, f"重做#{source_index}")
        return {"ok": True, "candidate": url, "draft": self.store.get_draft(draft_id)}

    def whiten_main(self, draft_id: int, source_index: int = 0) -> dict:
        """选第 source_index 张图做白底电商主图(gpt-image edit + 白底提示词),结果进候选区。"""
        from backend.gen_image import WHITE_MAIN_PROMPT  # noqa: PLC0415
        data = self._edit_source_image(draft_id, source_index, WHITE_MAIN_PROMPT)
        url = self._add_candidate(draft_id, data, f"白底主图#{source_index}")
        return {"ok": True, "candidate": url, "draft": self.store.get_draft(draft_id)}

    def scene_image(self, draft_id: int, source_index: int = 0, *, hint: str = "") -> dict:
        """选第 source_index 张图做场景/氛围图(保产品一致 + 放进使用场景),可带场景提示。结果进候选区。"""
        from backend.gen_image import SCENE_PROMPT  # noqa: PLC0415
        prompt = SCENE_PROMPT + (f" Scene hint: {str(hint).strip()}" if str(hint or "").strip() else "")
        data = self._edit_source_image(draft_id, source_index, prompt)
        url = self._add_candidate(draft_id, data, f"场景图#{source_index}")
        return {"ok": True, "candidate": url, "draft": self.store.get_draft(draft_id)}

    def _load_image_plan(self, draft_id: int, *, force: bool = False):
        """取/建图集计划(缓存 source_raw.image_plan)。返回 (draft, sr, plan)。"""
        from backend.drafts import loads_json  # noqa: PLC0415
        from backend.image_plan import build_image_plan  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = dict(sr or {})
        plan = sr.get("image_plan")
        if force or not isinstance(plan, list) or not plan:
            understanding = sr.get("understanding") if isinstance(sr.get("understanding"), dict) else None
            plan = build_image_plan(understanding, draft.get("images"))
            sr["image_plan"] = plan
            self.store.update_draft(draft_id, {"source_raw": sr})
        return draft, sr, plan

    def image_plan(self, draft_id: int, *, force: bool = False) -> dict:
        """图集计划 + 每槽状态(todo 待做 / candidate 候选中 / applied 已应用)。
        状态据"候选图的 slot 标记 + 是否已进 draft.images"推导。"""
        draft, sr, plan = self._load_image_plan(draft_id, force=force)
        images = {str(u) for u in (draft.get("images") or [])}
        by_slot: dict = {}
        for c in (sr.get("ai_image_candidates") or []):
            if isinstance(c, dict) and c.get("slot"):
                by_slot.setdefault(c["slot"], []).append(str(c.get("url") or ""))
        out = []
        for s in plan:
            urls = by_slot.get(s.get("slot_id"), [])
            applied = any(u in images for u in urls)
            out.append({**s, "status": "applied" if applied else ("candidate" if urls else "todo"),
                        "candidate_url": (urls[-1] if urls else "")})
        return {"ok": True, "plan": out}

    def generate_plan_slot(self, draft_id: int, slot_id: str) -> dict:
        """生成图集计划某个槽位的图(按 action 选生成器,用槽的 source_idx 为源)→ 进候选区,标 slot。"""
        from backend.gen_image import (  # noqa: PLC0415
            LOCALIZE_PROMPT, SCENE_PROMPT, WHITE_MAIN_PROMPT, build_infographic_prompt)
        _, _, plan = self._load_image_plan(draft_id)
        slot = next((s for s in plan if s.get("slot_id") == slot_id), None)
        if slot is None:
            raise ValueError(f"图集计划里没有槽位 {slot_id}")
        src = int(slot.get("source_idx") or 0)
        action = slot.get("action")
        if action == "white":
            prompt = WHITE_MAIN_PROMPT
        elif action == "localize":
            prompt = LOCALIZE_PROMPT
        elif action == "scene":
            prompt = SCENE_PROMPT
        elif action == "infographic":
            prompt = build_infographic_prompt(role=str(slot.get("role") or ""),
                                              heading=str(slot.get("heading") or ""),
                                              bullets=slot.get("bullets") or [])
        else:
            raise ValueError(f"未知 action {action}")
        data = self._edit_source_image(draft_id, src, prompt)
        url = self._add_candidate(draft_id, data, str(slot.get("label") or slot_id), slot=slot_id)
        return {"ok": True, "slot_id": slot_id, "candidate": url, "draft": self.store.get_draft(draft_id)}

    def start_ai_video(self, draft_id: int, *, prompt: str | None = None,
                       image_url: str | None = None) -> dict:
        """启动 Agnes 图生视频后台任务（全局单任务，约 5 秒 121帧@24fps）。
        默认用草稿主图 + 商品展示运镜提示词；完成后 _on_ai_video_done 回写草稿。"""
        from backend import agnes  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        settings = self.store.get_settings()
        agnes._conf(settings)   # 没配 key 在启动前就报错（不进后台线程才能让前端看到 400）
        img_in = str(image_url or "").strip()
        if not img_in:
            imgs = draft.get("images") or []
            if not imgs:
                raise ValueError("草稿没有图片，无法图生视频（先采集或上传图片）")
            img_in = str(imgs[0])
        image = self._resolve_image_input(img_in)
        title = str(draft.get("ozon_title") or draft.get("source_title") or "").strip()
        p = str(prompt or "").strip() or (
            "Smooth cinematic product showcase video, slow camera orbit around the product, "
            "soft studio lighting, the product stays centered and visually identical to the "
            f"reference image. Product: {title[:200]}"
        )
        create_fn = lambda: agnes.create_video_task(settings, p, image=image)  # noqa: E731
        query_fn = lambda vid: agnes.query_video(settings, vid)  # noqa: E731
        return ai_video.start_video(create_fn, query_fn, self._on_ai_video_done, draft_id)

    def _on_ai_video_done(self, draft_id: int, url: str) -> None:
        """视频任务完成回调（后台线程）：写 video_url + 下载本地预览副本。
        - 本地副本放 draft-{id}-ai key（避免覆盖采集来的源视频），overwrite=True 防重生成命中旧缓存
        - 下载完再重读草稿（下载可能数十秒，别用旧快照盖掉期间的编辑）
        - 下载失败把 video_local 置空：宁可前端回退播 video_url，也别播上一版的旧副本
        - 显式保留 status：这是后台异步写，不该把已发布草稿悄悄打回 ready/invalid"""
        try:
            vloc = _media.download_video(url, f"draft-{draft_id}-ai", overwrite=True)
        except Exception:  # noqa: BLE001  本地副本失败不影响 video_url 落库
            vloc = ""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            return
        patch: dict = {
            "video_url": url,
            "source_raw": {**(draft.get("source_raw") or {}), "video_local": vloc},
        }
        if draft.get("status"):
            patch["status"] = draft["status"]
        self.store.update_draft(draft_id, patch)

    def ai_video_status(self) -> dict:
        return ai_video.video_status()

    def stop_ai_video(self) -> dict:
        return ai_video.request_stop()

    def start_image_batch(self, draft_id: int, *, source_url: str | None = None) -> dict:
        """启动 Agnes 整套商品图后台任务（12 角度·全图生图·候选区）。
        参考图优先用本地副本（避 1688 防盗链 Agnes 拉不到）；生成结果进候选区不进正式图集。"""
        from backend import agnes  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        settings = self.store.get_settings()
        agnes._conf(settings)   # 没配 key 启动前就报错（不进后台线程，前端能看到 400）
        ref_in = str(source_url or "").strip()
        if not ref_in:
            locs = draft.get("local_images") or []
            imgs = draft.get("images") or []
            ref_in = str(locs[0]) if locs else (str(imgs[0]) if imgs else "")
        if not ref_in:
            raise ValueError("草稿没有图片，无法整套图生图（先采集或上传图片）")
        ref = self._resolve_image_input(ref_in)
        title = str(draft.get("ozon_title") or draft.get("source_title") or "")
        plan = agnes.plan_image_angles(title)
        # 开新批先清空旧候选（避免多次生成累积）
        sr = dict(draft.get("source_raw") or {})
        sr["ai_image_candidates"] = []
        self.store.update_draft(draft_id, {"source_raw": sr, "status": draft.get("status")})
        gen_fn = lambda prompt: agnes.generate_image(settings, prompt, source_images=[ref])  # noqa: E731
        return ai_image_batch.start_batch(gen_fn, self._on_image_candidate, plan, draft_id)

    def _on_image_candidate(self, draft_id: int, angle: str, url: str) -> None:
        """单张候选完成回调（后台线程，串行）：下载本地 + 追加 source_raw.ai_image_candidates。
        Agnes 图 URL 可能过期 → 下载到 /media（候选 key 与正式图分开）。失败抛出由批量层记 failed。"""
        data = _download_bytes(url, timeout=120)
        if len(data) > 20 * 1024 * 1024:
            raise RuntimeError("候选图过大(>20MB)")
        local = _media.save_upload(f"draft-{draft_id}-cand", "cand.png", data)
        draft = self.store.get_draft(draft_id)
        if draft is None:
            return
        sr = dict(draft.get("source_raw") or {})
        cands = list(sr.get("ai_image_candidates") or [])
        cands.append({"url": local, "angle": angle})
        sr["ai_image_candidates"] = cands
        self.store.update_draft(draft_id, {"source_raw": sr, "status": draft.get("status")})

    def image_batch_status(self) -> dict:
        return ai_image_batch.batch_status()

    def stop_image_batch(self) -> dict:
        return ai_image_batch.request_stop()

    def apply_image_candidates(self, draft_id: int, indices: list[int] | None = None) -> dict:
        """把候选区的图加入正式图集 draft.images，清空候选区。
        indices 为空 → 应用全部候选(前端"全部应用"/一键自动都是全应用，不靠前端传索引，
        避免前端响应式滞后导致传空 → 静默不应用)。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = dict(draft.get("source_raw") or {})
        cands = list(sr.get("ai_image_candidates") or [])
        if not cands:
            raise ValueError("没有候选图可应用")
        if indices:
            picked = [cands[i] for i in indices
                      if isinstance(i, int) and 0 <= i < len(cands)]
        else:
            picked = list(cands)   # 不传索引 = 全部应用
        sel = [str(c.get("url") or "") for c in picked if str(c.get("url") or "")]
        if not sel:
            raise ValueError("未选择任何候选图")
        # 记录每张应用图的类型(白底/细节/场景/尺寸/卖点…)，供图集分类/排序/标签
        types = dict(sr.get("image_types") or {})
        for c in picked:
            u = str(c.get("url") or "")
            if u:
                types[u] = _img_type_from_label(c.get("angle") or c.get("slot") or "")
        sr["image_types"] = types
        images = list(draft.get("images") or []) + sel
        sr["ai_image_candidates"] = []   # 应用后清空候选区
        updated = self.store.update_draft(draft_id, {"images": images, "source_raw": sr})
        return {"ok": True, "draft": updated, "added": len(sel)}

    def discard_image_candidates(self, draft_id: int) -> dict:
        """清空候选区（全部丢弃）。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = dict(draft.get("source_raw") or {})
        sr["ai_image_candidates"] = []
        updated = self.store.update_draft(draft_id, {"source_raw": sr, "status": draft.get("status")})
        return {"ok": True, "draft": updated}

    # ---------- 插件桥接（/api/ext/*）----------
    def ext_ping(self) -> dict:
        # 带上 rub_cny 汇率：插件采集 Ozon 卢布价后据此换算成人民币再回传（后端统一 CNY）
        rate = self.store.get_settings().get("rub_cny")
        try:
            rate = float(rate) if rate not in (None, "") else None
        except (TypeError, ValueError):
            rate = None
        return {"ok": True, "name": "ozon-listing-webui", "version": "1", "rub_cny": rate}

    def _ext_cache_video(self, draft_id: int, video_url: str) -> None:
        """为草稿下载视频本地副本（同 collect 流程）：Ozon/1688 CDN 的 <video> 跨域会卡死，
        落 /media/ 同源才能播。best-effort，失败不影响采集。"""
        vurl = str(video_url or "")
        if not vurl.startswith("http"):
            return
        cur = self.store.get_draft(draft_id)
        sr = cur.get("source_raw") if cur else None
        if isinstance(sr, str):
            from backend.drafts import loads_json  # noqa: PLC0415
            sr = loads_json(sr, {})
        sr = sr if isinstance(sr, dict) else {}
        if sr.get("video_local"):
            return
        try:
            from backend.media import download_video  # noqa: PLC0415
            vloc = download_video(vurl, f"draft-{draft_id}")
            if vloc:
                self.store.update_draft(draft_id, {"source_raw": {**sr, "video_local": vloc}})
        except Exception:  # noqa: BLE001
            pass

    def _auto_map_safe(self, draft_id: int) -> None:
        """采集后尽力自动映射属性(类目对上才有效)：把采集的名值对填进 Ozon 上架属性。
        无网/无key/无类目/失败都静默跳过，绝不阻断采集。"""
        try:
            d = self.store.get_draft(draft_id)
            if d and str(d.get("category_id") or "").strip() and str(d.get("type_id") or "").strip():
                self.auto_map_attributes(draft_id)
        except Exception:  # noqa: BLE001
            pass

    def _media_needs_upload(self, draft: dict) -> bool:
        """草稿里是否还有未传到我们 OSS 的媒体（需后台上传 → media_status=pending）。
        插件同步流已把图直传 OSS（URL 在 oss_public_base 下）→ 不需要 → done；
        计划三异步流推原始 ir.ozone.ru 链接 → 需要 → pending。
        没配 oss_public_base 时，有媒体即按 pending（兼容旧行为/媒体异步测试）。"""
        base = str((self.store.get_settings() or {}).get("oss_public_base") or "").rstrip("/")
        urls = list(draft.get("images") or [])
        v = str(draft.get("video_url") or "").strip()
        if v:
            urls.append(v)
        urls = [str(u).strip() for u in urls if str(u or "").strip()]
        if not urls:
            return False
        return any(not (base and u.startswith(base)) for u in urls)

    def _maybe_auto_publish(self, draft_id: int) -> None:
        """采集/媒体就绪后，若用户开了 auto_publish 且草稿可发，则后台发布到 Ozon。
        守卫：开关开 + 草稿存在 + status!=published（幂等）+ media_status!=pending（等媒体）。
        best-effort：整段吞异常，发不出去的草稿原样留 webui 等人工。"""
        try:
            if not bool((self.store.get_settings() or {}).get("auto_publish")):
                return
            draft = self.store.get_draft(draft_id)
            if draft is None:
                return
            if str(draft.get("status") or "") == "published":
                return
            if str(draft.get("media_status") or "done") == "pending":
                return
            self._dispatch_auto_publish(draft_id)
        except Exception:  # noqa: BLE001
            pass

    def _dispatch_auto_publish(self, draft_id: int) -> None:
        """派发后台线程跑 publish()，不阻塞采集（publish 会轮询 Ozon ~20s）。
        复制父线程 context 把 current_user_id 带进子线程（否则用错 Ozon 凭证）。
        抽成单独方法：测试 monkeypatch 本方法即可同步断言、不起真线程。"""
        import contextvars  # noqa: PLC0415
        import threading  # noqa: PLC0415
        ctx = contextvars.copy_context()

        def _run() -> None:
            try:
                ctx.run(self.publish, draft_id)
            except Exception:  # noqa: BLE001
                pass

        threading.Thread(target=_run, daemon=True).start()

    def ext_collect_parsed(self, payload: dict) -> dict:
        url = str(payload.get("url") or "").strip()
        if not url:
            raise ValueError("url required")
        data = payload.get("data") or {}
        _has_media = bool((data.get("images") or [])) or bool(str(data.get("video_url") or "").strip())
        from backend.drafts import create_draft_from_url  # noqa: PLC0415
        scraped = {
            "source_title": data.get("title"),
            "ozon_title": data.get("title"),  # Ozon 竞品标题已是俄语，直接用
            "description": data.get("description"),
            "price": data.get("price"),
            "old_price": data.get("old_price"),
            "images": data.get("images") or [],
            "video_url": data.get("video_url"),
            "weight_g": data.get("weight_g"),
            "length_mm": data.get("length_mm"),
            "width_mm": data.get("width_mm"),
            "height_mm": data.get("height_mm"),
            "category_path": data.get("category_path"),
        }
        # 采集的属性(名值对 {name,value}) → draft.attributes，供 collected_chars → auto-map/AI；
        # 发布时 to_ozon_import_item 只发 {id,values}，名值对自动丢弃（不会误发给 Ozon）。
        collected_attrs = [
            {"name": str(a["name"]).strip(), "value": str(a.get("value") or "").strip()}
            for a in (data.get("attributes") or [])
            if isinstance(a, dict) and a.get("name") and a.get("value")
        ]
        # 主题标签(webHashtags) → 属性 23171「#Хештеги」，多标签空格连接成单值（与 Ozon 卡片惯例一致）
        tags = [str(t).strip() for t in (data.get("hashtags") or []) if str(t).strip()]
        tag_attr = (
            {"id": 23171, "name": "#Хештеги", "values": [{"value": " ".join(tags)}]}
            if tags else None
        )
        attrs_list = list(collected_attrs)
        if tag_attr:
            attrs_list.append(tag_attr)
        if attrs_list:
            scraped["attributes"] = attrs_list
        # 按买家面包屑路径自动匹配卖家类目 → 填 category_id/type_id（没配 key/无匹配则跳过，回退编辑器手选）
        try:
            self._auto_match_category(scraped)   # 类目匹配走本地缓存，快(~0.06s)，保留
        except Exception:  # noqa: BLE001
            pass
        platform = str(data.get("source_platform") or "ozon").strip() or "ozon"
        # 1688 采集价=「进价」而非售价 → 进 cost_cny，按默认倍率算售价(进价×1.3×2)/划线价(售价×1.8)
        if platform == "1688":
            from backend.listing_build import default_pricing  # noqa: PLC0415
            try:
                cost = float(str(scraped.get("price") or "").strip() or 0)
            except (TypeError, ValueError):
                cost = 0.0
            if cost > 0:
                scraped["cost_cny"] = cost
                pr = default_pricing(cost)
                if pr:
                    scraped["price"], scraped["old_price"] = str(pr[0]), str(pr[1])
        new_draft = create_draft_from_url(url, source_platform=platform, scraped=scraped)
        # 草稿绑定店：插件带 store_client_id=当前店；缺省回退默认店(settings.ozon_client_id)，单店用户零感知
        store_cid = str(payload.get("store_client_id") or data.get("store_client_id") or "").strip()
        if not store_cid:
            store_cid = str((self.store.get_settings() or {}).get("ozon_client_id") or "")
        new_draft["store_client_id"] = store_cid
        det = data.get("detail_images") or []
        rcj = data.get("rich_content_json") or None
        vars_ = data.get("variants") or []
        vg = data.get("variant_group") or ""
        sa = data.get("selected_aspects") or []
        sr_new: dict = {}
        # 插件可直接带 source_raw（如 WB 的 options/brand_name，喂 auto-map/AI）
        incoming_sr = data.get("source_raw")
        if isinstance(incoming_sr, dict):
            sr_new.update(incoming_sr)
        if det:
            sr_new["detail_images"] = det
        if rcj:
            sr_new["rich_content_json"] = rcj  # 原始富文本(A+)，发布时复刻
        if vars_:
            sr_new["variants"] = vars_
        if vg:
            sr_new["variant_group"] = vg
        if sa:
            sr_new["selected_aspects"] = sa
        if sr_new:
            new_draft["source_raw"] = sr_new
        # 重复采集去重按店：同来源在不同店各存一份，刷新只补空当前店那一份
        existing = self.store.find_by_source_url(new_draft["source_url"], None, store_cid)
        if existing:
            # 重复采集：用新解析的源字段「补空」（不覆盖用户已编辑/已选的非空字段），
            # 这样旧的、缺描述/克重/尺寸的草稿能被重新采集刷新。
            patch = {}
            for k in ("ozon_title", "description", "price", "old_price", "images",
                      "video_url", "weight_g", "length_mm", "width_mm", "height_mm",
                      "category_id", "type_id"):
                cur = existing.get(k)
                is_empty = cur in (None, "", 0) or (isinstance(cur, list) and not cur)
                val = new_draft.get(k)
                has_val = val not in (None, "", 0) and not (isinstance(val, list) and not val)
                if is_empty and has_val:
                    patch[k] = val
            if sr_new:
                ex_sr = existing.get("source_raw")
                if not isinstance(ex_sr, dict):
                    ex_sr = {}
                merged = dict(ex_sr)
                for kk, vv in sr_new.items():
                    if not merged.get(kk):  # 旧草稿没有才补
                        merged[kk] = vv
                if merged != ex_sr:
                    patch["source_raw"] = merged
            # 属性/标签补空（不覆盖用户已填）：旧草稿没有采集名值对则补；没有 attr 23171 则补标签
            ex_attrs = existing.get("attributes")
            ex_attrs = ex_attrs if isinstance(ex_attrs, list) else []
            merged_attrs = list(ex_attrs)
            changed_attrs = False
            has_collected = any(a.get("name") and a.get("value") and "values" not in a for a in ex_attrs)
            if collected_attrs and not has_collected:
                merged_attrs = collected_attrs + merged_attrs
                changed_attrs = True
            if tag_attr and not any(_to_int(a.get("id")) == 23171 for a in ex_attrs):
                merged_attrs = merged_attrs + [tag_attr]
                changed_attrs = True
            if changed_attrs:
                patch["attributes"] = merged_attrs
            if patch:
                self.store.update_draft(existing["id"], patch)
            if _has_media and self._media_needs_upload(self.store.get_draft(existing["id"]) or existing):
                self.store.set_media_status(existing["id"], "pending")  # 媒体未传 OSS，待后台补
            self._auto_map_safe(existing["id"])   # 采集后自动映射属性（已本地缓存化，快）
            self._maybe_auto_publish(existing["id"])   # 开了 auto_publish 则后台发布
            return {"created": [{"id": existing["id"], "source_title": existing.get("source_title")}], "errors": [], "deduped": True,
                    "auto_publish": bool((self.store.get_settings() or {}).get("auto_publish"))}
        saved = self.store.insert_draft(new_draft)
        if _has_media and self._media_needs_upload(saved):
            self.store.set_media_status(saved["id"], "pending")  # 媒体未传 OSS，待后台补
        self._auto_map_safe(saved["id"])   # 采集后自动映射属性（已本地缓存化，快）
        self._maybe_auto_publish(saved["id"])   # 开了 auto_publish 则后台发布
        return {"created": [{"id": saved["id"], "source_title": saved.get("source_title")}], "errors": [],
                "auto_publish": bool((self.store.get_settings() or {}).get("auto_publish"))}

    def update_draft_media(self, payload: dict) -> dict:
        """插件后台把媒体传完 OSS 后回调：把草稿里命中的原 URL 换成 OSS URL，置 media_status=done。"""
        draft_id = _to_int(payload.get("draft_id"))
        media_map = payload.get("media_map") or {}
        if not draft_id:
            raise ValueError("draft_id required")
        self.store.apply_media_oss(draft_id, dict(media_map))
        self._maybe_auto_publish(draft_id)   # 媒体传完 → 开了 auto_publish 则后台发布
        return {"ok": True}

    def pending_media_drafts(self) -> dict:
        """当前用户媒体待传(pending)的草稿，供插件后台补传。"""
        from backend.store import current_user_id  # noqa: PLC0415
        return {"drafts": self.store.list_pending_media_drafts(current_user_id.get())}

    def ext_add_snapshot(self, payload: dict) -> dict:
        pid = str(payload.get("product_id") or "").strip()
        if not pid:
            raise ValueError("product_id required")
        fc = payload.get("follow_count")
        pmin = payload.get("price_min")
        pmax = payload.get("price_max")
        last = self.store.latest_offer_snapshot(pid)
        if last and last.get("follow_count") == fc and last.get("price_min") == pmin and last.get("price_max") == pmax:
            return {"id": last.get("id"), "deduped": True}
        from backend.drafts import dumps_json, utc_now_iso  # noqa: PLC0415
        snap = {
            "product_id": pid,
            "sku": (str(payload.get("sku")) if payload.get("sku") else None),
            "captured_at": utc_now_iso(),
            "follow_count": fc,
            "price_min": pmin,
            "price_max": pmax,
            "sellers_json": dumps_json(payload.get("sellers") or []),
        }
        return self.store.add_offer_snapshot(snap)

    def ext_snapshots(self, product_id: str) -> dict:
        return {"product_id": str(product_id), "snapshots": self.store.list_offer_snapshots(str(product_id))}

    def apply_ai_proposal(self, draft_id: int) -> dict:
        """把草稿的 AI 待确认草案合并进正式字段，清空草案。
        - fields 里存在的 key 才覆盖（删除的 key 保留正式原值）
        - attributes 有 value 的项解析成上架格式并按 id 合并；空 value 跳过
        返回 {ok, draft, unmapped}。无草案 → ValueError。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        proposal = draft.get("ai_proposal")
        if not proposal:
            raise ValueError("没有待确认的 AI 草案")
        fields = dict(proposal.get("fields") or {})
        patch: dict = {}
        for k in ("ozon_title", "description", "category_id", "type_id",
                  "weight_g", "length_mm", "width_mm", "height_mm"):
            if k in fields:
                patch[k] = fields[k]
        brand_id = fields.get("brand_id", draft.get("brand_id"))
        brand_name = str(fields.get("brand_name") or draft.get("brand_name") or "").strip()
        patch["brand_id"] = brand_id if brand_name == NO_BRAND and _to_int(brand_id) > 0 else None
        patch["brand_name"] = NO_BRAND
        cat = int(str(patch.get("category_id") or draft.get("category_id") or 0) or 0)
        typ = int(str(patch.get("type_id") or draft.get("type_id") or 0) or 0)
        new_attrs: list[dict] = []
        unmapped: list[dict] = []
        for a in (proposal.get("attributes") or []):
            val = str(a.get("value") or "").strip()
            if not val:
                continue
            aid = _to_int(a.get("id"))
            if aid <= 0:
                continue
            if aid == BRAND_ATTR_ID:
                continue
            if aid == 23171:   # #Хештеги 是自由文本，整串直存，不走字典解析(否则有类目时会被解析丢弃)
                new_attrs.append({"id": 23171, "values": [{"value": val}]})
                continue
            texts = [t.strip() for t in val.replace("，", ",").split(",") if t.strip()]
            vals = self._resolve_values(cat, typ, aid, texts, False) if cat and typ else [{"value": val}]
            if vals:
                new_attrs.append({"id": aid, "values": vals})
            else:
                unmapped.append({"id": aid, "name": a.get("name"), "value": val})
        attr_list = draft.get("attributes") if isinstance(draft.get("attributes"), list) else []
        passthrough = [a for a in attr_list if isinstance(a, dict) and not ("id" in a and "values" in a)]
        existing_pub = [a for a in attr_list if isinstance(a, dict) and "id" in a and "values" in a]
        merged = {_to_int(a["id"]): a for a in existing_pub if _to_int(a.get("id")) > 0}
        for a in new_attrs:
            merged[_to_int(a["id"])] = a
        patch["attributes"] = passthrough + list(merged.values())
        self.store.update_draft(draft_id, patch)
        self.store.set_ai_proposal(draft_id, None)
        updated = self.store.get_draft(draft_id)
        return {"ok": True, "draft": updated, "unmapped": unmapped}

    def patch_ai_proposal(self, draft_id: int, patch: dict) -> dict:
        """编辑/删除草案某项，或 discard 清空整份草案，写回 ai_proposal_json。
        op: edit_field{key,value} / delete_field{key} / edit_attr{id,value} /
            delete_attr{id} / discard。无草案 → ValueError；未知 op → ValueError。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        proposal = draft.get("ai_proposal")
        if not proposal:
            raise ValueError("没有待确认的 AI 草案")
        op = str((patch or {}).get("op") or "")
        if op == "discard":
            self.store.set_ai_proposal(draft_id, None)
            return {"ok": True, "proposal": None}
        fields = dict(proposal.get("fields") or {})
        attrs = list(proposal.get("attributes") or [])
        if op == "edit_field":
            fields[str(patch["key"])] = patch.get("value")
        elif op == "delete_field":
            fields.pop(str(patch.get("key")), None)
        elif op == "edit_attr":
            aid = _to_int(patch.get("id")); val = str(patch.get("value") or "")
            for a in attrs:
                if _to_int(a.get("id")) == aid:
                    a["value"] = val
                    break
        elif op == "delete_attr":
            aid = _to_int(patch.get("id"))
            attrs = [a for a in attrs if _to_int(a.get("id")) != aid]
        else:
            raise ValueError(f"未知操作: {op}")
        proposal["fields"] = fields
        proposal["attributes"] = attrs
        self.store.set_ai_proposal(draft_id, proposal)
        return {"ok": True, "proposal": proposal}

    # ===== realFBS 运费路线表（CSV 可维护；存 settings kv 的 JSON blob，首次读灌种子）=====
    _REALFBS_FIELDS = [
        "scoringGroup", "serviceLevel", "provider", "deliveryMethod", "ozonRating",
        "etaDays", "rateText", "batteries", "liquids", "measurements",
        "weightMinG", "weightMaxG", "valueRangeRub", "tarification",
        "volumeFormula", "compensationRub",
    ]
    _REALFBS_NUM = {"ozonRating", "weightMinG", "weightMaxG", "compensationRub"}

    def _realfbs_seed_routes(self) -> list[dict]:
        import json as _json  # noqa: PLC0415
        from pathlib import Path  # noqa: PLC0415
        seed = Path(__file__).resolve().parent / "realfbs_routes_seed.json"
        try:
            return _json.loads(seed.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            return []

    def realfbs_routes(self) -> dict:
        """返回 realFBS 运费路线给定价用。表空则灌种子(141 条)并持久化。"""
        routes = self.store.get_realfbs_routes()
        if routes is None:
            routes = self._realfbs_seed_routes()
            if routes:
                self.store.set_realfbs_routes(routes)
        return {"routes": routes or []}

    def import_realfbs_routes(self, csv_text: str) -> dict:
        """CSV 整表覆盖运费路线。表头须含 _REALFBS_FIELDS 各列；数值列空→None。"""
        import csv as _csv  # noqa: PLC0415
        import io as _io  # noqa: PLC0415

        def _num(v):
            s = str(v or "").strip()
            if s in ("", "None", "nan"):
                return None
            try:
                return float(s)
            except ValueError:
                return None

        reader = _csv.DictReader(_io.StringIO(csv_text or ""))
        routes: list[dict] = []
        for row in reader:
            r: dict = {}
            for f in self._REALFBS_FIELDS:
                raw = row.get(f)
                r[f] = _num(raw) if f in self._REALFBS_NUM else str(raw or "").strip()
            if not r.get("provider") and not r.get("deliveryMethod") and not r.get("rateText"):
                continue   # 跳过全空行
            routes.append(r)
        if not routes:
            raise ValueError("CSV 未解析到任何运费路线（请确认表头含 provider/rateText 等列）")
        self.store.set_realfbs_routes(routes)
        return {"count": len(routes)}

    def export_realfbs_routes_csv(self) -> str:
        """当前运费路线导出为 CSV（给用户下载→Excel 维护→再导入）。"""
        import csv as _csv  # noqa: PLC0415
        import io as _io  # noqa: PLC0415
        routes = self.realfbs_routes()["routes"]
        buf = _io.StringIO()
        w = _csv.DictWriter(buf, fieldnames=self._REALFBS_FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in routes:
            w.writerow({f: ("" if r.get(f) is None else r.get(f)) for f in self._REALFBS_FIELDS})
        return buf.getvalue()

    # ===== realFBS 佣金类目表（Excel 可维护；存 settings kv 的 JSON blob，首次读灌种子）=====
    #   只取 RFBS(=FBS) 三价格档佣金；导入认两种 Excel：Ozon 官方 Tarifs、本工具导出的模板。
    _COMMISSION_PRICE_TIERS_RUB = [1500, 5000]
    _COMMISSION_HEADER_ALIASES = {
        "parentZh": ("父类目(中)", "父类目", "parentzh"),
        "parentEn": ("父类目(英)", "parenten"),
        "subZh": ("子类目(中)", "子类目", "subzh"),
        "subEn": ("子类目(英)", "suben"),
        "r0": ("佣金% 0-1500", "0-1500", "rfbs0", "rfbs_0_1500"),
        "r1": ("佣金% 1500-5000", "1500-5000", "rfbs1"),
        "r2": ("佣金% 5000+", "5000+", "rfbs2"),
    }

    def _commission_seed(self) -> list[dict]:
        import json as _json  # noqa: PLC0415
        from pathlib import Path  # noqa: PLC0415
        seed = Path(__file__).resolve().parent / "commission_categories_seed.json"
        try:
            return _json.loads(seed.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            return []

    def commission_categories(self) -> dict:
        """返回佣金类目给定价用。表空则灌种子(80 类)并持久化。"""
        cats = self.store.get_commission_categories()
        if cats is None:
            cats = self._commission_seed()
            if cats:
                self.store.set_commission_categories(cats)
        return {"categories": cats or [], "priceTiersRub": list(self._COMMISSION_PRICE_TIERS_RUB)}

    @staticmethod
    def _commission_rate(v) -> float | None:
        """单元格 → 佣金小数：>1 视为百分比(/100)，否则视为小数；空/非数/≤0 → None。"""
        s = str(v).strip() if v is not None else ""
        if s in ("", "None", "nan"):
            return None
        try:
            x = float(s.replace("%", "").replace(",", "."))
        except ValueError:
            return None
        if x <= 0:
            return None
        return round(x / 100 if x > 1 else x, 4)

    def _parse_commission_ozon(self, wb) -> list[dict]:
        """认 Ozon 官方 Tarifs：sheet 'MP Tree Tarifs CN'，列 1=父EN 2=父ZH 4=子EN 5=子ZH 6-8=RFBS。"""
        name = next((s for s in wb.sheetnames if "MP Tree Tarifs" in s), None)
        if not name:
            return []
        ws = wb[name]
        out: list[dict] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2 or len(row) < 9:        # 跳过首行('Starts from')与表头行；列数不足跳过
                continue
            sub_en = str(row[4] or "").strip()
            rfbs = [self._commission_rate(row[6]), self._commission_rate(row[7]), self._commission_rate(row[8])]
            if not sub_en or any(r is None for r in rfbs):
                continue
            out.append({
                "parentEn": str(row[1] or "").strip(), "parentZh": str(row[2] or "").strip(),
                "subEn": sub_en, "subZh": str(row[5] or "").strip(), "rfbs": rfbs,
            })
        return out

    def _parse_commission_template(self, wb) -> list[dict]:
        """认本工具导出的模板：按表头名定位列（中英别名），佣金可填百分比或小数。"""
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        norm = [str(h or "").strip().lower() for h in header]

        def col(key) -> int | None:
            for alias in self._COMMISSION_HEADER_ALIASES[key]:
                a = alias.strip().lower()
                if a in norm:
                    return norm.index(a)
            return None

        idx = {k: col(k) for k in self._COMMISSION_HEADER_ALIASES}
        if idx["subEn"] is None or idx["r0"] is None:    # 表头对不上→不是我们的模板
            return []

        def at(row, key):
            j = idx[key]
            return row[j] if (j is not None and j < len(row)) else None

        out: list[dict] = []
        for row in rows:
            sub_en = str(at(row, "subEn") or "").strip()
            rfbs = [self._commission_rate(at(row, "r0")),
                    self._commission_rate(at(row, "r1")),
                    self._commission_rate(at(row, "r2"))]
            if not sub_en or any(r is None for r in rfbs):
                continue
            out.append({
                "parentEn": str(at(row, "parentEn") or "").strip(),
                "parentZh": str(at(row, "parentZh") or "").strip(),
                "subEn": sub_en, "subZh": str(at(row, "subZh") or "").strip(), "rfbs": rfbs,
            })
        return out

    def import_commission_categories_xlsx(self, data: bytes) -> dict:
        """整表覆盖佣金类目。先认 Ozon 官方 Tarifs，再认本工具模板；只取 RFBS(FBS)。"""
        import io as _io  # noqa: PLC0415
        import openpyxl  # noqa: PLC0415
        try:
            wb = openpyxl.load_workbook(_io.BytesIO(data), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"无法读取 Excel：{exc}")
        try:
            cats = self._parse_commission_ozon(wb) or self._parse_commission_template(wb)
        finally:
            wb.close()
        if not cats:
            raise ValueError("未从 Excel 解析到任何佣金类目（请用 Ozon 官方 Tarifs 文件，或本工具导出的模板）")
        # 按 (parentEn, subEn) 去重，保最后一条、保出现顺序
        seen: dict = {}
        order: list = []
        for c in cats:
            k = (c["parentEn"], c["subEn"])
            if k not in seen:
                order.append(k)
            seen[k] = c
        final = [seen[k] for k in order]
        self.store.set_commission_categories(final)
        return {"count": len(final)}

    @staticmethod
    def _commission_pct(x):
        return None if x is None else round(float(x) * 100, 2)

    def export_commission_categories_xlsx(self) -> bytes:
        """当前佣金类目导出为模板 xlsx（中文表头 + 百分比，给用户改完再导入）。"""
        import io as _io  # noqa: PLC0415
        import openpyxl  # noqa: PLC0415
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "佣金表(FBS)"
        ws.append(["父类目(中)", "父类目(英)", "子类目(中)", "子类目(英)",
                   "佣金% 0-1500", "佣金% 1500-5000", "佣金% 5000+"])
        for c in self.commission_categories()["categories"]:
            rfbs = (c.get("rfbs") or []) + [None, None, None]
            ws.append([c.get("parentZh"), c.get("parentEn"), c.get("subZh"), c.get("subEn"),
                       self._commission_pct(rfbs[0]), self._commission_pct(rfbs[1]), self._commission_pct(rfbs[2])])
        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def get_commission_map(self, cat_id: int, type_id: int) -> dict:
        """取某 Ozon 类目记住的 realFBS 佣金类目（命中返回 {parent_en, sub_en, rfbs}）。"""
        return self.store.load_commission_map(cat_id, type_id) or {}

    def save_commission_map(self, payload: dict) -> dict:
        cat = _to_int(payload.get("cat"))
        typ = _to_int(payload.get("type"))
        if not cat or not typ:
            return {"error": "cat/type 必填"}
        self.store.save_commission_map(
            cat, typ, str(payload.get("parent_en") or ""), str(payload.get("sub_en") or ""),
            payload.get("rfbs") if isinstance(payload.get("rfbs"), list) else [],
        )
        return {"ok": True}

    def brand_search(self, cat_id: int, type_id: int, attr_id: int, query: str, language: str = "ZH_HANS") -> dict:
        lang = _attr_language(language)
        # 本地优先：命中即返回；本地没有再调 Ozon 并写回缓存
        local = self.store.find_attribute_values(cat_id, type_id, attr_id, query, lang)
        if local:
            return {"result": local, "source": "local", "count": len(local), "language": lang}
        # Ozon 要求搜索词 ≥2 个字符(runes)，不足就只给本地缓存，不打 API（否则 400）
        if len(query.strip()) < 2:
            return {"result": [], "source": "local", "count": 0,
                    "hint": "输入至少 2 个字符搜索品牌", "language": lang}
        resp = search_attribute_values(self.store.get_settings(), cat_id, type_id, attr_id, query, language=lang)
        values = resp.get("result") or []
        if values:
            self.store.save_attribute_values(cat_id, type_id, attr_id, values, lang)
        return {"result": values, "source": "ozon", "count": len(values), "language": lang}

    def _scid_of(self, store_client_id: str | None) -> str:
        """把(可能为空的)目标店解析成实际 client_id：空→默认店。仓库/订单按它隔离。"""
        return str(self._settings_for_store(store_client_id).get("ozon_client_id") or "")

    def _warehouses_with_delivery(self, scid: str) -> list[dict]:
        """仓库列表，每个仓库挂上它的配送方式（本地按 warehouse_id 分组）。"""
        warehouses = self.store.list_warehouses(scid)
        methods = self.store.list_delivery_methods(scid)
        by_wh: dict[int, list[dict]] = {}
        for m in methods:
            by_wh.setdefault(m.get("warehouse_id"), []).append(m)
        for w in warehouses:
            w["delivery_methods"] = by_wh.get(w.get("warehouse_id"), [])
        return warehouses

    def list_warehouses(self, store_client_id: str | None = None) -> dict:
        scid = self._scid_of(store_client_id)
        return {"warehouses": self._warehouses_with_delivery(scid)}

    def sync_warehouses(self, store_client_id: str | None = None) -> dict:
        from backend.ozon_client_adapter import (  # noqa: PLC0415
            fetch_delivery_methods, fetch_warehouses,
        )
        settings = self._settings_for_store(store_client_id)
        scid = str(settings.get("ozon_client_id") or "")
        items = fetch_warehouses(settings)
        self.store.upsert_warehouses(items, scid)
        # 配送方式挂在仓库下：仓库同步完后逐仓拉取，按店全量替换
        wids = [w.get("warehouse_id") for w in self.store.list_warehouses(scid)]
        methods = fetch_delivery_methods(settings, wids)
        self.store.replace_delivery_methods(methods, scid)
        return {
            "synced": len(items),
            "delivery_methods": len(methods),
            "warehouses": self._warehouses_with_delivery(scid),
        }

    def set_default_warehouse(self, warehouse_id: int, store_client_id: str | None = None) -> dict:
        scid = self._scid_of(store_client_id)
        self.store.set_default_warehouse(int(warehouse_id), scid)
        return {"warehouses": self._warehouses_with_delivery(scid)}

    # ---------- 功能⑤：FBS 备货发货 ----------
    def pull_fbs(self, status: str = "awaiting_packaging", days: int = 14,
                 store_client_id: str | None = None) -> dict:
        from backend.ozon_client_adapter import pull_fbs_postings  # noqa: PLC0415
        settings = self._settings_for_store(store_client_id)
        scid = str(settings.get("ozon_client_id") or "")
        items = pull_fbs_postings(settings, status, days)
        self.store.upsert_postings(items, scid)
        self.store.rebuild_procurement(scid)
        return {"synced": len(items), "procurement": self.store.list_procurement(scid)}

    def list_procurement(self, store_client_id: str | None = None) -> dict:
        scid = self._scid_of(store_client_id)
        return {"procurement": self.store.list_procurement(scid)}

    def set_procurement_state(self, pid: int, purchase_state: str, note: str = "",
                              store_client_id: str | None = None) -> dict:
        self.store.set_procurement_state(int(pid), purchase_state, note=note)
        return {"procurement": self.store.list_procurement(self._scid_of(store_client_id))}

    def ship_posting(self, posting_number: str, store_client_id: str | None = None) -> dict:
        """按 posting 的商品组成一个包发货。不可逆。
        Ozon 发货请求要 product_id（swagger 里与 sku 是不同 ID）；而待发货单只带
        offer_id/sku，没有 product_id。故先用 offer_id 反查真实 product_id（info.id），
        绝不拿 sku 顶替——否则会发错货或被 Ozon 拒。"""
        from backend.ozon_client_adapter import get_ozon_info, ship_fbs  # noqa: PLC0415
        posting = self.store.get_posting(posting_number)
        if not posting:
            raise KeyError(f"posting {posting_number} not found")
        items = posting.get("products") or []
        offer_ids = [str(p.get("offer_id")) for p in items if p.get("offer_id")]
        if not offer_ids:
            raise ValueError("该 posting 无 offer_id，无法解析 product_id")
        settings = self._settings_for_store(store_client_id)
        try:
            info = get_ozon_info(settings, offer_ids)   # {offer_id: {id=product_id, ...}}
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"拉取商品信息失败，无法发货: {exc}") from exc
        products, missing = [], []
        for p in items:
            oid = str(p.get("offer_id") or "")
            pid = (info.get(oid) or {}).get("id")
            if not pid:
                missing.append(oid or "?")
                continue
            products.append({"product_id": int(pid), "quantity": int(p.get("quantity") or 1)})
        if missing:
            raise ValueError(f"无法解析 product_id 的商品: {', '.join(missing)}")
        if not products:
            raise ValueError("该 posting 无可发货商品")
        # warehouse_id 不传给 Ozon：/v4/posting/fbs/ship 接口只接受 posting_number + packages，
        # Ozon 从 posting 本身推断仓库，无需（也不接受）warehouse_id 参数。
        r = ship_fbs(settings, posting_number, [{"products": products}])
        return {"result": r.get("result") or [], "shipped": True, "response": r}

    def publish_variant_group(
        self,
        variant_group: str,
        store_client_id: str | None = None,
        model_name: str | None = None,
    ) -> dict:
        """把同一 variant_group 的所有草稿合并成一张 Ozon 多变体卡批量发布（批量 import）。
        颜色/尺寸字典值通过 search_attribute_values 实时解析；型号名(9048)作合并 key。
        **不可逆外部写**，由路由层二次确认后调用。"""
        from backend.variant_publish import build_group_items  # noqa: PLC0415
        drafts = self.store.list_drafts_by_variant_group(variant_group)
        if not drafts:
            raise ValueError("该分组没有草稿")
        store_settings = self._settings_for_store(store_client_id)
        # 类目取第一条（同组同类目）
        first = drafts[0]
        cat = int(str(first.get("category_id") or "0") or 0)
        typ = int(str(first.get("type_id") or "0") or 0)
        if not cat or not typ:
            raise ValueError("草稿缺类目，无法合并发布（先在编辑器选类目/AI匹配）")
        category_attrs = self._category_attrs(cat, typ)
        # 型号名(合并 key)：默认用分组键(主 SKU)，保证组内一致→合并
        mname = (model_name or "").strip() or str(variant_group)

        # 变体维度：采集没给 selected_aspects 时，从各变体 spec_attrs 跨组推导(颜色/规格变化维度)，
        # 颜色值中文 → 翻成俄语(与俄语 listing 一致 + 能匹配 Ozon 颜色字典)，写回内存 selected_aspects。
        from backend.variant_publish import derive_group_aspects  # noqa: PLC0415
        if not any((d.get("source_raw") or {}).get("selected_aspects") for d in drafts):
            derived = derive_group_aspects(drafts)
            if derived:
                from backend.settings_migrate import ai_config, migrate_ai  # noqa: PLC0415
                from backend.translate import get_engine  # noqa: PLC0415
                _tm = migrate_ai(store_settings)["translate_mode"]
                if _tm == "ai" and not ai_config(store_settings, "text")["key"]:
                    _tm = "manual"
                _engine = get_engine(_tm, store_settings)
                _cache: dict[str, str] = {}

                def _ru(v: str) -> str:
                    v = (v or "").strip()
                    if not v or not _has_cjk(v):
                        return v
                    if v not in _cache:
                        try:
                            _cache[v] = (_engine.translate(v) or v).strip() or v
                        except Exception:  # noqa: BLE001
                            _cache[v] = v
                    return _cache[v]

                for d in drafts:
                    asp = derived.get(d.get("id")) or []
                    if not asp:
                        continue
                    sr = dict(d.get("source_raw") or {})
                    sr["selected_aspects"] = [{"aspect_key": a.get("aspect_key"),
                                               "value": _ru(a.get("value"))} for a in asp]
                    d["source_raw"] = sr

        def resolve_dict(attr_id: int, dictionary_id: int, text: str) -> dict | None:
            try:
                r = search_attribute_values(store_settings, cat, typ, int(attr_id), str(text), limit=5)
                vals = (r or {}).get("result") if isinstance(r, dict) else None
                vals = vals or []
                if vals:
                    v0 = vals[0]
                    vid = v0.get("id") or v0.get("dictionary_value_id")
                    if vid:
                        return {"dictionary_value_id": int(vid), "value": v0.get("value") or text}
            except Exception:  # noqa: BLE001
                return None
            return None

        items = build_group_items(drafts, category_attrs, mname, resolve_dict)
        # 币种换算（内部统一 CNY → 目标店币种；与 publish() 保持一致）
        currency = str(store_settings.get("contract_currency") or "CNY").upper()
        if currency == "RUB":
            rate = float(store_settings.get("rub_cny") or 0) or None
            if rate:
                for it in items:
                    it["currency_code"] = "RUB"
                    it["price"] = str(round(float(it.get("price") or 0) / rate, 2))
                    if it.get("old_price"):
                        it["old_price"] = str(round(float(it["old_price"]) / rate, 2))
        else:
            for it in items:
                it["currency_code"] = currency
        response = publish_items(store_settings, items)
        return {
            "published": True,
            "count": len(items),
            "variant_group": variant_group,
            "model_name": mname,
            "response": response,
        }

    def fbs_label(self, posting_number: str, store_client_id: str | None = None) -> bytes:
        from backend.ozon_client_adapter import fbs_label_pdf  # noqa: PLC0415
        return fbs_label_pdf(self._settings_for_store(store_client_id), [posting_number])

    def delete(self, draft_id: int) -> dict:
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        # 删除草稿只清理本地工作台记录，不级联下架/删除 Ozon 线上商品。
        # 线上商品的下架、归档、删除必须走单独的显式操作，避免误删已发布商品。
        self.store.delete_draft(draft_id)
        return {
            "deleted": True,
            "id": draft_id,
            "ozon_deleted": False,
            "ozon_response": None,
            "ozon_error": None,
            "drafts": self.store.list_drafts(),
        }
