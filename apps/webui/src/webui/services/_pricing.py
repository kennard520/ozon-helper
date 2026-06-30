from __future__ import annotations


class PricingMixin:
    """realFBS 运费路线 + 佣金类目（CSV/Excel 可维护；存 settings kv 的 JSON blob）。"""

    # ===== realFBS 运费路线表 =====
    _REALFBS_FIELDS = [
        "scoringGroup", "serviceLevel", "provider", "deliveryMethod", "ozonRating",
        "etaDays", "rateText", "batteries", "liquids", "measurements",
        "weightMinG", "weightMaxG", "valueRangeRub", "tarification",
        "volumeFormula", "compensationRub",
    ]
    _REALFBS_NUM = {"ozonRating", "weightMinG", "weightMaxG", "compensationRub"}

    def _realfbs_seed_routes(self) -> list[dict]:
        import json as _json  # noqa: PLC0415
        from pathlib import Path  # noqa: PLC0415
        seed = Path(__file__).resolve().parent.parent / "realfbs_routes_seed.json"
        try:
            return _json.loads(seed.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            return []

    def realfbs_routes(self) -> dict:
        """返回 realFBS 运费路线给定价用。表空则灌种子(141 条)并持久化。"""
        routes = self.store.get_realfbs_routes()
        if routes is None:
            routes = self._realfbs_seed_routes()
            if routes:
                self.store.set_realfbs_routes(routes)
        return {"routes": routes or []}

    def import_realfbs_routes(self, csv_text: str) -> dict:
        """CSV 整表覆盖运费路线。表头须含 _REALFBS_FIELDS 各列；数值列空→None。"""
        import csv as _csv  # noqa: PLC0415
        import io as _io  # noqa: PLC0415

        def _num(v):
            s = str(v or "").strip()
            if s in ("", "None", "nan"):
                return None
            try:
                return float(s)
            except ValueError:
                return None

        reader = _csv.DictReader(_io.StringIO(csv_text or ""))
        routes: list[dict] = []
        for row in reader:
            r: dict = {}
            for f in self._REALFBS_FIELDS:
                raw = row.get(f)
                r[f] = _num(raw) if f in self._REALFBS_NUM else str(raw or "").strip()
            if not r.get("provider") and not r.get("deliveryMethod") and not r.get("rateText"):
                continue   # 跳过全空行
            routes.append(r)
        if not routes:
            raise ValueError("CSV 未解析到任何运费路线（请确认表头含 provider/rateText 等列）")
        self.store.set_realfbs_routes(routes)
        return {"count": len(routes)}

    def export_realfbs_routes_csv(self) -> str:
        """当前运费路线导出为 CSV（给用户下载→Excel 维护→再导入）。"""
        import csv as _csv  # noqa: PLC0415
        import io as _io  # noqa: PLC0415
        routes = self.realfbs_routes()["routes"]
        buf = _io.StringIO()
        w = _csv.DictWriter(buf, fieldnames=self._REALFBS_FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in routes:
            w.writerow({f: ("" if r.get(f) is None else r.get(f)) for f in self._REALFBS_FIELDS})
        return buf.getvalue()

    # ===== realFBS 佣金类目表 =====
    _COMMISSION_PRICE_TIERS_RUB = [1500, 5000]
    _COMMISSION_HEADER_ALIASES = {
        "parentZh": ("父类目(中)", "父类目", "parentzh"),
        "parentEn": ("父类目(英)", "parenten"),
        "subZh": ("子类目(中)", "子类目", "subzh"),
        "subEn": ("子类目(英)", "suben"),
        "r0": ("佣金% 0-1500", "0-1500", "rfbs0", "rfbs_0_1500"),
        "r1": ("佣金% 1500-5000", "1500-5000", "rfbs1"),
        "r2": ("佣金% 5000+", "5000+", "rfbs2"),
    }

    def _commission_seed(self) -> list[dict]:
        import json as _json  # noqa: PLC0415
        from pathlib import Path  # noqa: PLC0415
        seed = Path(__file__).resolve().parent.parent / "commission_categories_seed.json"
        try:
            return _json.loads(seed.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            return []

    def commission_categories(self) -> dict:
        """返回佣金类目给定价用。表空则灌种子(80 类)并持久化。"""
        cats = self.store.get_commission_categories()
        if cats is None:
            cats = self._commission_seed()
            if cats:
                self.store.set_commission_categories(cats)
        return {"categories": cats or [], "priceTiersRub": list(self._COMMISSION_PRICE_TIERS_RUB)}

    @staticmethod
    def _commission_rate(v) -> float | None:
        """单元格 → 佣金小数：>1 视为百分比(/100)，否则视为小数；空/非数/≤0 → None。"""
        s = str(v).strip() if v is not None else ""
        if s in ("", "None", "nan"):
            return None
        try:
            x = float(s.replace("%", "").replace(",", "."))
        except ValueError:
            return None
        if x <= 0:
            return None
        return round(x / 100 if x > 1 else x, 4)

    def _parse_commission_ozon(self, wb) -> list[dict]:
        """认 Ozon 官方 Tarifs：sheet 'MP Tree Tarifs CN'，列 1=父EN 2=父ZH 4=子EN 5=子ZH 6-8=RFBS。"""
        name = next((s for s in wb.sheetnames if "MP Tree Tarifs" in s), None)
        if not name:
            return []
        ws = wb[name]
        out: list[dict] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2 or len(row) < 9:        # 跳过首行('Starts from')与表头行；列数不足跳过
                continue
            sub_en = str(row[4] or "").strip()
            rfbs = [self._commission_rate(row[6]), self._commission_rate(row[7]), self._commission_rate(row[8])]
            if not sub_en or any(r is None for r in rfbs):
                continue
            out.append({
                "parentEn": str(row[1] or "").strip(), "parentZh": str(row[2] or "").strip(),
                "subEn": sub_en, "subZh": str(row[5] or "").strip(), "rfbs": rfbs,
            })
        return out

    def _parse_commission_template(self, wb) -> list[dict]:
        """认本工具导出的模板：按表头名定位列（中英别名），佣金可填百分比或小数。"""
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        norm = [str(h or "").strip().lower() for h in header]

        def col(key) -> int | None:
            for alias in self._COMMISSION_HEADER_ALIASES[key]:
                a = alias.strip().lower()
                if a in norm:
                    return norm.index(a)
            return None

        idx = {k: col(k) for k in self._COMMISSION_HEADER_ALIASES}
        if idx["subEn"] is None or idx["r0"] is None:    # 表头对不上→不是我们的模板
            return []

        def at(row, key):
            j = idx[key]
            return row[j] if (j is not None and j < len(row)) else None

        out: list[dict] = []
        for row in rows:
            sub_en = str(at(row, "subEn") or "").strip()
            rfbs = [self._commission_rate(at(row, "r0")),
                    self._commission_rate(at(row, "r1")),
                    self._commission_rate(at(row, "r2"))]
            if not sub_en or any(r is None for r in rfbs):
                continue
            out.append({
                "parentEn": str(at(row, "parentEn") or "").strip(),
                "parentZh": str(at(row, "parentZh") or "").strip(),
                "subEn": sub_en, "subZh": str(at(row, "subZh") or "").strip(), "rfbs": rfbs,
            })
        return out

    def import_commission_categories_xlsx(self, data: bytes) -> dict:
        """整表覆盖佣金类目。先认 Ozon 官方 Tarifs，再认本工具模板；只取 RFBS(FBS)。"""
        import io as _io  # noqa: PLC0415

        import openpyxl  # noqa: PLC0415
        try:
            wb = openpyxl.load_workbook(_io.BytesIO(data), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"无法读取 Excel：{exc}")
        try:
            cats = self._parse_commission_ozon(wb) or self._parse_commission_template(wb)
        finally:
            wb.close()
        if not cats:
            raise ValueError("未从 Excel 解析到任何佣金类目（请用 Ozon 官方 Tarifs 文件，或本工具导出的模板）")
        # 按 (parentEn, subEn) 去重，保最后一条、保出现顺序
        seen: dict = {}
        order: list = []
        for c in cats:
            k = (c["parentEn"], c["subEn"])
            if k not in seen:
                order.append(k)
            seen[k] = c
        final = [seen[k] for k in order]
        self.store.set_commission_categories(final)
        return {"count": len(final)}

    @staticmethod
    def _commission_pct(x):
        return None if x is None else round(float(x) * 100, 2)

    def export_commission_categories_xlsx(self) -> bytes:
        """当前佣金类目导出为模板 xlsx（中文表头 + 百分比，给用户改完再导入）。"""
        import io as _io  # noqa: PLC0415

        import openpyxl  # noqa: PLC0415
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "佣金表(FBS)"
        ws.append(["父类目(中)", "父类目(英)", "子类目(中)", "子类目(英)",
                   "佣金% 0-1500", "佣金% 1500-5000", "佣金% 5000+"])
        for c in self.commission_categories()["categories"]:
            rfbs = (c.get("rfbs") or []) + [None, None, None]
            ws.append([c.get("parentZh"), c.get("parentEn"), c.get("subZh"), c.get("subEn"),
                       self._commission_pct(rfbs[0]), self._commission_pct(rfbs[1]), self._commission_pct(rfbs[2])])
        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def get_commission_map(self, cat_id: int, type_id: int) -> dict:
        """取某 Ozon 类目记住的 realFBS 佣金类目（命中返回 {parent_en, sub_en, rfbs}）。"""
        return self.store.load_commission_map(cat_id, type_id) or {}

    def save_commission_map(self, payload: dict) -> dict:
        from webui.services._helpers import _to_int  # noqa: PLC0415
        cat = _to_int(payload.get("cat"))
        typ = _to_int(payload.get("type"))
        if not cat or not typ:
            return {"error": "cat/type 必填"}
        self.store.save_commission_map(
            cat, typ, str(payload.get("parent_en") or ""), str(payload.get("sub_en") or ""),
            payload.get("rfbs") if isinstance(payload.get("rfbs"), list) else [],
        )
        return {"ok": True}
