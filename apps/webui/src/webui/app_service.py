from __future__ import annotations

import logging
import os
import sys
import threading
from pathlib import Path

log = logging.getLogger("ozon.app")
if not log.handlers:
    h = logging.StreamHandler(sys.stderr)
    h.setFormatter(logging.Formatter("%(asctime)s [ozon] %(levelname)s %(message)s"))
    log.addHandler(h)
    log.setLevel(logging.INFO)
log.propagate = False  # 不往 root logger 传播，避免被 uvicorn 吞掉

ROOT = Path(__file__).resolve().parents[2]   # apps/webui/(frontend dist 在此下)
REPO = ROOT.parents[1]
AUTH_ROOT = REPO / ".auth"
PROFILE_1688 = AUTH_ROOT / "1688_profile"
FRONTEND_DIST = ROOT / "frontend" / "dist"
_FONT_PATH = str(Path(__file__).resolve().parent / "assets" / "Montserrat-VF.ttf")  # 俄语信息图字体（随镜像）
_GEN_SIZE = "1024x1536"   # 全站统一生图尺寸(gpt-image 竖版,生成时即指定,不裁)；与 gen_image.DEFAULT_SIZE 一致

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import time  # noqa: E402
import urllib.request  # noqa: E402
from urllib.parse import urlparse  # noqa: E402

import webui.ai_video as ai_video  # noqa: E402
import webui.media as _media  # noqa: E402
from ozon_common.oss import OssClient  # noqa: E402
from webui.catalog import Catalog  # noqa: E402
from webui.drafts import (  # noqa: E402
    BRAND_ATTR_ID,
    NO_BRAND,
    collected_chars,
    create_draft_from_url,
    dimension_warnings,
    match_chars_to_attributes,
    missing_required_attributes,
    normalize_category_attrs,
    split_collection_value,
    to_ozon_import_item,
    utc_now_iso,
    validate_draft,
)
from webui.media_rehost import needs_rehost, rehost_draft_media  # noqa: E402
from webui.ozon_client_adapter import (  # noqa: E402
    build_client,
    get_attribute_values,
    get_category_attributes,
    get_import_info,
    publish_items,
    search_attribute_values,
)
from webui.settings_migrate import migrate_ai, normalize_stores  # noqa: E402
from webui.store import Store  # noqa: E402

# 钱包出口处 Decimal→float 的字段（DAL 内部保持 Decimal 精确，仅 API 边界转 number）
_ACCOUNT_MONEY_KEYS = ("balance", "total_recharge", "total_consume")
_TXN_MONEY_KEYS = ("amount", "balance_after")

from webui.services._helpers import (  # noqa: E402
    OZON_ATTRIBUTE_LANGUAGES,
    _ATTR_EXCL,
    _DIM_KW,
    _VOL_KW,
    _WEIGHT_KW,
    _attr_language,
    _download_bytes,
    _has_cjk,
    _img_type_from_label,
    _is_country_attr,
    _models_url,
    _money_to_float,
    _parse_dims_mm,
    _parse_volume_ml,
    _parse_weight_g,
    _to_int,
    step_flags,
)  # noqa
from webui.services._auth import AuthMixin  # noqa: E402
from webui.services._settings import SettingsMixin  # noqa: E402
from webui.services._category import CategoryMixin  # noqa: E402
from webui.services._drafts import DraftMixin  # noqa: E402
from webui.services._publish import PublishMixin  # noqa: E402


class App(AuthMixin, SettingsMixin, CategoryMixin, DraftMixin, PublishMixin):
    def __init__(self) -> None:
        self.store = Store()
        self._cand_lock = threading.Lock()   # 候选区读-改-写串行化(图集并发出图时防丢候选)
        self.catalog = Catalog(store=self.store, language="ZH_HANS")
        # 俄语树用来对齐采集回来的俄语类目路径，做自动匹配
        self.catalog_ru = Catalog(store=self.store, language="RU")
        self._ensure_auth_bootstrap()

    # ---------- 鉴权（多用户）----------
    def _ensure_auth_bootstrap(self) -> None:
        """首次启动：生成稳定 JWT 密钥 + 建默认管理员 admin/admin(user_id=1，承接旧数据)。"""
        settings = self.store.get_settings()
        if not settings.get("jwt_secret"):
            self.store.save_settings({"jwt_secret": os.urandom(32).hex()})
        if self.store.count_users() == 0:
            from webui.auth import hash_password  # noqa: PLC0415
            self.store.create_user("admin", hash_password("admin"), role="admin")

    def _physical_facts(self, draft: dict) -> dict:
        """提取确定的物理量(尽量多来源、跨品类)：净重(g)、包装尺寸(mm)、产品尺寸(mm)、容量(ml)。
        包装尺寸/净重优先用草稿结构化字段；其余从 understanding.specs + 1688 原始参数广扫——
        键名各品类不一(容量/净含量/规格/重量/尺寸…)，只在**值带对应单位/LxWxH 格式**时才采纳，
        避免把型号/规格描述里的裸数字误当物理量。"""
        facts: dict = {}
        wg = _to_int(draft.get("weight_g"))
        if wg:
            facts["weight_g"] = wg
        L, W, H = (_to_int(draft.get("length_mm")), _to_int(draft.get("width_mm")),
                   _to_int(draft.get("height_mm")))
        if L and W and H:
            facts["pkg_dims_mm"] = (L, W, H)
        # 候选(键,值)对：理解 specs + 1688 原始属性
        sr = draft.get("source_raw") if isinstance(draft.get("source_raw"), dict) else {}
        und = sr.get("understanding") if isinstance(sr.get("understanding"), dict) else {}
        specs = und.get("specs") if isinstance(und.get("specs"), dict) else {}
        pairs: list = list(specs.items()) if isinstance(specs, dict) else []
        for a in (sr.get("attributes") or []):
            if isinstance(a, dict):
                pairs.append((str(a.get("name") or ""), a.get("value")))

        def _scan(keywords: tuple, parse) -> object:
            for k, v in pairs:
                if any(w in str(k) for w in keywords):
                    out = parse(v)
                    if out:
                        return out
            return None
        if "volume_ml" not in facts:
            ml = _scan(("容量", "体积", "容积", "净含量", "规格", "毫升", "升"), _parse_volume_ml)
            if ml:
                facts["volume_ml"] = ml
        if "weight_g" not in facts:
            g = _scan(("重量", "净重", "毛重", "克重", "净含量", "规格"), _parse_weight_g)
            if g:
                facts["weight_g"] = g
        if "prod_dims_mm" not in facts:
            dm = _scan(("尺寸", "规格", "大小", "长宽高", "外形"), _parse_dims_mm)
            if dm:
                facts["prod_dims_mm"] = dm
        return facts

    @staticmethod
    def _is_physical_attr(attr: dict) -> bool:
        """是否「重量/尺寸/容量」类数字物理属性(代码确定填，不让 AI 猜)。仅限非字典。"""
        if attr.get("dictionary_id"):
            return False
        name = str(attr.get("name") or "").lower()
        return any(k in name for k in (_WEIGHT_KW + _DIM_KW + _VOL_KW))

    def _physical_attr_value(self, attr: dict, facts: dict) -> str | None:
        """按属性名/描述的单位从 facts 算出该填的值；无法判定/无数据→None。单位写在名字或描述里。
        重量→克/千克(按单位)；尺寸→含"包装"用包装尺寸否则产品尺寸,按 mm/cm 输出 LxWxH；容量→毫升/升(按单位)。"""
        import re  # noqa: PLC0415
        name = str(attr.get("name") or "").lower()
        hint = (name + " " + str(attr.get("description") or "")).lower()   # 单位常写在名字「体积，升」或描述里
        if any(k in name for k in _WEIGHT_KW):
            wg = facts.get("weight_g")
            if not wg:
                return None
            # 先判千克(否则"кг"含"г"、"千克"含"克"会被当克)；都没写默认克
            if ("千克" in hint) or ("公斤" in hint) or ("кг" in hint) or re.search(r"(?<![a-zа-яё])kg(?![a-zа-яё])", hint):
                return f"{wg / 1000:g}"   # 单位千克
            return str(int(wg))           # 默认克(克/г/g)
        if any(k in name for k in _DIM_KW):
            # "包装"只在名字里判：描述常含"不带包装"会误判(如 #4382「不带包装的尺寸」是产品尺寸)
            is_pkg = ("包装" in name) or ("упаковк" in name)
            dims = facts.get("pkg_dims_mm") if is_pkg else (facts.get("prod_dims_mm") or facts.get("pkg_dims_mm"))
            if not dims:
                return None
            if ("厘米" in hint) or ("см" in hint) or ("cm" in hint):   # mm→cm
                return "x".join((str(v // 10) if v % 10 == 0 else f"{v / 10:.1f}") for v in dims)
            return "x".join(str(int(v)) for v in dims)
        if any(k in name for k in _VOL_KW):
            ml = facts.get("volume_ml")
            if not ml:
                return None
            # 先判毫升(否则"毫升"含"升"会被当升)；再判升;都没写默认毫升
            if ("毫升" in hint) or ("мл" in hint) or ("ml" in hint):
                return str(int(ml))       # 单位毫升
            if ("升" in hint) or ("литр" in hint) or re.search(r"(?<![а-яё])л(?![а-яё])", hint):
                return f"{ml / 1000:g}"   # 单位升 → ml 换算(如 20000ml→20)
            return str(int(ml))           # 没写单位默认毫升
        return None

    @staticmethod
    def _is_annotation_attr(attr: dict) -> bool:
        """「简介/Аннотация」类长文本营销字段——复用文案步骤的俄语描述，不让属性 AI 逐模型乱翻。"""
        if attr.get("dictionary_id"):
            return False
        name = str(attr.get("name") or "").lower().strip()
        return ("简介" in name) or ("аннотац" in name) or (name == "описание")

    def ai_fill_attributes(self, draft_id: int) -> dict:
        """用 AI 按草稿**当前类目**填属性(不重新导航类目，避免 AI 选错类目)。
        字典(选项)属性：把该属性的**全部合法选项**(id+俄文值)连同「单选/多选+上限」一起发给 AI，
        让它从选项里选、直接返回 dictionary_value_id —— 比旧「AI 自由写俄语词→逐值实时搜」准得多，
        也省掉实时搜。选项超大(oversized 或 >OPTION_CAP)回退「自由写俄语+实时搜」。
        非字典按 type 提示 AI：String→俄语文本 / Boolean→true·false。
        **重量/尺寸/容量等物理数字不让 AI 猜**：从 AI 清单剔除，由 _physical_* 按 Ozon 单位
        (名字/描述里写明 mm/cm/克/ml)从草稿+specs 确定填——单位永不出错、毛重不漏。
        已特殊处理的 9048/23171/85 不动；其余 AI 填的覆盖/补充。"""
        import json as _json  # noqa: PLC0415

        from webui.ai_card import _SYS_ATTRS_PICK, _SYS_TRANSLATE_RU, _extract_json, build_profile  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        cat = int(str(draft.get("category_id") or "0") or 0)
        typ = int(str(draft.get("type_id") or "0") or 0)
        if not cat or not typ:
            return {"error": "请先选好类目（类别识别）再用 AI 填特征"}
        raw = draft.get("source_raw") if isinstance(draft.get("source_raw"), dict) else {}
        settings = self.store.get_settings()
        profile = build_profile(raw or {}, understanding=(raw or {}).get("understanding"))
        # 本变体的区别规格(1688 spec_attrs，如「美式20升」)——喂给 AI 按本变体填区别特征(is_aspect)
        variant_spec = str((raw or {}).get("spec_attrs") or (raw or {}).get("variant_label") or "").strip()
        all_attrs = [a for a in (self._category_attrs(cat, typ) or []) if int(a.get("id") or 0) != 85]
        facts = self._physical_facts(draft)
        # 只把「代码真能填出值」的物理属性排除出 AI 并由代码填；提取不到值的**留给 AI**(避免两头落空)
        phys_fill: dict = {}
        for a in all_attrs:
            if self._is_physical_attr(a):
                v = self._physical_attr_value(a, facts)
                if v:
                    phys_fill[int(a.get("id") or 0)] = (a.get("name"), v)
        required = [a for a in all_attrs if a.get("is_required")]
        ordered = required + [a for a in all_attrs if not a.get("is_required")]
        OPTION_CAP = 150   # 选项数 ≤ 此值才整列发给 AI；超过则回退自由写+实时搜(绝不截断选项)
        brief: list[dict] = []
        opt_index: dict = {}   # aid -> {value_id: 俄文值}；AI 选 id 后据此回填 value(免实时搜)
        for a in ordered[:80]:
            aid = int(a.get("id") or 0)
            if not aid:
                continue
            if aid in phys_fill:
                continue   # 物理量代码已能确定填(下面填)；提取不到的物理属性仍留给 AI
            if self._is_annotation_attr(a):
                continue   # 简介(Аннotация)复用文案描述，不发给 AI
            item = {"id": aid, "name": a.get("name"), "required": bool(a.get("is_required")),
                    "is_aspect": bool(a.get("is_aspect")),   # 变体区别维度(颜色/尺寸/规格)，按本变体 spec 填
                    "hint": str(a.get("description") or "")[:100]}
            t = str(a.get("type") or "").strip().lower()
            if a.get("dictionary_id"):
                values, oversized = self._ensure_attr_values(cat, typ, aid, language="RU")
                if values and not oversized and len(values) <= OPTION_CAP:
                    opt_index[aid] = {int(v["id"]): str(v.get("value") or "")
                                      for v in values if int(v.get("id") or 0)}
                    item["kind"] = "select_many" if a.get("is_collection") else "select_one"
                    if a.get("is_collection") and a.get("max_value_count"):
                        item["max_values"] = int(a["max_value_count"])
                    item["options"] = [{"id": vid, "value": val} for vid, val in opt_index[aid].items()]
                else:
                    item["kind"] = "text_ru"   # 超大字典/选项过多 → 自由写俄语，事后实时搜
            elif t in ("decimal", "integer", "float", "number"):
                item["kind"] = "number"
            elif t == "boolean":
                item["kind"] = "boolean"
            else:
                item["kind"] = "text_ru"
            brief.append(item)
        chat = self._card_chat(settings, draft)
        try:
            user = "Attribute list:\n" + _json.dumps(brief, ensure_ascii=False) + "\n\nProduct:\n" + profile
            # 结构化变体维度(每轴一条)：1688 给 {axis}、Ozon 给 {aspect,aspect_key}、后端推导给 {aspect_key}
            # ——三源都读到轴名喂 AI(来源无关)。
            sel_aspects = (raw or {}).get("selected_aspects")
            axes_txt = ""
            if isinstance(sel_aspects, list) and sel_aspects:
                def _axis_name(a):
                    ax = a.get("axis") or a.get("aspect")
                    if ax:
                        return str(ax)
                    return {"color": "颜色", "size": "尺寸"}.get(str(a.get("aspect_key") or "").lower(), "维度")
                pairs = [f"{_axis_name(a)}={a.get('value')}" for a in sel_aspects
                         if isinstance(a, dict) and a.get("value")]
                axes_txt = "; ".join(pairs)
            if variant_spec or axes_txt:
                user += (f"\n\nThis is ONE VARIANT of the product. "
                         + (f"Its variant dimensions (axis=value, from 1688): {axes_txt}. " if axes_txt else "")
                         + (f"Its distinguishing spec (Chinese):「{variant_spec}」. " if variant_spec else "")
                         + "Fill the attributes marked is_aspect (the variant-distinguishing ones: color/size/volume/"
                         "material/cap/type) to match THIS variant — map each variant dimension to the matching aspect "
                         "attribute and value (e.g. 颜色=哑光白→color; 规格=20升→volume 20 L; 铝盖→cap aluminium). "
                         "Pick from each aspect's options when given; translate values to Russian.")
            card = _extract_json(chat(_SYS_ATTRS_PICK, user))
        except Exception as exc:  # noqa: BLE001
            return {"error": f"AI 属性输出解析失败: {exc}"}
        by_meta = {int(a["id"]): a for a in all_attrs if a.get("id") is not None}
        new_attrs: list[dict] = []
        mapped: list[dict] = []
        unmapped: list[dict] = []
        cjk_pending: list[tuple] = []   # 自由文本里 AI 照抄的中文(aid,name,val)→ 事后批量翻成俄语
        for ca in (card.get("attributes") or []):
            try:
                aid = int(ca.get("id"))
            except (TypeError, ValueError):
                continue
            meta = by_meta.get(aid)
            if not meta:
                continue
            if aid in phys_fill:
                continue   # 该物理量由代码填，忽略 AI 输出(提取不到的物理属性不在此集合,照常收 AI)
            if self._is_annotation_attr(meta):
                continue   # 简介复用文案描述，忽略 AI 输出
            if aid in opt_index:   # 选项属性：AI 返回 value_ids，按 id 直接回填(无需实时搜)
                ids = ca.get("value_ids")
                if not isinstance(ids, list):
                    ids = [ids] if ids not in (None, "") else []
                vals = []
                for i in ids:
                    try:
                        vid = int(i)
                    except (TypeError, ValueError):
                        continue
                    if vid in opt_index[aid]:
                        vals.append({"dictionary_value_id": vid, "value": opt_index[aid][vid]})
                if not meta.get("is_collection"):
                    vals = vals[:1]
                if vals:
                    new_attrs.append({"id": aid, "values": vals})
                    mapped.append({"id": aid, "name": meta.get("name"),
                                   "value": ", ".join(v["value"] for v in vals)})
                else:
                    unmapped.append({"id": aid, "name": meta.get("name"),
                                     "value": str(ca.get("value_ids") or ca.get("value") or "")})
                continue
            val = str(ca.get("value") or ca.get("value_ru") or "").strip()   # 容错:有的模型用 value_ru
            if not val:
                continue
            if meta.get("dictionary_id"):   # 超大字典回退：自由值→实时搜
                texts = [x.strip() for x in val.replace("，", ",").split(",") if x.strip()]
                resolved = self._resolve_values(cat, typ, aid, texts, bool(meta.get("is_collection")))
                if resolved:
                    new_attrs.append({"id": aid, "values": resolved})
                    mapped.append({"id": aid, "name": meta.get("name"), "value": val})
                else:
                    unmapped.append({"id": aid, "name": meta.get("name"), "value": val})
            elif _has_cjk(val):   # AI 没遵守"必须俄语"，自由文本照抄了中文 → 延后批量翻译，绝不直接填中文
                cjk_pending.append((aid, meta.get("name"), val))
            else:   # 自由文本/数字/布尔
                new_attrs.append({"id": aid, "values": [{"value": val}]})
                mapped.append({"id": aid, "name": meta.get("name"), "value": val})
        # 兜底翻译：自由文本里 AI 照抄的中文 → 一次性翻成俄语再填(如"配套/комплектация")；翻译失败的不填中文
        if cjk_pending:
            ru_map: dict = {}
            try:
                items = [{"id": a, "name": nm, "value_zh": v} for a, nm, v in cjk_pending]
                tr = _extract_json(chat(_SYS_TRANSLATE_RU, _json.dumps(items, ensure_ascii=False)))
                for x in (tr.get("items") or []):
                    try:
                        ru_map[int(x.get("id"))] = str(x.get("value_ru") or "").strip()
                    except (TypeError, ValueError):
                        continue
            except Exception:  # noqa: BLE001  翻译失败 → 全部记未映射，不填中文
                ru_map = {}
            for aid, nm, v in cjk_pending:
                ru = ru_map.get(aid, "")
                if ru and not _has_cjk(ru):
                    new_attrs.append({"id": aid, "values": [{"value": ru}]})
                    mapped.append({"id": aid, "name": nm, "value": ru})
                else:
                    unmapped.append({"id": aid, "name": nm, "value": v})
        # 物理量(重量/尺寸/容量)由代码按 Ozon 单位确定填——不经 AI，单位永不出错、毛重不漏
        for aid, (nm, v) in phys_fill.items():
            new_attrs.append({"id": aid, "values": [{"value": v}]})
            mapped.append({"id": aid, "name": nm, "value": v})
        # 简介(Аннотация)复用文案步骤生成的俄语描述，避免属性 AI 逐模型乱翻/重复；文案未跑则留空
        desc = str(draft.get("description") or "").strip()
        if desc:
            for a in all_attrs:
                if self._is_annotation_attr(a):
                    aid = int(a["id"])
                    new_attrs.append({"id": aid, "values": [{"value": desc}]})
                    mapped.append({"id": aid, "name": a.get("name"), "value": desc[:80]})
        keep = {9048, 23171, BRAND_ATTR_ID}  # 型号名/标签/品牌另有处理，AI 不动
        ai_ids = {int(a["id"]) for a in new_attrs if a.get("id") is not None} - keep
        existing = [a for a in (draft.get("attributes") or []) if isinstance(a, dict)]
        merged = [a for a in existing if not (a.get("id") is not None and int(a.get("id")) in ai_ids)]
        merged += [a for a in new_attrs if int(a.get("id")) not in keep]
        # 型号名称(9048,合并为一张卡用)：变体组**强制** = variant_group(整组一致→Ozon 合并)，覆盖任何旧值；
        # 非变体则随机 M-XXXX(每个单品独立卡)、已填则不覆盖。类目无此属性则不塞。
        import uuid  # noqa: PLC0415

        from webui.variant_publish import MODEL_NAME_ATTR_ID  # noqa: PLC0415  # 9048
        cat_has_model = any(_to_int(a.get("id")) == MODEL_NAME_ATTR_ID for a in all_attrs)
        sr = draft.get("source_raw")
        vg = str((sr or {}).get("variant_group") or "").strip() if isinstance(sr, dict) else ""
        extra_patch: dict = {}
        if cat_has_model:
            mv = ""
            if vg:
                mv = vg   # 变体组：强制 = variant_group
            else:
                has_model = any(_to_int(a.get("id")) == MODEL_NAME_ATTR_ID
                                and any(str(v.get("value") or "").strip() for v in (a.get("values") or []))
                                for a in merged)
                if not has_model:
                    mv = "M-" + uuid.uuid4().hex[:8].upper()
            if mv:
                merged = [a for a in merged if _to_int(a.get("id")) != MODEL_NAME_ATTR_ID]  # 去重旧 9048
                merged.append({"id": MODEL_NAME_ATTR_ID, "values": [{"value": mv}]})
                mapped.append({"id": MODEL_NAME_ATTR_ID, "name": "型号名称", "value": mv})
        # 货号：变体组里若该变体货号为空 → 随机 SP-，保证每个变体唯一货号(Ozon SKU 需唯一)
        if vg and not str(draft.get("offer_id") or "").strip():
            extra_patch["offer_id"] = "SP-" + uuid.uuid4().hex[:8].upper()
        updated = self.store.update_draft(draft_id, {"attributes": merged, **extra_patch})
        return {"ok": True, "draft": updated, "mapped_count": len(mapped),
                "mapped": mapped, "unmapped": unmapped}

    def auto_map_attributes(self, draft_id: int) -> dict:
        """采集特征(俄文) → 按俄文名对到类目属性(也取俄文) → 解析字典值，自动填进上架属性。
        属性 ID 与语言无关，填进去后中文界面的必填校验照样会减少。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        cat = str(draft.get("category_id") or "").strip()
        typ = str(draft.get("type_id") or "").strip()
        if not cat or not typ:
            return {"error": "请先选好类目再自动填充"}
        cat_i, typ_i = int(cat), int(typ)
        # 采集数据是俄文，属性表也取俄文才能按名对上；走 RU 维度缓存（_category_attrs 已带
        # 缓存+写回），避免每次采集都跨境拉一遍俄文属性表（实测每次省 ~5s）。不污染界面 ZH 缓存。
        meta = self._category_attrs(cat_i, typ_i, language="RU")
        pairs = match_chars_to_attributes(collected_chars(draft), meta)

        # 现有上架格式条目（{id, values}），按 id 索引，便于覆盖/合并
        raw = draft.get("attributes") if isinstance(draft.get("attributes"), list) else []
        publish_by_id: dict[int, dict] = {}
        passthrough: list[dict] = []
        for a in raw:
            if isinstance(a, dict) and "id" in a and "values" in a:
                publish_by_id[_to_int(a.get("id"))] = a
            elif isinstance(a, dict):
                passthrough.append(a)  # 采集来的 {name,value} 参考，原样保留

        mapped: list[dict] = []
        unmapped: list[dict] = []
        dict_tasks: list[tuple[int, list[str], bool]] = []     # (aid, texts, is_collection)
        dict_meta: dict[int, dict] = {}                        # aid -> {name, text}
        free_text: list[tuple[int, str, str]] = []             # (aid, text, name)
        for p in pairs:
            attr, ch = p["attr"], p["char"]
            aid = _to_int(attr.get("id"))
            if aid == BRAND_ATTR_ID:
                continue
            text = str(ch.get("value") or "")
            if attr.get("dictionary_id"):
                dict_tasks.append((aid, split_collection_value(text), bool(attr.get("is_collection"))))
                dict_meta[aid] = {"name": attr.get("name"), "text": text}
            else:  # 自由文本属性
                free_text.append((aid, text, str(attr.get("name") or "")))

        resolved = self._resolve_pairs_concurrent(cat_i, typ_i, dict_tasks)
        for aid, m in dict_meta.items():
            values = resolved.get(aid) or []
            if not values:
                unmapped.append({"id": aid, "name": m["name"], "value": m["text"]})
                continue
            publish_by_id[aid] = {"id": aid, "values": values}
            mapped.append({"id": aid, "name": m["name"],
                           "value": " , ".join(v["value"] for v in values)})
        for aid, text, name in free_text:
            publish_by_id[aid] = {"id": aid, "values": [{"value": text}]}
            mapped.append({"id": aid, "name": name, "value": text})

        self._force_country_to_china(cat_i, typ_i, meta, publish_by_id, mapped)  # 原产国一律中国
        self._fill_model_name(meta, draft, publish_by_id, mapped)  # 型号名称：单品自动填唯一随机值
        new_attributes = passthrough + list(publish_by_id.values())
        brand_id = draft.get("brand_id") if str(draft.get("brand_name") or "").strip() == NO_BRAND else None
        patch: dict = {
            "attributes": new_attributes,
            "brand_id": brand_id if _to_int(brand_id) > 0 else None,
            "brand_name": NO_BRAND,
        }
        updated = self.store.update_draft(draft_id, patch)
        return {"draft": updated, "mapped": mapped, "unmapped": unmapped,
                "mapped_count": len(mapped)}

    def _force_country_to_china(self, cat: int, typ: int, meta: list[dict],
                                publish_by_id: dict, mapped: list[dict]) -> None:
        """原产国一律「中国」：类目里只要有原产国属性(俄文名含 Страна)，就把它的值强制
        设成 Китай(中国)、覆盖竞品采到的任何值；竞品没采到也主动填。与采集内容无关。"""
        for attr in meta:
            if "страна" not in str(attr.get("name") or "").lower():
                continue
            caid = _to_int(attr.get("id"))
            if not caid or caid == BRAND_ATTR_ID:
                continue
            cvals = self._resolve_values(cat, typ, caid, ["Китай"], False)
            if cvals:
                publish_by_id[caid] = {"id": caid, "values": cvals}
                mapped.append({"id": caid, "name": attr.get("name"), "value": "Китай"})

    def _ensure_fixed_attrs(self, draft: dict) -> dict:
        """发布前**写死**：品牌=无品牌、原产国/制造国=中国(Китай)，覆盖采集/AI 的任何值。
        不在「特征」让用户填，要改去 Ozon 后台改。即使没跑过自动填充也保证这几项就位。"""
        cat = int(str(draft.get("category_id") or "0") or 0)
        typ = int(str(draft.get("type_id") or "0") or 0)
        if not cat or not typ:
            return draft
        try:
            meta = self._category_attrs(cat, typ, language="RU")
        except Exception:  # noqa: BLE001  拉不到属性表就不强填(不阻断发布)
            return draft
        attrs = [a for a in (draft.get("attributes") or []) if isinstance(a, dict)]
        patch: dict = {}
        # 原产国 / 制造国 等所有「国家」属性 → Китай(全部覆盖，不只第一个；非国家字典里"Китай"解析不到则跳过)
        country_ids = [_to_int(a.get("id")) for a in meta
                       if _is_country_attr(a) and _to_int(a.get("id")) not in (0, BRAND_ATTR_ID)]
        country_changed = False
        for cid in dict.fromkeys(country_ids):   # 去重保序
            cvals = self._resolve_values(cat, typ, cid, ["Китай"], False)
            if cvals:
                attrs = [a for a in attrs if _to_int(a.get("id")) != cid]
                attrs.append({"id": cid, "values": cvals})
                country_changed = True
        if country_changed:
            patch["attributes"] = attrs
        # 品牌 → 无品牌(brand_id 没解析过才解析一次)
        if not (_to_int(draft.get("brand_id")) > 0
                and str(draft.get("brand_name") or "").strip() == NO_BRAND):
            bvals = self._resolve_values(cat, typ, BRAND_ATTR_ID, [NO_BRAND], False)
            if bvals:
                patch["brand_id"] = bvals[0]["dictionary_value_id"]
                patch["brand_name"] = NO_BRAND
        if patch:
            return self.store.update_draft(int(draft["id"]), patch)
        return draft

    def _fill_model_name(self, meta: list[dict], draft: dict,
                         publish_by_id: dict, mapped: list[dict]) -> None:
        """型号名称(9048)：单品自动填一个唯一随机值，让每个单品成独立卡片、并满足必填。
        - 已有 9048（用户填过/竞品采到）不动；
        - 类目没有 9048 这个属性则不填（不塞无效属性）；
        - 变体组草稿跳过：其 9048 在合并发布时由 variant_group 统一填（保证组内合并为一张卡）。"""
        from webui.variant_publish import MODEL_NAME_ATTR_ID  # noqa: PLC0415  # 9048
        if MODEL_NAME_ATTR_ID in publish_by_id:
            return
        if not any(_to_int(a.get("id")) == MODEL_NAME_ATTR_ID for a in (meta or [])):
            return
        sr = draft.get("source_raw")
        vg = str((sr or {}).get("variant_group") or "").strip() if isinstance(sr, dict) else ""
        if vg:
            # 变体组草稿：型号名(9048)= variant_group，组内一致 → Ozon 合并为一张卡。
            # 单条发布也填(原来 return 跳过 → 单发变体时 9048 空缺、型号名发不上去)。
            publish_by_id[MODEL_NAME_ATTR_ID] = {"id": MODEL_NAME_ATTR_ID, "values": [{"value": vg}]}
            mapped.append({"id": MODEL_NAME_ATTR_ID, "name": "型号名称", "value": vg})
            return
        import uuid  # noqa: PLC0415
        token = "M-" + uuid.uuid4().hex[:8].upper()
        publish_by_id[MODEL_NAME_ATTR_ID] = {"id": MODEL_NAME_ATTR_ID, "values": [{"value": token}]}
        mapped.append({"id": MODEL_NAME_ATTR_ID, "name": "型号名称", "value": token})

    def _no_brand_value(self, cat: int, typ: int) -> dict | None:
        """解析"无品牌"(Нет бренда)的字典值，attr 85。找不到返回 None。"""
        vals = self._resolve_values(cat, typ, 85, [NO_BRAND], False)
        return vals[0] if vals else None

    def _category_roots_ru(self, settings: dict) -> list:
        client = None if self.catalog_ru.has_tree_cache() else build_client(settings)
        return self.catalog_ru.raw_tree(client)

    def _category_roots_zh(self, settings: dict) -> list:
        """中文(ZH_HANS)类目树根，喂给 AI 做类目下钻。产品 profile 是中文，
        中文↔中文匹配远比中文↔俄语可靠：「智能喂食器」直接对上「宠物餐具→宠物自动喂食器」，
        不会被摄像头/远程等电子卖点带去「宠物小工具→驱虫器」。返回的 cat/type_id 与语言无关。"""
        client = None if self.catalog.has_tree_cache() else build_client(settings)
        return self.catalog.raw_tree(client)

    def _card_chat(self, settings: dict, draft: dict):
        """卡片/类目下钻/提示词生成用的 chat 函数。
        - engine=agnes：agnes_chat；多模态时附公网主图
        - engine=openai：deepseek_chat；多模态时附公网主图(OpenAI vision)
        多模态由 ai_text.multimodal 决定（取代旧 ai_card_vision）。"""
        from webui.settings_migrate import ai_config, migrate_ai  # noqa: PLC0415
        engine = ai_config(settings, "text")["engine"]
        multimodal = bool(migrate_ai(settings)["ai_text"].get("multimodal"))
        images = None
        if multimodal:
            from webui import agnes  # noqa: PLC0415
            sr = (draft or {}).get("source_raw") or {}
            pics = agnes.pick_public_images((draft or {}).get("images"), sr.get("detail_images")) or []
            images = pics[:1] or None      # 只发主图
        if engine == "agnes":
            from webui import agnes  # noqa: PLC0415
            return lambda s, u: agnes.agnes_chat(settings, s, u, images=images)
        from webui.ai_card import deepseek_chat  # noqa: PLC0415  单测会 monkeypatch 该属性
        return lambda s, u: deepseek_chat(settings, s, u, images=images)

    def ai_generate(self, draft_id: int) -> dict:
        """生成 AI 卡片提案，但 **不写入草稿**。

        返回:
          {"ok": True,
           "proposal": { ...所有将被应用的字段... },
           "report": {"category_fallback":..., "brand_warning":...,
                      "unmapped":..., "mapped":..., "keywords":...,
                      "category_path":...}}

        调用方（前端）应向用户展示预览；用户点「应用」后把 proposal 字段发往
        PATCH /api/drafts/{id}（update_draft），点「放弃」则丢弃。
        """
        from webui.ai_card import generate_card  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        raw_sr = draft.get("source_raw") or {}
        if isinstance(raw_sr, str):
            from webui.drafts import loads_json  # noqa: PLC0415
            raw_sr = loads_json(raw_sr, {})
        raw = dict(raw_sr or {})
        if not str(raw.get("title", "")).strip():
            raw["title"] = draft.get("source_title") or draft.get("ozon_title") or ""
        # 1688 采集 source_raw.attributes = [{name,value}] → params；缺则用草稿 Ozon 格式兜底
        if not raw.get("params"):
            raw["params"] = raw.get("attributes") or (draft.get("attributes") if isinstance(draft.get("attributes"), list) else [])
        if not str(raw.get("description_text", "")).strip():
            raw["description_text"] = draft.get("description") or ""
        settings = self.store.get_settings()
        r = generate_card(
            raw,
            chat=self._card_chat(settings, draft),
            # AI 从根类目逐层下钻到末级类型（取代旧的"关键词→搜索候选→选一个"两步）。
            # 喂中文树：中文↔中文匹配，避免被电子卖点带偏类目（喂食器≠驱虫器）。
            category_roots=self._category_roots_zh(settings),
            fetch_required_attrs=self._category_attrs,
            resolve_values=self._resolve_values,
            # 理解层事实(若已生成)并入 profile：让文案基于图上卖点写，简介更厚
            understanding=(raw.get("understanding") if isinstance(raw, dict) else None),
        )
        if not r.get("ok"):
            return r
        cat, typ = int(r["category_id"]), int(r["type_id"])
        # 构造 proposal（仅包含将要被应用的字段，不写 DB）
        proposal: dict = {
            "category_id": r["category_id"],
            "type_id": r["type_id"],
            "ozon_title": r["ozon_title"],
            "description": r["description"],
            "attributes": r["attributes"],
        }
        # AI 从参数表解析到的毛重/尺寸：仅在 >0 时纳入 proposal
        for k in ("weight_g", "length_mm", "width_mm", "height_mm"):
            if _to_int(r.get(k)) > 0:
                proposal[k] = _to_int(r.get(k))
        nb = self._no_brand_value(cat, typ)
        brand_warning: str | None = None
        if nb:
            proposal["brand_id"] = nb["dictionary_value_id"]
            proposal["brand_name"] = nb["value"]
        else:
            proposal["brand_name"] = NO_BRAND
            brand_warning = "无品牌(Нет бренда)未能解析为字典值，发布前请手动确认品牌属性"
        unmapped = list(r["unmapped"])
        if brand_warning:
            unmapped = [{"id": 85, "name": "Бренд", "value": brand_warning}] + unmapped
        report = {
            "category_path": r.get("category_path"),
            "category_fallback": r.get("category_fallback", False),
            "brand_warning": brand_warning,
            "unmapped": unmapped,
            "mapped": r["mapped"],
            "keywords": [],   # 已改树形下钻，不再有关键词步骤（保留键以兼容下游）
        }
        # 组装待确认草案（含缺失必填/可选属性）
        from webui.ai_card import build_proposal_draft  # noqa: PLC0415
        try:
            attrs = self._category_attrs(cat, typ)
            required = [a for a in attrs if a.get("is_required")]
            optional = [a for a in attrs if not a.get("is_required")]
        except Exception:  # noqa: BLE001
            required, optional = [], []
        draft_json = build_proposal_draft(proposal, report, required, optional, ts=utc_now_iso())
        is_auto = bool(settings.get("ai_auto_apply"))
        self.store.set_ai_proposal(draft_id, draft_json)
        if is_auto:
            applied = self.apply_ai_proposal(draft_id)
            return {"ok": True, "mode": "applied", "draft": applied["draft"],
                    "unmapped": applied["unmapped"], "report": report}
        return {"ok": True, "mode": "draft", "proposal": draft_json, "report": report}

    def ai_copy(self, draft_id: int) -> dict:
        """只生成文案(标题/简介/标签)——**1 次 LLM 调用**，不下钻类目、不映射属性，比 ai_generate 快得多。
        结果写 ai_proposal(预览)，用户确认后应用。类目/属性走自动匹配或 Ozon 复制，单独处理。"""
        from webui.ai_card import (  # noqa: PLC0415
            _SYS_TITLE,
            NO_BRAND,
            _extract_json,
            build_profile,
            build_proposal_draft,
            clean_hashtags,
        )
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        raw_sr = draft.get("source_raw") or {}
        if isinstance(raw_sr, str):
            from webui.drafts import loads_json  # noqa: PLC0415
            raw_sr = loads_json(raw_sr, {})
        raw = dict(raw_sr or {})
        if not str(raw.get("title", "")).strip():
            raw["title"] = draft.get("source_title") or draft.get("ozon_title") or ""
        # 1688 采集 source_raw.attributes = [{name,value}] → params；缺则用草稿 Ozon 格式兜底
        if not raw.get("params"):
            raw["params"] = raw.get("attributes") or (draft.get("attributes") if isinstance(draft.get("attributes"), list) else [])
        if not str(raw.get("description_text", "")).strip():
            raw["description_text"] = draft.get("description") or ""
        understanding = raw.get("understanding") if isinstance(raw, dict) else None
        profile = build_profile(raw, understanding=understanding)
        chat = self._card_chat(self.store.get_settings(), draft)
        raw = ""
        try:
            log.info(f"[ai_copy] draft={draft_id} requesting, profile={(profile[:500])}")
            raw = chat(_SYS_TITLE, "Product:\n" + profile)
            log.info(f"[ai_copy] draft={draft_id} response, raw={(str(raw)[:500])}")
            if not str(raw).strip():
                log.error(f"[ai_copy] draft={draft_id} LLM returned empty")
                return {"ok": False, "error": "AI 文案输出为空（模型/网关异常）"}
            body = _extract_json(raw)
            log.info(f"[ai_copy] draft={draft_id} OK title={len(str(body.get('ozon_title','')))} desc={len(str(body.get('description','')))}")
        except Exception as exc:  # noqa: BLE001
            log.exception(f"[ai_copy] draft={draft_id} LLM error, raw={str(raw)[:1000]}")
            return {"ok": False, "error": f"AI 文案输出解析失败: {exc}"}
        proposal: dict = {"ozon_title": str(body.get("ozon_title") or ""),
                          "description": str(body.get("description") or ""),
                          "attributes": []}   # 文案只出 标题/简介/标签，不提取品牌/类型
        tags_value = clean_hashtags(body.get("hashtags"))
        mapped: list = []
        if tags_value:
            proposal["attributes"].append({"id": 23171, "values": [{"value": tags_value}]})
            mapped.append({"id": 23171, "name": "#Хештеги", "value": tags_value})
        report = {"category_path": "", "category_fallback": False, "brand_warning": None,
                  "unmapped": [], "mapped": mapped, "keywords": []}
        draft_json = build_proposal_draft(proposal, report, [], [], ts=utc_now_iso())
        for k in ("category_id", "type_id", "category_path", "brand_name", "brand_id"):
            (draft_json.get("fields") or {}).pop(k, None)
        # 自动应用模式：直接写入草稿，跳过人工确认
        if bool(self.store.get_settings().get("ai_auto_apply")):
            updated = self.apply_ai_proposal(draft_id)
            return {"ok": True, "mode": "applied", "draft": updated}
        self.store.set_ai_proposal(draft_id, draft_json)
        return {"ok": True, "mode": "draft", "proposal": draft_json, "report": report}

    def ai_image_prompts(self, draft_id: int, n_points: int = 3) -> dict:
        """生成 ChatGPT 出图提示词(主图 + n_points 张卖点图)，不写库。
        同时返回原始图片 URL，供用户手动上传 ChatGPT 当参考图。
        AI 未配置 → deepseek_chat 抛 RuntimeError(路由转 400)。"""
        from webui.ai_card import _SYS_IMG_PROMPTS, build_image_prompt_input, parse_image_prompts  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        # 注意：不能用 `n_points or 3`——n_points=0 是 falsy 会变 3。显式判 None。
        n = max(1, min(int(n_points if n_points is not None else 3), 6))
        settings = self.store.get_settings()
        system = _SYS_IMG_PROMPTS.format(n=n)
        out = self._card_chat(settings, draft)(system, build_image_prompt_input(draft))
        parsed = parse_image_prompts(out, n)
        _sr = draft.get("source_raw") or {}
        detail_images = _sr.get("detail_images") or []
        detail_local = _sr.get("detail_local") or []   # 本地副本(避防盗链)，给缩略图显示
        return {
            "ok": True,
            "main": parsed["main"],
            "selling_points": parsed["selling_points"],
            # 两池化后:采集图在 materials(in_gallery=0),用所有 materials URL 作参考图
            "source_images": [m["url"] for m in (draft.get("materials") or [])
                              if m.get("source") == "collected"]
                             or (draft.get("images") or []),
            "local_images": draft.get("local_images") or [],
            "detail_images": detail_images,    # 源 URL，给 ChatGPT 当参考(复制全部用)
            "detail_local": detail_local,      # 本地副本，缩略图显示用(避防盗链)
        }

    def _resolve_image_input(self, url: str) -> str:
        """把前端选的源图变成 Agnes 能取到的输入：http(s) 直接用；/media/ 本地副本
        外网不可达 → 读字节转 data URI（图生图官方支持 Data URI Base64；图生视频
        官方只写了 URL，data URI 是尽力而为，被拒会以 Agnes 400 浮出来）。"""
        from webui import agnes  # noqa: PLC0415
        u = str(url or "").strip()
        if u.startswith("http"):
            return u
        if u.startswith("/media/"):
            ext = u.rsplit(".", 1)[-1].lower() if "." in u else ""
            if ext in ("mp4", "mov", "webm", "m4v"):
                raise ValueError(f"源图不能是视频文件: {u}")
            data = _media.read_media_bytes(u)
            if not data:
                raise ValueError(f"本地图读取失败: {u}")
            return agnes.to_data_uri(data, u)
        raise ValueError("源图必须是 http(s) URL 或 /media/ 本地图")

    def ai_generate_image(self, draft_id: int, *, mode: str = "text2img", prompt: str = "",
                          source_url: str | None = None, size: str | None = None,
                          as_main: bool = False) -> dict:
        """Agnes 生成商品图：text2img(营销图) / img2img(白底主图等，保构图改场景)。
        生成结果下载到本地 /media/（不依赖 Agnes URL 存活期），挂进 draft.images；
        发布时由 OSS rehost 链路把本地图上传到你的 OSS。as_main=True 插到首位当主图。"""
        from webui import agnes  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        p = str(prompt or "").strip()
        if not p:
            raise ValueError("提示词不能为空")
        m = str(mode or "text2img").strip().lower()
        settings = self.store.get_settings()
        sources = None
        if m == "img2img":
            if not str(source_url or "").strip():
                raise ValueError("图生图需要选择一张源图")
            sources = [self._resolve_image_input(str(source_url))]
        # 出图引擎 admin 可切：ai_image.engine == gptimage → 走 gen_image；默认 agnes
        from webui.settings_migrate import ai_config  # noqa: PLC0415
        imgcfg = ai_config(settings, "image")
        if imgcfg["engine"] == "gptimage":
            from ozon_common.gen_image import create_image, edit_image, images_from_response  # noqa: PLC0415
            gcfg = self._gen_image_cfg(settings)
            tmp = None
            try:
                if str(source_url or "").strip():       # img2img：源图落临时文件给 edit
                    tmp = self._local_image_path(str(source_url))
                    resp = edit_image(gcfg, p, [tmp], size=size or _GEN_SIZE, output_format="png")
                else:                                    # text2img
                    resp = create_image(gcfg, p, size=size or _GEN_SIZE, output_format="png")
            finally:
                if tmp and os.path.exists(tmp):
                    os.remove(tmp)
            picked = images_from_response(resp)
            if not picked:
                raise RuntimeError("gptimage 未返回图片")
            data = picked[0]
            remote_url = ""
            fname = "gptimage-ai.png"
        else:
            remote_url = agnes.generate_image(settings, p, size=size or _GEN_SIZE,
                                              source_images=sources)
            data = _download_bytes(remote_url, timeout=120)
            ext = remote_url.rsplit("?", 1)[0].rsplit(".", 1)[-1].lower()
            fname = f"agnes-ai.{ext if ext in ('png', 'jpg', 'jpeg', 'webp') else 'png'}"
        if len(data) > 20 * 1024 * 1024:   # 与上传路由同口径的 20MB 上限
            raise RuntimeError(f"生成图过大({len(data) // 1024 // 1024}MB > 20MB)")
        local = _media.save_upload(f"draft-{draft_id}", fname, data)
        # 生图可能阻塞数分钟——写前重读草稿，别用进入时的旧快照盖掉期间的用户编辑/并发生成
        cur = self.store.get_draft(draft_id)
        if cur is None:
            raise KeyError(f"draft {draft_id} not found")
        images = list(cur.get("images") or [])
        patch: dict = {"images": images}
        if as_main:
            images.insert(0, local)
            # images↔local_images 是按下标配对的平行数组（前端 localMap），头插必须同步补位
            locs = list(cur.get("local_images") or [])
            if locs:
                patch["local_images"] = ["", *locs]
        else:
            images.append(local)
        updated = self.store.update_draft(draft_id, patch)
        return {"ok": True, "draft": updated, "image": local, "remote_url": remote_url}

    def _image_bytes(self, src: str) -> bytes:
        """读一张草稿图的字节：http(s)→下载；/media/→本地副本。"""
        u = str(src or "").strip()
        if u.startswith("http"):
            return _download_bytes(u)
        if u.startswith("/media/"):
            data = _media.read_media_bytes(u)
            if not data:
                raise ValueError(f"本地图读取失败: {u}")
            return data
        raise ValueError("图片必须是 http(s) URL 或 /media/ 本地图")

    def make_infographic(self, draft_id: int, *, source_index: int = 0, heading: str = "",
                         bullets: list | None = None, watermark: str = "") -> dict:
        """把草稿第 source_index 张图做成俄语信息图（底部面板 + 标题 + 要点），可选叠店铺水印，
        存 /media/ 并追加到 draft.images。标题缺省取 ozon_title。引擎无关、纯 PIL（Montserrat）。"""
        from webui.image_compose import add_watermark, compose_infographic  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        imgs = list(draft.get("images") or [])
        if not imgs:
            raise ValueError("草稿没有图片，无法生成信息图（先采集或生成主图）")
        idx = max(0, min(int(source_index), len(imgs) - 1))
        data = self._image_bytes(str(imgs[idx]))
        h = str(heading or draft.get("ozon_title") or draft.get("source_title") or "")
        bl = [str(b) for b in (bullets or []) if str(b).strip()]
        ig = compose_infographic(data, heading=h, bullets=bl, font_path=_FONT_PATH)
        wm = str(watermark or "").strip()
        if wm:
            ig = add_watermark(ig, wm, font_path=_FONT_PATH)
        local = _media.save_upload(f"draft-{draft_id}", "infographic.jpg", ig)
        # 生成可能耗时——写前重读草稿，别盖掉期间编辑
        cur = self.store.get_draft(draft_id)
        if cur is None:
            raise KeyError(f"draft {draft_id} not found")
        images = list(cur.get("images") or [])
        images.append(local)
        updated = self.store.update_draft(draft_id, {"images": images})
        return {"ok": True, "draft": updated, "image": local}

    def _local_image_path(self, src: str) -> str:
        """把一张图(http 或 /media/)落成本地临时文件，返回路径（供 gen_image 的 edit 用）。调用方负责删。"""
        import tempfile  # noqa: PLC0415
        data = self._image_bytes(src)
        fd, path = tempfile.mkstemp(suffix=".png", prefix="ozsrc-")
        try:
            os.write(fd, data)
        finally:
            os.close(fd)
        return path

    def try_copy(self, draft_id: int) -> dict:
        """Ozon 来源草稿：试官方复制(import-by-sku)。可复制→在目标店建复制卡并标记草稿；
        不可复制→返回 copyable=False（前端据此转「原创建卡」分支）。会在你店里真建一张卡。"""
        import re  # noqa: PLC0415

        from webui.listing_build import random_offer_id  # noqa: PLC0415
        from webui.ozon_client_adapter import copy_by_sku  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        # SKU：显式 sku → 商品链接末尾数字（不用 ozon_product_id，那是 Seller API 的另一种 id）
        sku = str(draft.get("sku") or "").strip()
        if not sku:
            m = re.search(r"/product/[^/?#]*?(\d+)/?(?:[?#]|$)", str(draft.get("source_url") or ""))
            sku = m.group(1) if m else ""
        if not sku.isdigit():
            raise ValueError("草稿没有可用的 Ozon SKU（只有带 Ozon 商品链接的草稿能走官方复制）")
        settings = self._settings_for_store(draft.get("store_client_id"))
        offer_id = random_offer_id()
        verdict = copy_by_sku(
            settings, sku=int(sku), offer_id=offer_id,
            price=str(draft.get("price") or "") or None,
            old_price=str(draft.get("old_price") or "") or None,
            currency_code="RUB",
            name=str(draft.get("ozon_title") or draft.get("source_title") or "")[:200] or None)
        if verdict["copyable"]:
            self.store.update_draft(draft_id, {
                "offer_id": offer_id,
                "status": "published" if verdict["status"] == "created" else "draft",
                "publish_response": {"copy": verdict},
            })
        return {"copyable": verdict["copyable"], "status": verdict["status"],
                "offer_id": offer_id if verdict["copyable"] else None,
                "task_id": verdict.get("task_id")}

    def make_rich_content(self, draft_id: int, *, image_indexes: list | None = None) -> dict:
        """把草稿图拼成 Ozon 富文本(billboard 大图序列)，存 draft.source_raw.rich_content_json。
        默认跳过主图(索引0)、其余全进；俄语文字烤在图里(先用「生成俄语信息图」做几张再来这步)。
        发布时 to_ozon_import_item 自动塞属性 11254、rewrite_item_media 把图链换 OSS 直链。"""
        from webui.drafts import loads_json  # noqa: PLC0415
        from webui.listing_build import build_rich_content  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        imgs = list(draft.get("images") or [])
        if not imgs:
            raise ValueError("草稿没有图片，无法生成富文本（先采集/生成图片）")
        if image_indexes:
            sel = [imgs[i] for i in image_indexes if isinstance(i, int) and 0 <= i < len(imgs)]
        else:
            sel = imgs[1:] if len(imgs) > 1 else imgs   # 默认跳过主图，其余进富文本
        if not sel:
            raise ValueError("没有可用于富文本的图片")
        rc = build_rich_content(sel)
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = dict(sr or {})
        sr["rich_content_json"] = rc
        updated = self.store.update_draft(draft_id, {"source_raw": sr})
        return {"ok": True, "draft": updated, "blocks": len(rc["content"])}

    def _multimodal_chat(self, settings: dict):
        """理解层多模态 chat。优先用单独配的「多模态 AI」槽;**没配则自动复用「文本 AI」**
        ——Agnes 本就多模态,翻译那套配置直接拿来看图理解,用户无需再配一个模型。
        返回 chat(system, user, image_urls) -> str。"""
        from webui.settings_migrate import ai_config  # noqa: PLC0415
        kind = "multimodal" if ai_config(settings, "multimodal").get("key") else "text"
        engine = ai_config(settings, kind)["engine"]
        if engine == "agnes":
            from webui import agnes  # noqa: PLC0415
            return lambda s, u, images: agnes.agnes_chat(settings, s, u, images=images, kind=kind)
        from webui.ai_card import deepseek_chat  # noqa: PLC0415
        return lambda s, u, images: deepseek_chat(settings, s, u, images=images, kind=kind)

    def understand_draft(self, draft_id: int, *, force: bool = False) -> dict:
        """理解层:多模态"看图理解"→ understanding,缓存 source_raw.understanding。
        force=True 重抽;已有缓存且非 force → 直接返回(cached=True)。图先经 _resolve_image_input
        转成模型可取链接(http 直用 / /media/ 转 data URI)。"""
        from webui.drafts import loads_json  # noqa: PLC0415
        from webui.understand import understand  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = dict(sr or {})
        cached = isinstance(sr.get("understanding"), dict) and bool(sr["understanding"])
        if force or not cached:
            chat_fn = self._multimodal_chat(self.store.get_settings())
            u = understand(draft, chat_fn, resolve_image=self._resolve_image_input)
            sr["understanding"] = u
        else:
            u = sr["understanding"]
        # 把理解到的尺寸/克重 + 默认定价写进结构化草稿字段（即使缓存命中也回填，便于补旧草稿）
        patch = self._autofill_from_understanding(draft, u)
        patch["source_raw"] = sr
        updated = self.store.update_draft(draft_id, patch)
        autofill = {k: v for k, v in patch.items() if k != "source_raw"}
        return {"ok": True, "understanding": u, "cached": cached and not force,
                "draft": updated, "autofill": autofill}

    def _cost_basis_cny(self, draft: dict) -> float | None:
        """进价(CNY)：cost_cny>0 优先；否则取 1688 采集的 price_display 区间最低价。"""
        try:
            c = float(draft.get("cost_cny") or 0)
            if c > 0:
                return c
        except (TypeError, ValueError):
            pass
        import re  # noqa: PLC0415

        from webui.drafts import loads_json  # noqa: PLC0415
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = sr if isinstance(sr, dict) else {}
        nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", str(sr.get("price_display") or ""))]
        return min(nums) if nums else None

    def _autofill_from_understanding(self, draft: dict, understanding: dict) -> dict:
        """理解 → 结构化字段补丁：尺寸/克重以**采集的包装件重尺为准**，理解(看图)的值仅在草稿没有时兜底
        (看图读到的多是产品尺寸而非物流箱尺寸，Ozon 要的是包装尺寸)；默认定价仅当售价为空时填。"""
        from webui.listing_build import default_pricing, parse_dims_mm, parse_weight_g  # noqa: PLC0415
        specs = (understanding or {}).get("specs") or {}
        patch: dict = {}

        def _empty(v):
            return v in (None, 0, "", "0")
        dims = parse_dims_mm(specs.get("尺寸") or specs.get("size") or "")
        if dims and all(_empty(draft.get(k)) for k in ("length_mm", "width_mm", "height_mm")):
            patch["length_mm"], patch["width_mm"], patch["height_mm"] = dims
        w = parse_weight_g(specs.get("重量") or specs.get("weight") or "")
        if w and _empty(draft.get("weight_g")):
            patch["weight_g"] = w
        if not str(draft.get("price") or "").strip():   # 不覆盖用户已填的售价
            cost = self._cost_basis_cny(draft)
            pr = default_pricing(cost) if cost else None
            if pr:
                patch["price"], patch["old_price"] = str(pr[0]), str(pr[1])
                try:
                    has_cost = float(draft.get("cost_cny") or 0) > 0
                except (TypeError, ValueError):
                    has_cost = False
                if not has_cost:
                    patch["cost_cny"] = cost   # 进价回填，便于用户看到/二次定价稳定
        return patch

    def recommend(self, draft_id: int) -> dict:
        """智能推荐(纯逻辑,不调 AI):据来源 + 缓存的 understanding → 推荐路径(复制/俄化/重做)
        + 逐图默认处理。understanding 未生成时 per_image 为空、仅按来源给路径默认。"""
        from webui.drafts import loads_json  # noqa: PLC0415
        from webui.recommend import recommend_path  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = sr if isinstance(sr, dict) else {}
        understanding = sr.get("understanding") if isinstance(sr.get("understanding"), dict) else None
        source = draft.get("source") or draft.get("source_platform") or ""
        rec = recommend_path(source=source, understanding=understanding, copyable=None)
        return {"ok": True, "recommendation": rec, "has_understanding": understanding is not None}

    def _add_candidate(self, draft_id: int, data: bytes, label: str, *, slot: str = "") -> str:
        """生成的图直接 INSERT draft_images 表行；记 image_types + slot_images 进 source_raw。
        返回本地 URL。方法名沿用以减少调用点改动。"""
        if len(data) > 20 * 1024 * 1024:
            raise RuntimeError("生成图过大(>20MB)")
        local = _media.save_upload(f"draft-{draft_id}-gen", "gen.png", data)
        img_type = _img_type_from_label(label)
        with self._cand_lock:
            self.store.add_draft_image(draft_id, local, type=img_type, source="generated")
            draft = self.store.get_draft(draft_id)
            if draft is None:
                raise KeyError(f"draft {draft_id} not found")
            sr = dict(draft.get("source_raw") or {})
            types = dict(sr.get("image_types") or {})
            types[local] = img_type
            sr["image_types"] = types
            if slot:
                slot_imgs = dict(sr.get("slot_images") or {})
                slot_imgs[str(slot)] = local
                sr["slot_images"] = slot_imgs
            self.store.update_draft(draft_id, {"source_raw": sr, "status": draft.get("status")})
        return local

    def _gen_image_cfg(self, settings: dict | None = None):
        """按 DB 的 ai_image 槽（接口地址/Key/模型）构造 GenImageConfig；字段留空回退 GPTPLUS5_* 环境变量。
        俄化/重做/AI生图统一走它 → 生图 AI 配置可全部存数据库。"""
        from ozon_common.gen_image import GenImageConfig  # noqa: PLC0415
        from webui.settings_migrate import ai_config  # noqa: PLC0415
        s = settings if settings is not None else self.store.get_settings()
        c = ai_config(s, "image")
        return GenImageConfig(api_key=(c.get("key") or None),
                              base_url=(c.get("base") or None),
                              model=(c.get("model") or None))

    def _edit_source_image(self, draft_id: int, source_index: int, prompt: str) -> bytes:
        """对草稿第 source_index 张图做 gpt-image edit(传源图保产品一致),返回结果字节。"""
        from ozon_common.gen_image import edit_image, images_from_response  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        imgs = list(draft.get("images") or [])
        if not imgs:
            raise ValueError("草稿没有图片（先采集/生成）")
        idx = max(0, min(int(source_index), len(imgs) - 1))
        tmp = self._local_image_path(str(imgs[idx]))
        try:
            resp = edit_image(self._gen_image_cfg(), prompt, [tmp], size=_GEN_SIZE, output_format="png")
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)
        picked = images_from_response(resp)
        if not picked:
            raise RuntimeError("出图未返回图片")
        return picked[0]

    def localize_image(self, draft_id: int, source_index: int = 0) -> dict:
        """单张俄化:把第 source_index 张图上的中文换成俄语(gpt-image edit,保图不变),结果进候选区。"""
        from ozon_common.gen_image import LOCALIZE_PROMPT  # noqa: PLC0415
        data = self._edit_source_image(draft_id, source_index, LOCALIZE_PROMPT)
        url = self._add_candidate(draft_id, data, f"俄化#{source_index}")
        return {"ok": True, "candidate": url, "draft": self.store.get_draft(draft_id)}

    def regen_image(self, draft_id: int, source_index: int = 0, *, role: str = "",
                    heading: str = "", bullets: list | None = None) -> dict:
        """单张重做:按角色 + 俄语文字 重新生成(gpt-image edit 源图),结果进候选区。数字需 QC。"""
        from ozon_common.gen_image import build_infographic_prompt  # noqa: PLC0415
        prompt = build_infographic_prompt(role=role, heading=heading, bullets=bullets)
        data = self._edit_source_image(draft_id, source_index, prompt)
        url = self._add_candidate(draft_id, data, f"重做#{source_index}")
        return {"ok": True, "candidate": url, "draft": self.store.get_draft(draft_id)}

    def whiten_main(self, draft_id: int, source_index: int = 0) -> dict:
        """选第 source_index 张图做白底电商主图(gpt-image edit + 白底提示词),结果进候选区。"""
        from ozon_common.gen_image import WHITE_MAIN_PROMPT  # noqa: PLC0415
        data = self._edit_source_image(draft_id, source_index, WHITE_MAIN_PROMPT)
        url = self._add_candidate(draft_id, data, f"白底主图#{source_index}")
        return {"ok": True, "candidate": url, "draft": self.store.get_draft(draft_id)}

    def scene_image(self, draft_id: int, source_index: int = 0, *, hint: str = "") -> dict:
        """选第 source_index 张图做场景/氛围图(保产品一致 + 放进使用场景),可带场景提示。结果进候选区。"""
        from ozon_common.gen_image import SCENE_PROMPT  # noqa: PLC0415
        prompt = SCENE_PROMPT + (f" Scene hint: {str(hint).strip()}" if str(hint or "").strip() else "")
        data = self._edit_source_image(draft_id, source_index, prompt)
        url = self._add_candidate(draft_id, data, f"场景图#{source_index}")
        return {"ok": True, "candidate": url, "draft": self.store.get_draft(draft_id)}

    def _load_image_plan(self, draft_id: int, *, force: bool = False):
        """取/建图集计划(缓存 source_raw.image_plan)。返回 (draft, sr, plan)。"""
        from ozon_common.image_plan import build_image_plan  # noqa: PLC0415
        from webui.drafts import loads_json  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = dict(sr or {})
        plan = sr.get("image_plan")
        if force or not isinstance(plan, list) or not plan:
            understanding = sr.get("understanding") if isinstance(sr.get("understanding"), dict) else None
            plan = build_image_plan(understanding, draft.get("images"))
            sr["image_plan"] = plan
            self.store.update_draft(draft_id, {"source_raw": sr})
        return draft, sr, plan

    def image_plan(self, draft_id: int, *, force: bool = False) -> dict:
        """图集计划 + 每槽状态(todo 待做 / applied 已生成)。生成图直接进 images、不再有候选态；
        据 slot_images(槽→已生成图url) 是否仍在 draft.images 判断该槽是否已出图。"""
        draft, sr, plan = self._load_image_plan(draft_id, force=force)
        images = {str(u) for u in (draft.get("images") or [])}
        slot_imgs = sr.get("slot_images") or {}
        out = []
        for s in plan:
            u = str(slot_imgs.get(str(s.get("slot_id"))) or "")
            done = bool(u and u in images)   # 用户删了该图 → 回到 todo，可重出
            out.append({**s, "status": "applied" if done else "todo",
                        "candidate_url": u if done else ""})
        return {"ok": True, "plan": out}

    def design_image_plan(self, draft_id: int, *, target: int = 10) -> dict:
        """**AI 设计图集**：把看图理解 + 源图清单(角色) 喂给 LLM 当"美术总监"，设计 ~target 张
        符合 Ozon 的商品图方案(白底主图/细节俄化/场景/卖点·规格信息图)，产出与规则版同形状的槽位，
        覆盖写入 source_raw.image_plan(后续渲染/界面照旧用)。LLM 失败/空 → 回退规则版 build_image_plan。"""
        import json as _json  # noqa: PLC0415

        from ozon_common.image_plan import build_image_plan  # noqa: PLC0415
        from webui.ai_card import _SYS_IMG_PLAN, _extract_json, build_profile  # noqa: PLC0415
        from webui.drafts import loads_json  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = draft.get("source_raw")
        if isinstance(sr, str):
            sr = loads_json(sr, {})
        sr = dict(sr or {})
        images = list(draft.get("images") or [])
        if not images:
            raise ValueError("草稿没有图片，无法设计图集（先采集/生成图片）")
        und = sr.get("understanding") if isinstance(sr.get("understanding"), dict) else {}
        if not und:   # 设计强依赖看图理解(图片角色/卖点/规格)——没有就先自动跑一遍
            self.understand_draft(draft_id)
            draft = self.store.get_draft(draft_id)
            sr = draft.get("source_raw")
            if isinstance(sr, str):
                sr = loads_json(sr, {})
            sr = dict(sr or {})
            und = sr.get("understanding") if isinstance(sr.get("understanding"), dict) else {}
        n = len(images)
        roles = {im["idx"]: str(im.get("role") or "") for im in (und.get("images") or [])
                 if isinstance(im, dict) and isinstance(im.get("idx"), int) and 0 <= im["idx"] < n}
        inventory = [{"idx": i, "role": roles.get(i, "")} for i in range(n)]
        profile = build_profile(sr, understanding=und)
        user = (f"Target image count: {int(target)}\nSource photos (use source_idx from this inventory):\n"
                + _json.dumps(inventory, ensure_ascii=False) + "\n\nProduct understanding:\n" + profile)
        chat = self._card_chat(self.store.get_settings(), draft)
        valid: list[dict] = []
        try:
            out = _extract_json(chat(_SYS_IMG_PLAN, user))
            seen: set = set()
            for i, s in enumerate(out.get("slots") or []):
                action = str(s.get("action") or "").strip().lower()
                if action not in ("white", "localize", "scene", "infographic"):
                    continue
                try:
                    si = int(s.get("source_idx"))
                except (TypeError, ValueError):
                    si = 0
                si = si if 0 <= si < n else 0
                sid = str(s.get("slot_id") or "").strip() or f"s{i}"
                if sid in seen:
                    sid = f"{sid}_{i}"
                seen.add(sid)
                valid.append({"slot_id": sid, "role": str(s.get("role") or ""),
                              "label": str(s.get("label") or sid), "action": action, "source_idx": si,
                              "heading": str(s.get("heading") or ""),
                              "bullets": [str(b) for b in (s.get("bullets") or []) if str(b).strip()],
                              "scene_hint": str(s.get("scene_hint") or ""),
                              "prompt": str(s.get("prompt") or "").strip()})
        except Exception as exc:  # noqa: BLE001  设计失败 → 回退规则版，不阻断
            valid = []
        if not valid:
            valid = build_image_plan(und, images)
            fallback = True
        else:
            fallback = False
        sr["image_plan"] = valid
        self.store.update_draft(draft_id, {"source_raw": sr})
        return {"ok": True, "plan": valid, "count": len(valid), "fallback": fallback}

    def generate_plan_slot(self, draft_id: int, slot_id: str) -> dict:
        """生成图集计划某个槽位的图(按 action 选生成器,用槽的 source_idx 为源)→ 进候选区,标 slot。"""
        from ozon_common.gen_image import (  # noqa: PLC0415
            LOCALIZE_PROMPT,
            NON_PRODUCT_RULE,
            OZON_RU_RULE,
            SCENE_PROMPT,
            WHITE_MAIN_PROMPT,
            build_infographic_prompt,
        )
        _, _, plan = self._load_image_plan(draft_id)
        slot = next((s for s in plan if s.get("slot_id") == slot_id), None)
        if slot is None:
            raise ValueError(f"图集计划里没有槽位 {slot_id}")
        src = int(slot.get("source_idx") or 0)
        action = slot.get("action")
        designed = str(slot.get("prompt") or "").strip()
        if designed:   # 设计模型为这张图输出的整段提示词 → 直接用，强制追加去非产品+Ozon俄语硬规则(双保险)
            prompt = designed + " " + NON_PRODUCT_RULE + OZON_RU_RULE
        elif action == "white":
            prompt = WHITE_MAIN_PROMPT
        elif action == "localize":
            prompt = LOCALIZE_PROMPT
        elif action == "scene":
            hint = str(slot.get("scene_hint") or slot.get("heading") or "").strip()
            prompt = SCENE_PROMPT + (f" Scene context: {hint}" if hint else "")
        elif action == "infographic":
            prompt = build_infographic_prompt(role=str(slot.get("role") or ""),
                                              heading=str(slot.get("heading") or ""),
                                              bullets=slot.get("bullets") or [])
        else:
            raise ValueError(f"未知 action {action}")
        data = self._edit_source_image(draft_id, src, prompt)
        url = self._add_candidate(draft_id, data, str(slot.get("label") or slot_id), slot=slot_id)
        return {"ok": True, "slot_id": slot_id, "candidate": url, "draft": self.store.get_draft(draft_id)}

    def start_ai_video(self, draft_id: int, *, prompt: str | None = None,
                       image_url: str | None = None) -> dict:
        """启动 Agnes 图生视频后台任务（全局单任务，约 5 秒 121帧@24fps）。
        默认用草稿主图 + 商品展示运镜提示词；完成后 _on_ai_video_done 回写草稿。"""
        from webui import agnes  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        settings = self.store.get_settings()
        agnes._conf(settings)   # 没配 key 在启动前就报错（不进后台线程才能让前端看到 400）
        img_in = str(image_url or "").strip()
        if not img_in:
            imgs = draft.get("images") or []
            if not imgs:
                raise ValueError("草稿没有图片，无法图生视频（先采集或上传图片）")
            img_in = str(imgs[0])
        image = self._resolve_image_input(img_in)
        title = str(draft.get("ozon_title") or draft.get("source_title") or "").strip()
        p = str(prompt or "").strip() or (
            "Smooth cinematic product showcase video, slow camera orbit around the product, "
            "soft studio lighting, the product stays centered and visually identical to the "
            f"reference image. Product: {title[:200]}"
        )
        create_fn = lambda: agnes.create_video_task(settings, p, image=image)  # noqa: E731
        query_fn = lambda vid: agnes.query_video(settings, vid)  # noqa: E731
        return ai_video.start_video(create_fn, query_fn, self._on_ai_video_done, draft_id)

    def _on_ai_video_done(self, draft_id: int, url: str) -> None:
        """视频任务完成回调（后台线程）：写 video_url + 下载本地预览副本。
        - 本地副本放 draft-{id}-ai key（避免覆盖采集来的源视频），overwrite=True 防重生成命中旧缓存
        - 下载完再重读草稿（下载可能数十秒，别用旧快照盖掉期间的编辑）
        - 下载失败把 video_local 置空：宁可前端回退播 video_url，也别播上一版的旧副本
        - 显式保留 status：这是后台异步写，不该把已发布草稿悄悄打回 ready/invalid"""
        try:
            vloc = _media.download_video(url, f"draft-{draft_id}-ai", overwrite=True)
        except Exception:  # noqa: BLE001  本地副本失败不影响 video_url 落库
            vloc = ""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            return
        patch: dict = {
            "video_url": url,
            "source_raw": {**(draft.get("source_raw") or {}), "video_local": vloc},
        }
        if draft.get("status"):
            patch["status"] = draft["status"]
        self.store.update_draft(draft_id, patch)

    def ai_video_status(self) -> dict:
        return ai_video.video_status()

    def stop_ai_video(self) -> dict:
        return ai_video.request_stop()

    def start_image_batch(self, draft_id: int, *, source_url: str | None = None) -> dict:
        """启动 Agnes 整套商品图后台任务（12 角度·全图生图·候选区）。
        参考图优先用本地副本（避 1688 防盗链 Agnes 拉不到）；生成结果进候选区不进正式图集。"""
        from webui import agnes  # noqa: PLC0415
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        settings = self.store.get_settings()
        agnes._conf(settings)   # 没配 key 启动前就报错（不进后台线程，前端能看到 400）
        ref_in = str(source_url or "").strip()
        if not ref_in:
            locs = draft.get("local_images") or []
            imgs = draft.get("images") or []
            ref_in = str(locs[0]) if locs else (str(imgs[0]) if imgs else "")
        if not ref_in:
            raise ValueError("草稿没有图片，无法整套图生图（先采集或上传图片）")
        ref = self._resolve_image_input(ref_in)
        title = str(draft.get("ozon_title") or draft.get("source_title") or "")
        plan = agnes.plan_image_angles(title)
        # 开新批先清空旧候选（避免多次生成累积）
        sr = dict(draft.get("source_raw") or {})
        sr["ai_image_candidates"] = []
        self.store.update_draft(draft_id, {"source_raw": sr, "status": draft.get("status")})
        gen_fn = lambda prompt: agnes.generate_image(settings, prompt, source_images=[ref])  # noqa: E731
        return ai_image_batch.start_batch(gen_fn, self._on_image_candidate, plan, draft_id)

    def _on_image_candidate(self, draft_id: int, angle: str, url: str) -> None:
        """单张候选完成回调（后台线程，串行）：下载本地 + 追加 source_raw.ai_image_candidates。
        Agnes 图 URL 可能过期 → 下载到 /media（候选 key 与正式图分开）。失败抛出由批量层记 failed。"""
        data = _download_bytes(url, timeout=120)
        if len(data) > 20 * 1024 * 1024:
            raise RuntimeError("候选图过大(>20MB)")
        local = _media.save_upload(f"draft-{draft_id}-cand", "cand.png", data)
        draft = self.store.get_draft(draft_id)
        if draft is None:
            return
        sr = dict(draft.get("source_raw") or {})
        cands = list(sr.get("ai_image_candidates") or [])
        cands.append({"url": local, "angle": angle})
        sr["ai_image_candidates"] = cands
        self.store.update_draft(draft_id, {"source_raw": sr, "status": draft.get("status")})

    def image_batch_status(self) -> dict:
        return ai_image_batch.batch_status()

    def stop_image_batch(self) -> dict:
        return ai_image_batch.request_stop()

    def apply_image_candidates(self, draft_id: int, indices: list[int] | None = None) -> dict:
        """把候选区的图加入正式图集 draft.images，清空候选区。
        indices 为空 → 应用全部候选(前端"全部应用"/一键自动都是全应用，不靠前端传索引，
        避免前端响应式滞后导致传空 → 静默不应用)。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = dict(draft.get("source_raw") or {})
        cands = list(sr.get("ai_image_candidates") or [])
        if not cands:
            raise ValueError("没有候选图可应用")
        if indices:
            picked = [cands[i] for i in indices
                      if isinstance(i, int) and 0 <= i < len(cands)]
        else:
            picked = list(cands)   # 不传索引 = 全部应用
        sel = [str(c.get("url") or "") for c in picked if str(c.get("url") or "")]
        if not sel:
            raise ValueError("未选择任何候选图")
        # 记录每张应用图的类型(白底/细节/场景/尺寸/卖点…)，供图集分类/排序/标签
        types = dict(sr.get("image_types") or {})
        for c in picked:
            u = str(c.get("url") or "")
            if u:
                types[u] = _img_type_from_label(c.get("angle") or c.get("slot") or "")
        sr["image_types"] = types
        images = list(draft.get("images") or []) + sel
        sr["ai_image_candidates"] = []   # 应用后清空候选区
        updated = self.store.update_draft(draft_id, {"images": images, "source_raw": sr})
        return {"ok": True, "draft": updated, "added": len(sel)}

    def discard_image_candidates(self, draft_id: int) -> dict:
        """清空候选区（全部丢弃）。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        sr = dict(draft.get("source_raw") or {})
        sr["ai_image_candidates"] = []
        updated = self.store.update_draft(draft_id, {"source_raw": sr, "status": draft.get("status")})
        return {"ok": True, "draft": updated}

    # ---------- 插件桥接（/api/ext/*）----------
    def ext_ping(self) -> dict:
        # 带上 rub_cny 汇率：插件采集 Ozon 卢布价后据此换算成人民币再回传（后端统一 CNY）
        rate = self.store.get_settings().get("rub_cny")
        try:
            rate = float(rate) if rate not in (None, "") else None
        except (TypeError, ValueError):
            rate = None
        return {"ok": True, "name": "ozon-listing-webui", "version": "1", "rub_cny": rate}

    def _ext_cache_video(self, draft_id: int, video_url: str) -> None:
        """为草稿下载视频本地副本（同 collect 流程）：Ozon/1688 CDN 的 <video> 跨域会卡死，
        落 /media/ 同源才能播。best-effort，失败不影响采集。"""
        vurl = str(video_url or "")
        if not vurl.startswith("http"):
            return
        cur = self.store.get_draft(draft_id)
        sr = cur.get("source_raw") if cur else None
        if isinstance(sr, str):
            from webui.drafts import loads_json  # noqa: PLC0415
            sr = loads_json(sr, {})
        sr = sr if isinstance(sr, dict) else {}
        if sr.get("video_local"):
            return
        try:
            from webui.media import download_video  # noqa: PLC0415
            vloc = download_video(vurl, f"draft-{draft_id}")
            if vloc:
                self.store.update_draft(draft_id, {"source_raw": {**sr, "video_local": vloc}})
        except Exception:  # noqa: BLE001
            pass

    def ext_collect_parsed(self, payload: dict) -> dict:
        url = str(payload.get("url") or "").strip()
        if not url:
            raise ValueError("url required")
        data = payload.get("data") or {}
        _has_media = bool((data.get("images") or [])) or bool(str(data.get("video_url") or "").strip())
        from webui.drafts import create_draft_from_url  # noqa: PLC0415
        scraped = {
            "source_title": data.get("title"),
            "ozon_title": data.get("title"),  # Ozon 竞品标题已是俄语，直接用
            "description": data.get("description"),
            "price": data.get("price"),
            "old_price": data.get("old_price"),
            "images": data.get("images") or [],
            "video_url": data.get("video_url"),
            "weight_g": data.get("weight_g"),
            "length_mm": data.get("length_mm"),
            "width_mm": data.get("width_mm"),
            "height_mm": data.get("height_mm"),
            "category_path": data.get("category_path"),
        }
        # 采集的属性(名值对 {name,value}) → draft.attributes，供 collected_chars → auto-map/AI；
        # 发布时 to_ozon_import_item 只发 {id,values}，名值对自动丢弃（不会误发给 Ozon）。
        collected_attrs = [
            {"name": str(a["name"]).strip(), "value": str(a.get("value") or "").strip()}
            for a in (data.get("attributes") or [])
            if isinstance(a, dict) and a.get("name") and a.get("value")
        ]
        # 主题标签(webHashtags) → 属性 23171「#Хештеги」，多标签空格连接成单值（与 Ozon 卡片惯例一致）
        tags = [str(t).strip() for t in (data.get("hashtags") or []) if str(t).strip()]
        tag_attr = (
            {"id": 23171, "name": "#Хештеги", "values": [{"value": " ".join(tags)}]}
            if tags else None
        )
        attrs_list = list(collected_attrs)
        if tag_attr:
            attrs_list.append(tag_attr)
        if attrs_list:
            scraped["attributes"] = attrs_list
        # 按买家面包屑路径自动匹配卖家类目 → 填 category_id/type_id（没配 key/无匹配则跳过，回退编辑器手选）
        try:
            self._auto_match_category(scraped)   # 类目匹配走本地缓存，快(~0.06s)，保留
        except Exception:  # noqa: BLE001
            pass
        platform = str(data.get("source_platform") or "ozon").strip() or "ozon"
        # 1688 采集价=「进价」而非售价 → 进 cost_cny，按默认倍率算售价(进价×1.3×2)/划线价(售价×1.8)
        if platform == "1688":
            from webui.listing_build import default_pricing  # noqa: PLC0415
            try:
                cost = float(str(scraped.get("price") or "").strip() or 0)
            except (TypeError, ValueError):
                cost = 0.0
            if cost > 0:
                scraped["cost_cny"] = cost
                pr = default_pricing(cost)
                if pr:
                    scraped["price"], scraped["old_price"] = str(pr[0]), str(pr[1])
        new_draft = create_draft_from_url(url, source_platform=platform, scraped=scraped)
        # 草稿绑定店：插件带 store_client_id=当前店；缺省回退默认店(settings.ozon_client_id)，单店用户零感知
        store_cid = str(payload.get("store_client_id") or data.get("store_client_id") or "").strip()
        if not store_cid:
            store_cid = str((self.store.get_settings() or {}).get("ozon_client_id") or "")
        new_draft["store_client_id"] = store_cid
        det = data.get("detail_images") or []
        rcj = data.get("rich_content_json") or None
        vars_ = data.get("variants") or []
        vg = data.get("variant_group") or ""
        sa = data.get("selected_aspects") or []
        sr_new: dict = {}
        # 插件可直接带 source_raw（如 WB 的 options/brand_name，喂 auto-map/AI）
        incoming_sr = data.get("source_raw")
        if isinstance(incoming_sr, dict):
            sr_new.update(incoming_sr)
        if det:
            sr_new["detail_images"] = det
        if rcj:
            sr_new["rich_content_json"] = rcj  # 原始富文本(A+)，发布时复刻
        if vars_:
            sr_new["variants"] = vars_
        if vg:
            sr_new["variant_group"] = vg
        if sa:
            sr_new["selected_aspects"] = sa
        if sr_new:
            new_draft["source_raw"] = sr_new
        # 重复采集去重按店：同来源在不同店各存一份，刷新只补空当前店那一份。
        # 注意按完整 source_url 去重——多变体已由 normalize 保留 #sku 片段而各自唯一(各建一张草稿)，
        # 不能按 source_offer_id 兜底去重(那会把同商品的 N 个变体误并成一张)。
        existing = self.store.find_by_source_url(new_draft["source_url"], None, store_cid)
        if existing:
            # 重复采集：用新解析的源字段「补空」（不覆盖用户已编辑/已选的非空字段），
            # 这样旧的、缺描述/克重/尺寸的草稿能被重新采集刷新。
            patch = {}
            for k in ("ozon_title", "description", "price", "old_price", "images",
                      "video_url", "weight_g", "length_mm", "width_mm", "height_mm",
                      "category_id", "type_id"):
                cur = existing.get(k)
                is_empty = cur in (None, "", 0) or (isinstance(cur, list) and not cur)
                val = new_draft.get(k)
                has_val = val not in (None, "", 0) and not (isinstance(val, list) and not val)
                if is_empty and has_val:
                    patch[k] = val
            if sr_new:
                ex_sr = existing.get("source_raw")
                if not isinstance(ex_sr, dict):
                    ex_sr = {}
                merged = dict(ex_sr)
                for kk, vv in sr_new.items():
                    if not merged.get(kk):  # 旧草稿没有才补
                        merged[kk] = vv
                if merged != ex_sr:
                    patch["source_raw"] = merged
            # 属性/标签补空（不覆盖用户已填）：旧草稿没有采集名值对则补；没有 attr 23171 则补标签
            ex_attrs = existing.get("attributes")
            ex_attrs = ex_attrs if isinstance(ex_attrs, list) else []
            merged_attrs = list(ex_attrs)
            changed_attrs = False
            has_collected = any(a.get("name") and a.get("value") and "values" not in a for a in ex_attrs)
            if collected_attrs and not has_collected:
                merged_attrs = collected_attrs + merged_attrs
                changed_attrs = True
            if tag_attr and not any(_to_int(a.get("id")) == 23171 for a in ex_attrs):
                merged_attrs = merged_attrs + [tag_attr]
                changed_attrs = True
            if changed_attrs:
                patch["attributes"] = merged_attrs
            if patch:
                self.store.update_draft(existing["id"], patch)
            if _has_media and self._media_needs_upload(self.store.get_draft(existing["id"]) or existing):
                self.store.set_media_status(existing["id"], "pending")  # 媒体未传 OSS，待后台补
            self._auto_map_safe(existing["id"])   # 采集后自动映射属性（已本地缓存化，快）
            self._maybe_auto_publish(existing["id"])   # 开了 auto_publish 则后台发布
            return {"created": [{"id": existing["id"], "source_title": existing.get("source_title")}], "errors": [], "deduped": True,
                    "auto_publish": bool((self.store.get_settings() or {}).get("auto_publish"))}
        try:
            saved = self.store.insert_draft(new_draft)
        except Exception as exc:  # noqa: BLE001  兜底:并发同 url 重采等极端撞唯一键 → 找回同 url 已存在的,不再 500
            msg = str(exc)
            ex = (self.store.find_by_source_url(new_draft["source_url"], None, store_cid)
                  if (("uq_draft" in msg) or ("Duplicate" in msg) or ("UNIQUE" in msg)) else None)
            if ex:
                return {"created": [{"id": ex["id"], "source_title": ex.get("source_title")}],
                        "errors": [], "deduped": True,
                        "auto_publish": bool((self.store.get_settings() or {}).get("auto_publish"))}
            raise
        if _has_media and self._media_needs_upload(saved):
            self.store.set_media_status(saved["id"], "pending")  # 媒体未传 OSS，待后台补
        self._auto_map_safe(saved["id"])   # 采集后自动映射属性（已本地缓存化，快）
        self._maybe_auto_publish(saved["id"])   # 开了 auto_publish 则后台发布
        return {"created": [{"id": saved["id"], "source_title": saved.get("source_title")}], "errors": [],
                "auto_publish": bool((self.store.get_settings() or {}).get("auto_publish"))}

    def update_draft_media(self, payload: dict) -> dict:
        """插件后台把媒体传完 OSS 后回调：把草稿里命中的原 URL 换成 OSS URL，置 media_status=done。"""
        draft_id = _to_int(payload.get("draft_id"))
        media_map = payload.get("media_map") or {}
        if not draft_id:
            raise ValueError("draft_id required")
        self.store.apply_media_oss(draft_id, dict(media_map))
        self._maybe_auto_publish(draft_id)   # 媒体传完 → 开了 auto_publish 则后台发布
        return {"ok": True}

    def pending_media_drafts(self) -> dict:
        """当前用户媒体待传(pending)的草稿，供插件后台补传。"""
        from webui.store import current_user_id  # noqa: PLC0415
        return {"drafts": self.store.list_pending_media_drafts(current_user_id.get())}

    def ext_add_snapshot(self, payload: dict) -> dict:
        pid = str(payload.get("product_id") or "").strip()
        if not pid:
            raise ValueError("product_id required")
        fc = payload.get("follow_count")
        pmin = payload.get("price_min")
        pmax = payload.get("price_max")
        last = self.store.latest_offer_snapshot(pid)
        if last and last.get("follow_count") == fc and last.get("price_min") == pmin and last.get("price_max") == pmax:
            return {"id": last.get("id"), "deduped": True}
        from webui.drafts import dumps_json, utc_now_iso  # noqa: PLC0415
        snap = {
            "product_id": pid,
            "sku": (str(payload.get("sku")) if payload.get("sku") else None),
            "captured_at": utc_now_iso(),
            "follow_count": fc,
            "price_min": pmin,
            "price_max": pmax,
            "sellers_json": dumps_json(payload.get("sellers") or []),
        }
        return self.store.add_offer_snapshot(snap)

    def ext_snapshots(self, product_id: str) -> dict:
        # price_min/max 列已是 Numeric→Decimal；API 出口 float 化保证 JSON number（前端/插件期望 number）
        snaps = [
            _money_to_float(s, ("price_min", "price_max"))
            for s in self.store.list_offer_snapshots(str(product_id))
        ]
        return {"product_id": str(product_id), "snapshots": snaps}

    def apply_ai_proposal(self, draft_id: int) -> dict:
        """把草稿的 AI 待确认草案合并进正式字段，清空草案。
        - fields 里存在的 key 才覆盖（删除的 key 保留正式原值）
        - attributes 有 value 的项解析成上架格式并按 id 合并；空 value 跳过
        返回 {ok, draft, unmapped}。无草案 → ValueError。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        proposal = draft.get("ai_proposal")
        if not proposal:
            raise ValueError("没有待确认的 AI 草案")
        fields = dict(proposal.get("fields") or {})
        patch: dict = {}
        for k in ("ozon_title", "description", "category_id", "type_id",
                  "weight_g", "length_mm", "width_mm", "height_mm"):
            if k in fields:
                patch[k] = fields[k]
        brand_id = fields.get("brand_id", draft.get("brand_id"))
        brand_name = str(fields.get("brand_name") or draft.get("brand_name") or "").strip()
        patch["brand_id"] = brand_id if brand_name == NO_BRAND and _to_int(brand_id) > 0 else None
        patch["brand_name"] = NO_BRAND
        cat = int(str(patch.get("category_id") or draft.get("category_id") or 0) or 0)
        typ = int(str(patch.get("type_id") or draft.get("type_id") or 0) or 0)
        new_attrs: list[dict] = []
        unmapped: list[dict] = []
        for a in (proposal.get("attributes") or []):
            val = str(a.get("value") or "").strip()
            if not val:
                continue
            aid = _to_int(a.get("id"))
            if aid <= 0:
                continue
            if aid == BRAND_ATTR_ID:
                continue
            if aid == 23171:   # #Хештеги 是自由文本，整串直存，不走字典解析(否则有类目时会被解析丢弃)
                new_attrs.append({"id": 23171, "values": [{"value": val}]})
                continue
            texts = [t.strip() for t in val.replace("，", ",").split(",") if t.strip()]
            vals = self._resolve_values(cat, typ, aid, texts, False) if cat and typ else [{"value": val}]
            if vals:
                new_attrs.append({"id": aid, "values": vals})
            else:
                unmapped.append({"id": aid, "name": a.get("name"), "value": val})
        attr_list = draft.get("attributes") if isinstance(draft.get("attributes"), list) else []
        passthrough = [a for a in attr_list if isinstance(a, dict) and not ("id" in a and "values" in a)]
        existing_pub = [a for a in attr_list if isinstance(a, dict) and "id" in a and "values" in a]
        merged = {_to_int(a["id"]): a for a in existing_pub if _to_int(a.get("id")) > 0}
        for a in new_attrs:
            merged[_to_int(a["id"])] = a
        patch["attributes"] = passthrough + list(merged.values())
        self.store.update_draft(draft_id, patch)
        self.store.set_ai_proposal(draft_id, None)
        updated = self.store.get_draft(draft_id)
        return {"ok": True, "draft": updated, "unmapped": unmapped}

    def patch_ai_proposal(self, draft_id: int, patch: dict) -> dict:
        """编辑/删除草案某项，或 discard 清空整份草案，写回 ai_proposal_json。
        op: edit_field{key,value} / delete_field{key} / edit_attr{id,value} /
            delete_attr{id} / discard。无草案 → ValueError；未知 op → ValueError。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        proposal = draft.get("ai_proposal")
        if not proposal:
            raise ValueError("没有待确认的 AI 草案")
        op = str((patch or {}).get("op") or "")
        if op == "discard":
            self.store.set_ai_proposal(draft_id, None)
            return {"ok": True, "proposal": None}
        fields = dict(proposal.get("fields") or {})
        attrs = list(proposal.get("attributes") or [])
        if op == "edit_field":
            fields[str(patch["key"])] = patch.get("value")
        elif op == "delete_field":
            fields.pop(str(patch.get("key")), None)
        elif op == "edit_attr":
            aid = _to_int(patch.get("id")); val = str(patch.get("value") or "")
            for a in attrs:
                if _to_int(a.get("id")) == aid:
                    a["value"] = val
                    break
        elif op == "delete_attr":
            aid = _to_int(patch.get("id"))
            attrs = [a for a in attrs if _to_int(a.get("id")) != aid]
        else:
            raise ValueError(f"未知操作: {op}")
        proposal["fields"] = fields
        proposal["attributes"] = attrs
        self.store.set_ai_proposal(draft_id, proposal)
        return {"ok": True, "proposal": proposal}

    # ===== realFBS 运费路线表（CSV 可维护；存 settings kv 的 JSON blob，首次读灌种子）=====
    _REALFBS_FIELDS = [
        "scoringGroup", "serviceLevel", "provider", "deliveryMethod", "ozonRating",
        "etaDays", "rateText", "batteries", "liquids", "measurements",
        "weightMinG", "weightMaxG", "valueRangeRub", "tarification",
        "volumeFormula", "compensationRub",
    ]
    _REALFBS_NUM = {"ozonRating", "weightMinG", "weightMaxG", "compensationRub"}

    def _realfbs_seed_routes(self) -> list[dict]:
        import json as _json  # noqa: PLC0415
        from pathlib import Path  # noqa: PLC0415
        seed = Path(__file__).resolve().parent / "realfbs_routes_seed.json"
        try:
            return _json.loads(seed.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            return []

    def realfbs_routes(self) -> dict:
        """返回 realFBS 运费路线给定价用。表空则灌种子(141 条)并持久化。"""
        routes = self.store.get_realfbs_routes()
        if routes is None:
            routes = self._realfbs_seed_routes()
            if routes:
                self.store.set_realfbs_routes(routes)
        return {"routes": routes or []}

    def import_realfbs_routes(self, csv_text: str) -> dict:
        """CSV 整表覆盖运费路线。表头须含 _REALFBS_FIELDS 各列；数值列空→None。"""
        import csv as _csv  # noqa: PLC0415
        import io as _io  # noqa: PLC0415

        def _num(v):
            s = str(v or "").strip()
            if s in ("", "None", "nan"):
                return None
            try:
                return float(s)
            except ValueError:
                return None

        reader = _csv.DictReader(_io.StringIO(csv_text or ""))
        routes: list[dict] = []
        for row in reader:
            r: dict = {}
            for f in self._REALFBS_FIELDS:
                raw = row.get(f)
                r[f] = _num(raw) if f in self._REALFBS_NUM else str(raw or "").strip()
            if not r.get("provider") and not r.get("deliveryMethod") and not r.get("rateText"):
                continue   # 跳过全空行
            routes.append(r)
        if not routes:
            raise ValueError("CSV 未解析到任何运费路线（请确认表头含 provider/rateText 等列）")
        self.store.set_realfbs_routes(routes)
        return {"count": len(routes)}

    def export_realfbs_routes_csv(self) -> str:
        """当前运费路线导出为 CSV（给用户下载→Excel 维护→再导入）。"""
        import csv as _csv  # noqa: PLC0415
        import io as _io  # noqa: PLC0415
        routes = self.realfbs_routes()["routes"]
        buf = _io.StringIO()
        w = _csv.DictWriter(buf, fieldnames=self._REALFBS_FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in routes:
            w.writerow({f: ("" if r.get(f) is None else r.get(f)) for f in self._REALFBS_FIELDS})
        return buf.getvalue()

    # ===== realFBS 佣金类目表（Excel 可维护；存 settings kv 的 JSON blob，首次读灌种子）=====
    #   只取 RFBS(=FBS) 三价格档佣金；导入认两种 Excel：Ozon 官方 Tarifs、本工具导出的模板。
    _COMMISSION_PRICE_TIERS_RUB = [1500, 5000]
    _COMMISSION_HEADER_ALIASES = {
        "parentZh": ("父类目(中)", "父类目", "parentzh"),
        "parentEn": ("父类目(英)", "parenten"),
        "subZh": ("子类目(中)", "子类目", "subzh"),
        "subEn": ("子类目(英)", "suben"),
        "r0": ("佣金% 0-1500", "0-1500", "rfbs0", "rfbs_0_1500"),
        "r1": ("佣金% 1500-5000", "1500-5000", "rfbs1"),
        "r2": ("佣金% 5000+", "5000+", "rfbs2"),
    }

    def _commission_seed(self) -> list[dict]:
        import json as _json  # noqa: PLC0415
        from pathlib import Path  # noqa: PLC0415
        seed = Path(__file__).resolve().parent / "commission_categories_seed.json"
        try:
            return _json.loads(seed.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            return []

    def commission_categories(self) -> dict:
        """返回佣金类目给定价用。表空则灌种子(80 类)并持久化。"""
        cats = self.store.get_commission_categories()
        if cats is None:
            cats = self._commission_seed()
            if cats:
                self.store.set_commission_categories(cats)
        return {"categories": cats or [], "priceTiersRub": list(self._COMMISSION_PRICE_TIERS_RUB)}

    @staticmethod
    def _commission_rate(v) -> float | None:
        """单元格 → 佣金小数：>1 视为百分比(/100)，否则视为小数；空/非数/≤0 → None。"""
        s = str(v).strip() if v is not None else ""
        if s in ("", "None", "nan"):
            return None
        try:
            x = float(s.replace("%", "").replace(",", "."))
        except ValueError:
            return None
        if x <= 0:
            return None
        return round(x / 100 if x > 1 else x, 4)

    def _parse_commission_ozon(self, wb) -> list[dict]:
        """认 Ozon 官方 Tarifs：sheet 'MP Tree Tarifs CN'，列 1=父EN 2=父ZH 4=子EN 5=子ZH 6-8=RFBS。"""
        name = next((s for s in wb.sheetnames if "MP Tree Tarifs" in s), None)
        if not name:
            return []
        ws = wb[name]
        out: list[dict] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2 or len(row) < 9:        # 跳过首行('Starts from')与表头行；列数不足跳过
                continue
            sub_en = str(row[4] or "").strip()
            rfbs = [self._commission_rate(row[6]), self._commission_rate(row[7]), self._commission_rate(row[8])]
            if not sub_en or any(r is None for r in rfbs):
                continue
            out.append({
                "parentEn": str(row[1] or "").strip(), "parentZh": str(row[2] or "").strip(),
                "subEn": sub_en, "subZh": str(row[5] or "").strip(), "rfbs": rfbs,
            })
        return out

    def _parse_commission_template(self, wb) -> list[dict]:
        """认本工具导出的模板：按表头名定位列（中英别名），佣金可填百分比或小数。"""
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        norm = [str(h or "").strip().lower() for h in header]

        def col(key) -> int | None:
            for alias in self._COMMISSION_HEADER_ALIASES[key]:
                a = alias.strip().lower()
                if a in norm:
                    return norm.index(a)
            return None

        idx = {k: col(k) for k in self._COMMISSION_HEADER_ALIASES}
        if idx["subEn"] is None or idx["r0"] is None:    # 表头对不上→不是我们的模板
            return []

        def at(row, key):
            j = idx[key]
            return row[j] if (j is not None and j < len(row)) else None

        out: list[dict] = []
        for row in rows:
            sub_en = str(at(row, "subEn") or "").strip()
            rfbs = [self._commission_rate(at(row, "r0")),
                    self._commission_rate(at(row, "r1")),
                    self._commission_rate(at(row, "r2"))]
            if not sub_en or any(r is None for r in rfbs):
                continue
            out.append({
                "parentEn": str(at(row, "parentEn") or "").strip(),
                "parentZh": str(at(row, "parentZh") or "").strip(),
                "subEn": sub_en, "subZh": str(at(row, "subZh") or "").strip(), "rfbs": rfbs,
            })
        return out

    def import_commission_categories_xlsx(self, data: bytes) -> dict:
        """整表覆盖佣金类目。先认 Ozon 官方 Tarifs，再认本工具模板；只取 RFBS(FBS)。"""
        import io as _io  # noqa: PLC0415

        import openpyxl  # noqa: PLC0415
        try:
            wb = openpyxl.load_workbook(_io.BytesIO(data), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"无法读取 Excel：{exc}")
        try:
            cats = self._parse_commission_ozon(wb) or self._parse_commission_template(wb)
        finally:
            wb.close()
        if not cats:
            raise ValueError("未从 Excel 解析到任何佣金类目（请用 Ozon 官方 Tarifs 文件，或本工具导出的模板）")
        # 按 (parentEn, subEn) 去重，保最后一条、保出现顺序
        seen: dict = {}
        order: list = []
        for c in cats:
            k = (c["parentEn"], c["subEn"])
            if k not in seen:
                order.append(k)
            seen[k] = c
        final = [seen[k] for k in order]
        self.store.set_commission_categories(final)
        return {"count": len(final)}

    @staticmethod
    def _commission_pct(x):
        return None if x is None else round(float(x) * 100, 2)

    def export_commission_categories_xlsx(self) -> bytes:
        """当前佣金类目导出为模板 xlsx（中文表头 + 百分比，给用户改完再导入）。"""
        import io as _io  # noqa: PLC0415

        import openpyxl  # noqa: PLC0415
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "佣金表(FBS)"
        ws.append(["父类目(中)", "父类目(英)", "子类目(中)", "子类目(英)",
                   "佣金% 0-1500", "佣金% 1500-5000", "佣金% 5000+"])
        for c in self.commission_categories()["categories"]:
            rfbs = (c.get("rfbs") or []) + [None, None, None]
            ws.append([c.get("parentZh"), c.get("parentEn"), c.get("subZh"), c.get("subEn"),
                       self._commission_pct(rfbs[0]), self._commission_pct(rfbs[1]), self._commission_pct(rfbs[2])])
        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def get_commission_map(self, cat_id: int, type_id: int) -> dict:
        """取某 Ozon 类目记住的 realFBS 佣金类目（命中返回 {parent_en, sub_en, rfbs}）。"""
        return self.store.load_commission_map(cat_id, type_id) or {}

    def save_commission_map(self, payload: dict) -> dict:
        cat = _to_int(payload.get("cat"))
        typ = _to_int(payload.get("type"))
        if not cat or not typ:
            return {"error": "cat/type 必填"}
        self.store.save_commission_map(
            cat, typ, str(payload.get("parent_en") or ""), str(payload.get("sub_en") or ""),
            payload.get("rfbs") if isinstance(payload.get("rfbs"), list) else [],
        )
        return {"ok": True}

    def _scid_of(self, store_client_id: str | None) -> str:
        """把(可能为空的)目标店解析成实际 client_id：空→默认店。仓库/订单按它隔离。"""
        return str(self._settings_for_store(store_client_id).get("ozon_client_id") or "")

    def _warehouses_with_delivery(self, scid: str) -> list[dict]:
        """仓库列表，每个仓库挂上它的配送方式（本地按 warehouse_id 分组）。"""
        warehouses = self.store.list_warehouses(scid)
        methods = self.store.list_delivery_methods(scid)
        by_wh: dict[int, list[dict]] = {}
        for m in methods:
            by_wh.setdefault(m.get("warehouse_id"), []).append(m)
        for w in warehouses:
            w["delivery_methods"] = by_wh.get(w.get("warehouse_id"), [])
        return warehouses

    def list_warehouses(self, store_client_id: str | None = None) -> dict:
        scid = self._scid_of(store_client_id)
        return {"warehouses": self._warehouses_with_delivery(scid)}

    def sync_warehouses(self, store_client_id: str | None = None) -> dict:
        from webui.ozon_client_adapter import (  # noqa: PLC0415
            fetch_delivery_methods,
            fetch_warehouses,
        )
        settings = self._settings_for_store(store_client_id)
        scid = str(settings.get("ozon_client_id") or "")
        items = fetch_warehouses(settings)
        self.store.upsert_warehouses(items, scid)
        # 配送方式挂在仓库下：仓库同步完后逐仓拉取，按店全量替换
        wids = [w.get("warehouse_id") for w in self.store.list_warehouses(scid)]
        methods = fetch_delivery_methods(settings, wids)
        self.store.replace_delivery_methods(methods, scid)
        return {
            "synced": len(items),
            "delivery_methods": len(methods),
            "warehouses": self._warehouses_with_delivery(scid),
        }

    def set_default_warehouse(self, warehouse_id: int, store_client_id: str | None = None) -> dict:
        scid = self._scid_of(store_client_id)
        self.store.set_default_warehouse(int(warehouse_id), scid)
        return {"warehouses": self._warehouses_with_delivery(scid)}

    # ---------- 功能⑤：FBS 备货发货 ----------
    def pull_fbs(self, status: str = "awaiting_packaging", days: int = 14,
                 store_client_id: str | None = None) -> dict:
        from webui.ozon_client_adapter import pull_fbs_postings  # noqa: PLC0415
        settings = self._settings_for_store(store_client_id)
        scid = str(settings.get("ozon_client_id") or "")
        items = pull_fbs_postings(settings, status, days)
        self.store.upsert_postings(items, scid)
        self.store.rebuild_procurement(scid)
        return {"synced": len(items), "procurement": self.store.list_procurement(scid)}

    def list_procurement(self, store_client_id: str | None = None) -> dict:
        scid = self._scid_of(store_client_id)
        return {"procurement": self.store.list_procurement(scid)}

    def set_procurement_state(self, pid: int, purchase_state: str, note: str = "",
                              store_client_id: str | None = None) -> dict:
        self.store.set_procurement_state(int(pid), purchase_state, note=note)
        return {"procurement": self.store.list_procurement(self._scid_of(store_client_id))}

    def ship_posting(self, posting_number: str, store_client_id: str | None = None) -> dict:
        """按 posting 的商品组成一个包发货。不可逆。
        Ozon 发货请求要 product_id（swagger 里与 sku 是不同 ID）；而待发货单只带
        offer_id/sku，没有 product_id。故先用 offer_id 反查真实 product_id（info.id），
        绝不拿 sku 顶替——否则会发错货或被 Ozon 拒。"""
        from webui.ozon_client_adapter import get_ozon_info, ship_fbs  # noqa: PLC0415
        posting = self.store.get_posting(posting_number)
        if not posting:
            raise KeyError(f"posting {posting_number} not found")
        items = posting.get("products") or []
        offer_ids = [str(p.get("offer_id")) for p in items if p.get("offer_id")]
        if not offer_ids:
            raise ValueError("该 posting 无 offer_id，无法解析 product_id")
        settings = self._settings_for_store(store_client_id)
        try:
            info = get_ozon_info(settings, offer_ids)   # {offer_id: {id=product_id, ...}}
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"拉取商品信息失败，无法发货: {exc}") from exc
        products, missing = [], []
        for p in items:
            oid = str(p.get("offer_id") or "")
            pid = (info.get(oid) or {}).get("id")
            if not pid:
                missing.append(oid or "?")
                continue
            products.append({"product_id": int(pid), "quantity": int(p.get("quantity") or 1)})
        if missing:
            raise ValueError(f"无法解析 product_id 的商品: {', '.join(missing)}")
        if not products:
            raise ValueError("该 posting 无可发货商品")
        # warehouse_id 不传给 Ozon：/v4/posting/fbs/ship 接口只接受 posting_number + packages，
        # Ozon 从 posting 本身推断仓库，无需（也不接受）warehouse_id 参数。
        r = ship_fbs(settings, posting_number, [{"products": products}])
        return {"result": r.get("result") or [], "shipped": True, "response": r}

    # ---------- 出图任务（gen_jobs）----------

    def submit_gen_job(self, draft_id: int, target: int) -> dict:
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        if self.store.has_active_gen_job(draft_id):
            raise ValueError("该草稿已有进行中的出图任务，请等待完成")
        job = self.store.create_gen_job(draft_id, target)
        try:
            from ozon_common.mq import publish_gen_job  # noqa: PLC0415
            publish_gen_job(job["id"], draft_id, target)
        except Exception as exc:
            self.store.update_gen_job(job["id"], {"status": "failed", "error": f"MQ 发送失败: {exc}"})
            raise RuntimeError(f"消息队列不可用: {exc}")
        return {"job_id": job["id"], "status": job["status"]}

    def submit_gen_images_custom(self, draft_id: int, payload: dict) -> dict:
        """用户自定义出图（AiImageDialog）：选 N 张源图 + 提示词 + 可选参考图 → 建 gen_job → MQ → worker。
        每张源图一个槽，源图为空(文生图)则建 1 个空槽。"""
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        source_urls = [str(u).strip() for u in (payload.get("source_urls") or []) if str(u).strip()]
        if not source_urls:
            source_urls = [""]   # text2img: 一个空槽
        ref_url = str(payload.get("ref_url") or "").strip() or None
        prompt = str(payload.get("prompt") or "").strip()
        size = str(payload.get("size") or "1024x1536").strip()
        as_main = bool(payload.get("as_main"))
        target = len(source_urls)
        job = self.store.create_gen_job(draft_id, target)
        # 构造 image_plan（worker 读它来获取 prompt 和 source_url）
        slots = []
        for i, surl in enumerate(source_urls):
            sid = f"custom_{i}"
            slots.append({
                "slot_id": sid, "label": f"自定义出图{i+1}",
                "action": "custom", "source_idx": 0,
                "prompt": prompt, "size": size, "as_main": (as_main and i == 0),
                "source_url": surl or None, "ref_url": ref_url,
            })
        self.store.create_gen_job_images(job["id"], slots)
        # 存 image_plan 到 source_raw 供 worker 读取
        sr = dict(draft.get("source_raw") or {})
        sr["image_plan"] = [(sr.get("image_plan") or []) + slots][0]
        self.store.update_draft(draft_id, {"source_raw": sr})
        try:
            from ozon_common.mq import publish_gen_job  # noqa: PLC0415
            publish_gen_job(job["id"], draft_id, target, mode="custom")
        except Exception as exc:
            self.store.update_gen_job(job["id"], {"status": "failed", "error": f"MQ 发送失败: {exc}"})
            raise RuntimeError(f"消息队列不可用: {exc}")
        return {"job_id": job["id"], "status": job["status"]}

    def submit_batch_gen_job(self, draft_id: int, source_indices: list[int], action: str) -> dict:
        """批量出图（白底/俄化/场景/细节/重做）：每个 source_index 一个槽，走 worker 异步。"""
        from ozon_common.gen_image import (  # noqa: PLC0415
            LOCALIZE_PROMPT,
            SCENE_PROMPT,
            WHITE_MAIN_PROMPT,
            build_infographic_prompt,
        )
        draft = self.store.get_draft(draft_id)
        if draft is None:
            raise KeyError(f"draft {draft_id} not found")
        if self.store.has_active_gen_job(draft_id):
            raise ValueError("该草稿已有进行中的出图任务")
        indices = [int(i) for i in source_indices if isinstance(i, int)]
        if not indices:
            raise ValueError("未选择图片")
        label_map = {"white": "白底图", "localize": "俄化", "scene": "场景图",
                     "detail": "细节图", "redo": "重做"}
        label = label_map.get(action, action)
        job = self.store.create_gen_job(draft_id, len(indices))
        slots = []
        for si in indices:
            p = ""
            if action == "white":
                p = WHITE_MAIN_PROMPT
            elif action == "localize":
                p = LOCALIZE_PROMPT
            elif action == "scene":
                p = SCENE_PROMPT
            elif action in ("detail", "redo"):
                p = build_infographic_prompt(role="细节" if action == "detail" else "")
            sid = f"batch_{action}_{si}"
            slots.append({"slot_id": sid, "label": f"{label}#{si}", "action": "batch",
                         "source_idx": si, "prompt": p})
        self.store.create_gen_job_images(job["id"], slots)
        sr = dict(draft.get("source_raw") or {})
        sr["image_plan"] = (sr.get("image_plan") or []) + slots
        self.store.update_draft(draft_id, {"source_raw": sr})
        try:
            from ozon_common.mq import publish_gen_job  # noqa: PLC0415
            publish_gen_job(job["id"], draft_id, len(indices), mode="batch")
        except Exception as exc:
            self.store.update_gen_job(job["id"], {"status": "failed", "error": f"MQ 发送失败: {exc}"})
            raise RuntimeError(f"消息队列不可用: {exc}")
        return {"job_id": job["id"], "status": job["status"]}

    def get_gen_job_status(self, job_id: int) -> dict:
        job = self.store.get_gen_job(job_id)
        if job is None:
            raise KeyError(f"job {job_id} not found")
        images = self.store.get_gen_job_images(job_id)
        return {"job_id": job["id"], "status": job["status"],
                "target": job["target"], "total": job["total"],
                "succeeded": job["succeeded"], "failed": job["failed"],
                "error": job.get("error"), "created_at": job["created_at"],
                "updated_at": job["updated_at"],
                "images": [{"slot_id": i["slot_id"], "label": i["label"],
                            "status": i["status"], "url": i.get("url"),
                            "error": i.get("error")} for i in images]}

    def get_latest_gen_job(self, draft_id: int) -> dict:
        job = self.store.get_latest_gen_job(draft_id)
        if job is None:
            raise KeyError(f"draft {draft_id} 没有出图任务")
        return self.get_gen_job_status(job["id"])

    def batch_latest_gen_jobs(self, draft_ids: list[int]) -> dict:
        """批量查多个草稿的全部出图任务（卡片列表展示用）。返回 {draft_id: [jobs]}。"""
        out = {}
        for did in draft_ids:
            try:
                jobs = self.store.list_gen_jobs(did)
            except Exception:
                continue
            if jobs:
                summaries = []
                for j in jobs:
                    imgs = self.store.get_gen_job_images(j["id"])
                    done = sum(1 for i in imgs if i["status"] == "done")
                    fail = sum(1 for i in imgs if i["status"] == "failed")
                    run = sum(1 for i in imgs if i["status"] == "running")
                    summaries.append({
                        "job_id": j["id"], "status": j["status"],
                        "target": j["target"],
                        "total": j["total"] or len(imgs),
                        "succeeded": done, "failed": fail, "running": run,
                        "created_at": j.get("created_at"),
                        "updated_at": j.get("updated_at"),
                    })
                out[str(did)] = summaries
        return {"jobs": out}

    def copy_images_to_draft(self, draft_id, image_urls, target_draft_ids):
        if not image_urls:
            raise ValueError("未选择图片")
        targets = [int(t) for t in (target_draft_ids or []) if int(t)]
        if not targets:
            raise ValueError("未指定目标变体")
        src = self.store.get_draft(draft_id)
        if src is None:
            raise KeyError(f"draft {draft_id} not found")
        vg_src = str((src.get("source_raw") or {}).get("variant_group") or "")
        for tid in targets:
            tgt = self.store.get_draft(tid)
            if tgt is None:
                raise KeyError(f"draft {tid} not found")
            vg_tgt = str((tgt.get("source_raw") or {}).get("variant_group") or "")
            if vg_src and vg_src != vg_tgt:
                raise ValueError("只能在同一变体组内复制图片")
        added = self.store.copy_images(draft_id, image_urls, targets)
        return {"ok": True, "added": added}

    def gallery_add(self, draft_id, image_ids):
        self.store.gallery_add(int(draft_id), [int(i) for i in (image_ids or [])])
        return self.store.get_draft(int(draft_id))

    def gallery_remove(self, draft_id, image_ids):
        self.store.gallery_remove(int(draft_id), [int(i) for i in (image_ids or [])])
        return self.store.get_draft(int(draft_id))

    def gallery_reorder(self, draft_id, image_ids):
        self.store.gallery_reorder(int(draft_id), [int(i) for i in (image_ids or [])])
        return self.store.get_draft(int(draft_id))

    def gallery_delete(self, draft_id, image_id):
        self.store.gallery_delete(int(draft_id), int(image_id))
        return self.store.get_draft(int(draft_id))
