from __future__ import annotations

import io
import tempfile
import unittest
from pathlib import Path

import openpyxl

from webui.app_service import App
from webui.store import Store


def _app(tmp: str) -> App:
    app = App()
    app.store.close()                          # 关掉默认库连接
    app.store = Store(Path(tmp) / "test.db")   # 隔离到临时库
    return app


def _ozon_tarifs_xlsx() -> bytes:
    """造一个 Ozon 官方 Tarifs 风格的最小 xlsx：sheet 'MP Tree Tarifs CN'，
    列布局与真实文件一致(RU,EN,ZH 父 | RU,EN,ZH 子 | RFBS×3 | FBP×3)。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MP Tree Tarifs CN"
    ws.append([None] * 12)                                          # 第1行：'Starts from ...'
    ws.append(["cat_ru", "ZH", "EN", "mp_ru", "ZH", "EN",
               "RFBS -> 0 - 1500", "RFBS -> 1500.01 - 5000", "RFBS -> 5000.01+",
               "FBP -> 0 - 1500", "FBP -> 1500.01 - 5000", "FBP -> 5000.01+"])   # 第2行表头
    # 数据行：列1=parentEn 列2=parentZh 列4=subEn 列5=subZh 列6-8=RFBS 列9-11=FBP
    ws.append(["药房_ru", "Pharmacy Products", "药房商品", "药店_ru", "Pharmacy", "药店",
               0.12, 0.14, 0.18, 0.99, 0.99, 0.99])
    ws.append(["家居_ru", "House & Car Products", "家居与汽车用品", "装饰_ru",
               "Decor, Cleaning & Storage", "装饰、清洁与储物",
               0.12, 0.14, 0.20, 0.88, 0.88, 0.88])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _template_xlsx() -> bytes:
    """我们导出的模板风格：中文表头 + 百分比数字（人手维护友好）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "佣金表(FBS)"
    ws.append(["父类目(中)", "父类目(英)", "子类目(中)", "子类目(英)",
               "佣金% 0-1500", "佣金% 1500-5000", "佣金% 5000+"])
    ws.append(["药房商品", "Pharmacy Products", "药店", "Pharmacy", 12, 14, 18])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class CommissionCategoriesTest(unittest.TestCase):
    def test_seed_on_first_read_and_persist(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            app = _app(tmp)
            try:
                res = app.commission_categories()
                cats = res["categories"]
                self.assertEqual(len(cats), 80)                    # 种子=80 个 MP 类目
                self.assertEqual(res["priceTiersRub"], [1500, 5000])
                self.assertEqual(len(cats[0]["rfbs"]), 3)          # 三价格档
                # 中文不是乱码：至少一个类目名含中日韩汉字
                self.assertTrue(any(
                    any("一" <= ch <= "鿿" for ch in (c.get("parentZh") or ""))
                    for c in cats
                ))
                # 已持久化：第二次读数量一致（来自 DB，不再灌种子）
                self.assertEqual(len(app.commission_categories()["categories"]), 80)
            finally:
                app.store.close()

    def test_store_get_set_roundtrip(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store = Store(Path(tmp) / "test.db")
            try:
                self.assertIsNone(store.get_commission_categories())     # 空 → None
                store.set_commission_categories([{"subEn": "A"}, {"subEn": "B"}])
                self.assertEqual([c["subEn"] for c in store.get_commission_categories()], ["A", "B"])
                store.set_commission_categories([{"subEn": "C"}])        # 覆盖
                self.assertEqual([c["subEn"] for c in store.get_commission_categories()], ["C"])
            finally:
                store.close()

    def test_import_ozon_tarifs_fbs_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            app = _app(tmp)
            try:
                r = app.import_commission_categories_xlsx(_ozon_tarifs_xlsx())
                self.assertEqual(r["count"], 2)                    # 整表覆盖
                cats = app.commission_categories()["categories"]
                self.assertEqual(len(cats), 2)
                ph = next(c for c in cats if c["subEn"] == "Pharmacy")
                self.assertEqual(ph["parentEn"], "Pharmacy Products")
                self.assertEqual(ph["parentZh"], "药房商品")
                self.assertEqual(ph["subZh"], "药店")
                self.assertEqual(ph["rfbs"], [0.12, 0.14, 0.18])   # 只取 RFBS，忽略 FBP(0.99)
            finally:
                app.store.close()

    def test_import_template_percent_to_decimal(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            app = _app(tmp)
            try:
                r = app.import_commission_categories_xlsx(_template_xlsx())
                self.assertEqual(r["count"], 1)
                c = app.commission_categories()["categories"][0]
                self.assertEqual(c["subEn"], "Pharmacy")
                self.assertEqual(c["parentZh"], "药房商品")
                self.assertEqual(c["rfbs"], [0.12, 0.14, 0.18])    # 12/14/18 → 小数
            finally:
                app.store.close()

    def test_export_roundtrips(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            app = _app(tmp)
            try:
                app.import_commission_categories_xlsx(_ozon_tarifs_xlsx())
                data = app.export_commission_categories_xlsx()
                self.assertEqual(data[:2], b"PK")                  # xlsx 是 zip 包
                app.import_commission_categories_xlsx(data)        # 导出 → 再导入 等价
                cats = app.commission_categories()["categories"]
                self.assertEqual(len(cats), 2)
                ph = next(c for c in cats if c["subEn"] == "Pharmacy")
                self.assertEqual(ph["rfbs"], [0.12, 0.14, 0.18])
                self.assertEqual(ph["parentZh"], "药房商品")
            finally:
                app.store.close()

    def test_import_empty_raises(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            app = _app(tmp)
            try:
                wb = openpyxl.Workbook()
                wb.active.append(["nonsense", "header"])
                buf = io.BytesIO()
                wb.save(buf)
                with self.assertRaises(ValueError):
                    app.import_commission_categories_xlsx(buf.getvalue())
            finally:
                app.store.close()


if __name__ == "__main__":
    unittest.main()
