"""强化筛选 v2：绿区 + 剔大件重货 + 偏小件 + 价不过低 + 缺口纳入排序。
仍是启发式（搜索词无克重，用品类词近似大小），输出需人眼复核 + 后续真克重验证。
"""
from __future__ import annotations

import json
import pathlib

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = pathlib.Path(__file__).resolve().parent
SRC = HERE / "data" / "search_queries_days7.jsonl"
OUT = HERE / "data" / "Ozon轻小绿区_候选.xlsx"

# 非绿区（命中即排除）
BLOCK = [
    "плать", "футболк", "шорт", "джинс", "носк", "трус", "юбк", "костюм", "куртк",
    "пальто", "свитер", "кофт", "рубашк", "брюк", "бель", "купальник", "кроссовк",
    "туфл", "босоножк", "сандал", "ботин", "сапог", "крокс", "обувь", "шапк", "шарф",
    "сумк", "рюкзак", "часы", "майк", "сарафан",
    "телефон", "смартфон", "наушник", "зарядк", "кабель", "адаптер", "павербанк",
    "колонк", "телевизор", "ноутбук", "планшет", "мышь", "клавиатур", "камер", "чехол",
    "лампа", "фен", "утюг", "блендер", "пылесос", "триммер", "шуруповерт", "дрель",
    "аккумулят", "батарейк", "220", "электр", "робот", "монитор", "смарт", "айфон",
    "шампунь", "крем", "маск", "помад", "духи", "парфюм", "тушь", "масло", "витамин",
    "таблетк", "лекарств", "чай", "кофе", "конфет", "шоколад", "корм", "бальзам", "паст",
    "детск", "игрушк", "кукл", "констру", "подгузник", "коляск", "медицин", "виниры",
    "отбелив", "зуб",
]

# 绿区白名单（必须命中）
GREEN = [
    "органайзер", "хранени", "коробк", "ящик", "контейнер", "корзин", "держатель",
    "подставк", "разделит", "крючок", "вешалк", "ваза", "рамк", "свеч", "подсвечник",
    "наклейк", "статуэтк", "кашпо", "панно", "пенал", "папк", "канцеляр", "скрепк",
    "степлер", "блокнот", "дозатор", "крышк", "брелок", "зажим", "клипс", "сушилк",
    "мешочек", "лоток", "когтеточк", "поводок", "миск", "намордник", "шлейк",
    "салфетк", "скатерт", "крюк",
]

# 大件/重货（命中即剔除——直发运费陷阱）
HEAVY = [
    "сетка", "сетк", "коврик", "ковёр", "ковер", "палас", "штор", "мебель", "стол",
    "стул", "шкаф", "кроват", "диван", "матрас", "стеллаж", "зеркало", "гамак",
    "качел", "тент", "навес", "зонт", "лестниц", "ванна", "раковин", "дверь",
    "забор", "плитк", "обои", "полка для ванной", "напольн", "гарнитур", "столешниц",
    "покрывал", "плед", "подушк",
]

WIN_SUMMER = ["лет", "пляж", "солнц", "дача", "сад", "пикник", "комар", "путешеств", "чемодан"]
WIN_SCHOOL = ["школ", "пенал", "тетрад", "канцеляр", "папк", "ручк", "карандаш"]


def _win(q):
    if any(s in q for s in WIN_SCHOOL):
        return "学校季"
    if any(s in q for s in WIN_SUMMER):
        return "夏季"
    return "四季"


def _rank(vals):
    order = sorted(range(len(vals)), key=lambda i: vals[i])
    n = len(vals)
    out = [0.0] * n
    for pos, i in enumerate(order):
        out[i] = pos / (n - 1) if n > 1 else 0.0
    return out


def main():
    rows = [json.loads(l) for l in SRC.open(encoding="utf-8")]
    kept = []
    for r in rows:
        q = (r.get("query") or "").lower()
        if not q:
            continue
        if any(b in q for b in BLOCK):
            continue
        if not any(g in q for g in GREEN):
            continue
        if any(h in q for h in HEAVY):
            continue
        if (r.get("uniqSellers", 0) or 0) <= 0:
            continue
        price = r.get("avgCaRub", 0) or 0
        if price < 350 or price > 4000:   # 太便宜盖不住固定成本 / 太贵多半大件
            continue
        kept.append(r)
    print(f"[i] 10000 → 轻小绿区(价350~4000) {len(kept)} 行")

    dps = [r.get("count", 0) / max(r.get("uniqSellers", 0) or 1, 1) for r in kept]
    gap = [r.get("zrShare", 0) or 0 for r in kept]
    dem = [r.get("count", 0) or 0 for r in kept]
    rd, rg, re_ = _rank(dps), _rank(gap), _rank(dem)
    recs = []
    for i, r in enumerate(kept):
        opp = 100 * (0.35 * rd[i] + 0.35 * rg[i] + 0.30 * re_[i])  # 缺口权重提高
        recs.append({
            "机会分": round(opp, 1), "窗口": _win((r.get("query") or "").lower()),
            "搜索词": r.get("query", ""), "搜索量": r.get("count", 0),
            "卖家数": r.get("uniqSellers", 0), "需求/卖家": round(dps[i], 1),
            "缺口率%": round(r.get("zrShare", 0) or 0, 1),
            "加购均价₽": round(r.get("avgCaRub", 0) or 0),
            "年趋势%": round(r.get("trendForYear", 0) or 0, 1),
        })
    recs.sort(key=lambda x: x["机会分"], reverse=True)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "轻小绿区候选"
    cols = list(recs[0].keys()); ws.append(cols)
    for r in recs:
        ws.append([r[c] for c in cols])
    for c in range(1, len(cols) + 1):
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"
    for i, w in enumerate([8, 7, 34, 9, 8, 10, 8, 11, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(OUT)
    print(f"[+] 导出 {len(recs)} 行 -> {OUT}\n")
    print(f"=== 轻小绿区 Top 30（缺口权重提高）===")
    print(f"{'机会':>5}{'窗口':<5}{'搜索词':<32}{'量':>6}{'卖家':>5}{'缺口%':>6}{'均价₽':>7}")
    for r in recs[:30]:
        print(f"{r['机会分']:>5} {r['窗口']:<5}{r['搜索词'][:32]:<32}{r['搜索量']:>6}{r['卖家数']:>5}{r['缺口率%']:>6}{r['加购均价₽']:>7}")


if __name__ == "__main__":
    main()
