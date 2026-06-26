"""把 search_queries_days7.jsonl 转成带派生指标的 Excel，按机会分排序。

机会分逻辑（方法论：高需求 + 低竞争 + 高缺口）：
  需求/卖家比 = count / max(uniqSellers,1)   越高=越供不应求
  缺口率      = zrShare（无结果搜索占比）     越高=平台没货
  opp 分      = 0.4·rank(需求/卖家比) + 0.3·rank(缺口率) + 0.3·rank(需求)
"""
from __future__ import annotations

import json
import pathlib

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = pathlib.Path(__file__).resolve().parent
SRC = HERE / "data" / "search_queries_days7.jsonl"
OUT = HERE / "data" / "Ozon搜索查询_机会分析_days7.xlsx"


def _pct_rank(values: list[float]) -> dict[int, float]:
    order = sorted(range(len(values)), key=lambda i: values[i])
    n = len(values)
    rank = {}
    for pos, idx in enumerate(order):
        rank[idx] = pos / (n - 1) if n > 1 else 0.0
    return rank


def main() -> None:
    rows = [json.loads(l) for l in SRC.open(encoding="utf-8")]
    dps = [r.get("count", 0) / max(r.get("uniqSellers", 0) or 1, 1) for r in rows]
    gap = [r.get("zrShare", 0) or 0 for r in rows]
    dem = [r.get("count", 0) or 0 for r in rows]
    r_dps, r_gap, r_dem = _pct_rank(dps), _pct_rank(gap), _pct_rank(dem)

    recs = []
    for i, r in enumerate(rows):
        opp = 100 * (0.4 * r_dps[i] + 0.3 * r_gap[i] + 0.3 * r_dem[i])
        recs.append({
            "机会分": round(opp, 1),
            "搜索词": r.get("query", ""),
            "搜索量": r.get("count", 0),
            "卖家数": r.get("uniqSellers", 0),
            "需求/卖家比": round(dps[i], 1),
            "缺口率%": round(r.get("zrShare", 0) or 0, 1),
            "加购均价₽": round(r.get("avgCaRub", 0) or 0),
            "加购转化%": round(r.get("ca", 0) or 0, 2),
            "年趋势%": round(r.get("trendForYear", 0) or 0, 1),
            "GMV₽": round(r.get("gmv", 0) or 0),
            "订单数": r.get("ord", 0),
        })
    recs.sort(key=lambda x: x["机会分"], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全部(按机会分)"
    cols = list(recs[0].keys())
    ws.append(cols)
    for r in recs:
        ws.append([r[c] for c in cols])

    # 表头样式 + 冻结 + 列宽
    fill = PatternFill("solid", fgColor="1F4E78")
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    ws.freeze_panes = "A2"
    widths = [8, 34, 10, 8, 12, 9, 12, 10, 9, 14, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    print(f"[+] 已导出 {len(recs)} 行 -> {OUT}")
    print("\n机会分 Top 20（高需求+低竞争+高缺口）：")
    print(f"{'机会':>5} {'搜索词':<28} {'搜索量':>7} {'卖家':>5} {'需求/卖家':>8} {'缺口%':>6} {'均价₽':>7}")
    for r in recs[:20]:
        print(f"{r['机会分']:>5} {r['搜索词'][:28]:<28} {r['搜索量']:>7} {r['卖家数']:>5} "
              f"{r['需求/卖家比']:>8} {r['缺口率%']:>6} {r['加购均价₽']:>7}")


if __name__ == "__main__":
    main()
